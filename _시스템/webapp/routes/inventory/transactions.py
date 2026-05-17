"""[I] /inventory/{inbound,outbound,adjust,move,history} — 4 거래 + 히스토리.

ai-workflow STEP 7 Sprint 2 Task 2.1~2.5
2026-05-08 보강: 멀티라인 거래 + 엑셀 export + pending 임시저장 (박스히어로 1:1 복제).
v1.5: Bundle 출고 자동 차감 — 묶음 SKU 입력 시 구성 SKU 로 자동 expand.
"""
import io
import json
from pathlib import Path
from datetime import datetime, timezone
from flask import render_template, request, redirect, url_for, flash, send_file, jsonify, abort

from shared.db import SessionLocal
from lemouton.sourcing.models import Option, Model
from lemouton.inventory.models import InventoryTx, InventoryLocation, InventoryPending
from lemouton.inventory import inbound as tx_svc

from . import bp


# ============ Bundle 자동 차감 ============
_BUNDLE_FILE = Path(__file__).resolve().parents[3] / 'data' / 'bundles.json'


def _load_bundles_map() -> dict:
    """Bundle SKU → list[(component_sku, component_qty)] 맵."""
    if not _BUNDLE_FILE.exists():
        return {}
    try:
        items = json.loads(_BUNDLE_FILE.read_text(encoding='utf-8'))
    except Exception:
        return {}
    out = {}
    for b in items:
        sku = b.get('sku')
        comps = [(c.get('sku'), int(c.get('qty', 1))) for c in b.get('components', []) if c.get('sku')]
        if sku and comps:
            out[sku] = {'name': b.get('name', sku), 'components': comps}
    return out


def _expand_bundle_lines(lines: list[dict], bundles: dict) -> tuple[list[dict], list[str]]:
    """묶음 SKU 라인을 구성 SKU 라인으로 확장.

    lines: [{sku, qty, price?}]
    return: (expanded_lines, expansion_notes) — 호출자가 메모에 추가 가능.
    """
    out = []
    notes = []
    for line in lines:
        sku = line.get('sku')
        qty = line.get('qty', 0)
        if not sku or qty <= 0:
            continue
        b = bundles.get(sku)
        if b:
            for comp_sku, comp_qty in b['components']:
                out.append({
                    'sku': comp_sku,
                    'qty': qty * comp_qty,
                    'price': 0,
                    'bundle_origin': sku,
                })
            notes.append(f"묶음 {sku} ({b['name']}) × {qty} → 구성 {len(b['components'])}종 자동 차감")
        else:
            out.append(line)
    return out, notes


# ============ 임시 저장 (Pending) — 박스히어로 1:1 ============

@bp.post('/pending/save')
def pending_save():
    """거래 작성 중간 저장 — tx_type · 폼 데이터 직렬화."""
    tx_type = request.form.get('tx_type', '').strip()
    if tx_type not in ('in', 'out', 'adjust', 'move'):
        flash('잘못된 tx_type', 'error')
        return redirect(url_for('inventory.history'))
    payload = {k: request.form.getlist(k) if len(request.form.getlist(k)) > 1 else request.form.get(k, '')
               for k in request.form.keys() if k != 'tx_type'}
    s = SessionLocal()
    try:
        p = InventoryPending(
            tx_type=tx_type,
            payload_json=json.dumps(payload, ensure_ascii=False),
            created_by=request.form.get('created_by', '운영자'),
        )
        s.add(p)
        s.commit()
        flash(f'임시 저장 완료 (#{p.id})', 'success')
    finally:
        s.close()
    target = {'in': 'inventory.inbound_list', 'out': 'inventory.outbound_list',
              'adjust': 'inventory.adjust_list', 'move': 'inventory.move_list'}[tx_type]
    return redirect(url_for(target))


@bp.get('/pending')
def pending_list():
    """임시 저장 목록."""
    s = SessionLocal()
    try:
        items = (s.query(InventoryPending)
                 .order_by(InventoryPending.created_at.desc()).limit(100).all())
        rows = []
        for p in items:
            try:
                payload = json.loads(p.payload_json) if p.payload_json else {}
            except (ValueError, TypeError):
                payload = {}
            rows.append({'p': p, 'payload': payload,
                         'sku_count': len([s for s in payload.get('option_canonical_sku', []) if s] if isinstance(payload.get('option_canonical_sku'), list) else [payload.get('option_canonical_sku')] if payload.get('option_canonical_sku') else [])})
        return render_template('inventory/pending_list.html',
                               active='pending', rows=rows)
    finally:
        s.close()


@bp.post('/pending/<int:pid>/delete')
def pending_delete(pid):
    s = SessionLocal()
    try:
        p = s.query(InventoryPending).filter(InventoryPending.id == pid).first()
        if p:
            s.delete(p)
            s.commit()
            flash(f'임시저장 #{pid} 삭제', 'success')
    finally:
        s.close()
    return redirect(url_for('inventory.pending_list'))


# ============ 입고 (Task 2.1) ============

@bp.get('/inbound')
def inbound_list():
    s = SessionLocal()
    try:
        page = max(1, int(request.args.get('page', 1)))
        per_page = 30
        status = request.args.get('status', 'all')  # all | draft | completed
        f = _list_filter(s, 'in', page, per_page, status)
        # 카운트 (탭별)
        all_total = s.query(InventoryTx).filter(InventoryTx.tx_type == 'in').count()
        draft_total = s.query(InventoryPending).filter(InventoryPending.tx_type == 'in').count()
        return render_template(
            'inventory/inbound/list.html',
            active='inbound', items=f['items'], total=f['total'], page=page,
            total_pages=max(1, (f['total'] + per_page - 1) // per_page),
            loc_map=_loc_map(s), per_page=per_page, status=status, mode=f['mode'],
            counts={'all': all_total, 'draft': draft_total, 'completed': all_total},
        )
    finally:
        s.close()


@bp.get('/inbound/<int:tx_id>')
def inbound_detail(tx_id):
    """입고 상세 — 박스히어로 1:1 행 클릭 시 detail 페이지."""
    s = SessionLocal()
    try:
        tx = s.query(InventoryTx).filter(InventoryTx.id == tx_id, InventoryTx.tx_type == 'in').first()
        if not tx:
            flash('입고 거래 없음', 'error')
            return redirect(url_for('inventory.inbound_list'))
        # 옵션 정보
        opt = s.query(Option).filter(Option.canonical_sku == tx.option_canonical_sku).first()
        loc = s.query(InventoryLocation).filter(InventoryLocation.id == tx.location_id).first()
        return render_template(
            'inventory/inbound/detail.html',
            active='inbound', tx=tx, opt=opt, loc=loc,
        )
    finally:
        s.close()


@bp.get('/inbound/new')
def inbound_new():
    s = SessionLocal()
    try:
        from lemouton.inventory.locations import list_active
        locations = list_active(s)
        # 제품 추가 모달용 — 모든 옵션 (model_name + canonical_sku + 색상/사이즈)
        options = (s.query(Option).order_by(Option.model_code, Option.canonical_sku).limit(500).all())
        opt_data = [{
            'sku': o.canonical_sku, 'model': o.model_code,
            'color': o.color_display or o.color_code,
            'size': o.size_display or o.size_code,
            'bh': o.boxhero_sku or '', 'stock': o.boxhero_stock_total or 0,
            'avg': o.boxhero_avg_purchase_price or 0,
        } for o in options]
        return render_template(
            'inventory/inbound/form.html', active='inbound',
            locations=locations, opt_data=opt_data,
        )
    finally:
        s.close()


@bp.post('/inbound/create')
def inbound_create():
    """다중 라인 입고 — 박스히어로 1:1 N라인 거래."""
    s = SessionLocal()
    try:
        skus = request.form.getlist('option_canonical_sku')
        qtys = request.form.getlist('qty')
        prices = request.form.getlist('unit_purchase_price')
        # 단일 라인 호환 (멀티 라인 없을 때)
        if not skus and request.form.get('option_canonical_sku'):
            skus = [request.form['option_canonical_sku']]
            qtys = [request.form.get('qty', '0')]
            prices = [request.form.get('unit_purchase_price', '0')]
        loc_id = int(request.form['location_id'])
        partner = request.form.get('partner_label', '')
        memo = request.form.get('memo', '')
        created_by = request.form.get('created_by', '운영자')
        ok = err = 0
        for i, sku in enumerate(skus):
            sku = (sku or '').strip()
            if not sku:
                continue
            try:
                qty = int((qtys[i] if i < len(qtys) else '0') or 0)
                if qty <= 0:
                    continue
                price = int((prices[i] if i < len(prices) else '0') or 0)
                tx_svc.create_inbound(s, location_id=loc_id, option_canonical_sku=sku,
                                      qty=qty, unit_purchase_price=price,
                                      partner_label=partner, memo=memo, created_by=created_by)
                ok += 1
            except (ValueError, KeyError) as e:
                err += 1
        if ok:
            s.commit()
            flash(f"입고 완료 — {ok} 라인" + (f" / 실패 {err}" if err else ""), 'success')
        else:
            s.rollback()
            flash("입고 실패: 유효한 라인 없음", 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.inbound_list'))


# ============ 출고 (Task 2.2) ============

def _opt_data_all(s, include_bundles: bool = False):
    """제품 추가 모달용 — 전체 옵션 JSON. include_bundles=True 면 묶음 SKU 포함 (출고용)."""
    options = (s.query(Option).order_by(Option.model_code, Option.canonical_sku).limit(500).all())
    out = [{
        'sku': o.canonical_sku, 'model': o.model_code,
        'color': o.color_display or o.color_code,
        'size': o.size_display or o.size_code,
        'bh': o.boxhero_sku or '', 'stock': o.boxhero_stock_total or 0,
        'avg': o.boxhero_avg_purchase_price or 0,
        'img': o.image_url or '',
        'is_bundle': False,
    } for o in options]
    if include_bundles:
        bundles = _load_bundles_map()
        for sku, b in bundles.items():
            comps_str = ' + '.join(f"{c[0]}×{c[1]}" for c in b['components'])
            out.append({
                'sku': sku, 'model': '',
                'color': b['name'], 'size': '묶음',
                'bh': '', 'stock': 0, 'avg': 0, 'img': '',
                'is_bundle': True, 'components': comps_str,
            })
    return out


def _list_filter(s, tx_type, page, per_page, status):
    """거래 list 박스히어로 5세부탭 필터 (전체/임시저장/완료)."""
    if status == 'draft':
        # 임시 저장 (InventoryPending)
        q = s.query(InventoryPending).filter(InventoryPending.tx_type == tx_type)
        total = q.count()
        items = q.order_by(InventoryPending.created_at.desc()).offset((page-1)*per_page).limit(per_page).all()
        return {'mode': 'pending', 'items': items, 'total': total}
    # 완료 거래
    items, total = tx_svc.list_txs(s, tx_type, page, per_page)
    return {'mode': 'tx', 'items': items, 'total': total}


# ============ 거래 수정 (박스히어로 1:1) ============
def _update_tx(tx_id, tx_type, redirect_endpoint):
    """공통 거래 메타 수정 — 거래일·거래처·메모·담당자."""
    s = SessionLocal()
    try:
        tx = s.query(InventoryTx).filter(InventoryTx.id == tx_id, InventoryTx.tx_type == tx_type).first()
        if not tx:
            flash('거래 없음', 'error')
            return redirect(url_for(redirect_endpoint))
        # 거래일
        dt = (request.form.get('created_at') or '').strip()
        if dt:
            try:
                tx.created_at = datetime.fromisoformat(dt.replace('T', ' ')[:19])
            except ValueError:
                flash('잘못된 거래일 형식', 'error')
                return redirect(url_for(redirect_endpoint.replace('_list', '_detail'), tx_id=tx_id))
        tx.partner_label = request.form.get('partner_label', tx.partner_label)
        tx.memo = request.form.get('memo', tx.memo)
        tx.created_by = request.form.get('created_by', tx.created_by)
        s.commit()
        flash(f'거래 #{tx_id} 수정됨', 'success')
        return redirect(url_for(redirect_endpoint.replace('_list', '_detail'), tx_id=tx_id))
    finally:
        s.close()


@bp.post('/inbound/<int:tx_id>/update')
def inbound_update(tx_id):
    return _update_tx(tx_id, 'in', 'inventory.inbound_list')


@bp.post('/outbound/<int:tx_id>/update')
def outbound_update(tx_id):
    return _update_tx(tx_id, 'out', 'inventory.outbound_list')


@bp.post('/adjust/<int:tx_id>/update')
def adjust_update(tx_id):
    return _update_tx(tx_id, 'adjust', 'inventory.adjust_list')


@bp.post('/move/<int:tx_id>/update')
def move_update(tx_id):
    return _update_tx(tx_id, 'move', 'inventory.move_list')


def _loc_map(s):
    return {l.id: l.name for l in s.query(InventoryLocation).all()}


@bp.get('/outbound')
def outbound_list():
    s = SessionLocal()
    try:
        page = max(1, int(request.args.get('page', 1)))
        per_page = 30
        status = request.args.get('status', 'all')
        f = _list_filter(s, 'out', page, per_page, status)
        all_total = s.query(InventoryTx).filter(InventoryTx.tx_type == 'out').count()
        draft_total = s.query(InventoryPending).filter(InventoryPending.tx_type == 'out').count()
        return render_template(
            'inventory/outbound/list.html',
            active='outbound', items=f['items'], total=f['total'], page=page,
            total_pages=max(1, (f['total'] + per_page - 1) // per_page),
            loc_map=_loc_map(s), per_page=per_page, status=status, mode=f['mode'],
            counts={'all': all_total, 'draft': draft_total, 'completed': all_total},
        )
    finally:
        s.close()


@bp.get('/outbound/<int:tx_id>')
def outbound_detail(tx_id):
    s = SessionLocal()
    try:
        tx = s.query(InventoryTx).filter(InventoryTx.id == tx_id, InventoryTx.tx_type == 'out').first()
        if not tx:
            flash('출고 거래 없음', 'error')
            return redirect(url_for('inventory.outbound_list'))
        opt = s.query(Option).filter(Option.canonical_sku == tx.option_canonical_sku).first()
        loc = s.query(InventoryLocation).filter(InventoryLocation.id == tx.location_id).first()
        return render_template('inventory/outbound/detail.html',
                               active='outbound', tx=tx, opt=opt, loc=loc)
    finally:
        s.close()


@bp.get('/outbound/new')
def outbound_new():
    s = SessionLocal()
    try:
        from lemouton.inventory.locations import list_active
        locations = list_active(s)
        return render_template('inventory/outbound/form.html',
                               active='outbound', locations=locations,
                               opt_data=_opt_data_all(s, include_bundles=True))
    finally:
        s.close()


@bp.post('/outbound/create')
def outbound_create():
    """다중 라인 출고 — 묶음 SKU 자동 차감 지원."""
    s = SessionLocal()
    try:
        skus = request.form.getlist('option_canonical_sku')
        qtys = request.form.getlist('qty')
        prices = request.form.getlist('unit_sale_price')
        if not skus and request.form.get('option_canonical_sku'):
            skus = [request.form['option_canonical_sku']]
            qtys = [request.form.get('qty', '0')]
            prices = [request.form.get('unit_sale_price', '0')]
        loc_id = int(request.form['location_id'])
        partner = request.form.get('partner_label', '')
        memo = request.form.get('memo', '')
        created_by = request.form.get('created_by', '운영자')

        raw_lines = []
        for i, sku in enumerate(skus):
            sku = (sku or '').strip()
            if not sku:
                continue
            try:
                qty = int((qtys[i] if i < len(qtys) else '0') or 0)
                if qty <= 0:
                    continue
                price = int((prices[i] if i < len(prices) else '0') or 0)
                raw_lines.append({'sku': sku, 'qty': qty, 'price': price})
            except (ValueError, KeyError):
                pass

        bundles = _load_bundles_map()
        expanded, expand_notes = _expand_bundle_lines(raw_lines, bundles)
        full_memo = memo
        if expand_notes:
            full_memo = (memo + '\n' if memo else '') + '\n'.join(expand_notes)

        ok = err = bundle_expanded = 0
        for line in expanded:
            try:
                tx_svc.create_outbound(s, location_id=loc_id,
                                       option_canonical_sku=line['sku'],
                                       qty=line['qty'],
                                       unit_sale_price=line.get('price', 0),
                                       partner_label=partner, memo=full_memo,
                                       created_by=created_by)
                ok += 1
                if line.get('bundle_origin'):
                    bundle_expanded += 1
            except (ValueError, KeyError):
                err += 1

        if ok:
            s.commit()
            msg = f"출고 완료 — {ok} 라인"
            if bundle_expanded:
                msg += f" (묶음 자동 차감 {bundle_expanded} 라인 포함)"
            if err:
                msg += f" / 실패 {err}"
            flash(msg, 'success')
        else:
            s.rollback()
            flash("출고 실패: 유효한 라인 없음", 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.outbound_list'))


# ============ 조정 (Task 2.3) ============

@bp.get('/adjust')
def adjust_list():
    s = SessionLocal()
    try:
        page = max(1, int(request.args.get('page', 1)))
        per_page = 30
        status = request.args.get('status', 'all')
        f = _list_filter(s, 'adjust', page, per_page, status)
        all_total = s.query(InventoryTx).filter(InventoryTx.tx_type == 'adjust').count()
        draft_total = s.query(InventoryPending).filter(InventoryPending.tx_type == 'adjust').count()
        return render_template(
            'inventory/adjust/list.html',
            active='adjust', items=f['items'], total=f['total'], page=page,
            total_pages=max(1, (f['total'] + per_page - 1) // per_page),
            loc_map=_loc_map(s), per_page=per_page, status=status, mode=f['mode'],
            counts={'all': all_total, 'draft': draft_total, 'completed': all_total},
        )
    finally:
        s.close()


@bp.get('/adjust/<int:tx_id>')
def adjust_detail(tx_id):
    s = SessionLocal()
    try:
        tx = s.query(InventoryTx).filter(InventoryTx.id == tx_id, InventoryTx.tx_type == 'adjust').first()
        if not tx:
            flash('조정 거래 없음', 'error')
            return redirect(url_for('inventory.adjust_list'))
        opt = s.query(Option).filter(Option.canonical_sku == tx.option_canonical_sku).first()
        loc = s.query(InventoryLocation).filter(InventoryLocation.id == tx.location_id).first()
        return render_template('inventory/adjust/detail.html',
                               active='adjust', tx=tx, opt=opt, loc=loc)
    finally:
        s.close()


@bp.get('/adjust/new')
def adjust_new():
    s = SessionLocal()
    try:
        from lemouton.inventory.locations import list_active
        locations = list_active(s)
        return render_template('inventory/adjust/form.html',
                               active='adjust', locations=locations,
                               opt_data=_opt_data_all(s))
    finally:
        s.close()


@bp.post('/adjust/create')
def adjust_create():
    """다중 라인 조정 — ± 모드 지원 (박스히어로 1:1).

    delta_mode = 'set' (절대값) | 'plus' (습득 +) | 'minus' (분실 -)
    """
    s = SessionLocal()
    try:
        skus = request.form.getlist('option_canonical_sku')
        qtys = request.form.getlist('qty')
        modes = request.form.getlist('delta_mode')  # 라인별 ±/= 모드
        if not skus and request.form.get('option_canonical_sku'):
            skus = [request.form['option_canonical_sku']]
            qtys = [request.form.get('qty', '0')]
            modes = [request.form.get('delta_mode', 'set')]
        loc_id = int(request.form['location_id'])
        memo = request.form.get('memo', '')
        created_by = request.form.get('created_by', '운영자')
        ok = err = 0
        for i, sku in enumerate(skus):
            sku = (sku or '').strip()
            if not sku:
                continue
            try:
                input_qty = int((qtys[i] if i < len(qtys) else '0') or 0)
                mode = (modes[i] if i < len(modes) else 'set').strip() or 'set'
                # 현재 stock 조회
                opt = s.query(Option).filter(Option.canonical_sku == sku).first()
                if not opt:
                    err += 1
                    continue
                cur_stock = opt.boxhero_stock_total or 0
                if mode == 'plus':
                    new_qty = cur_stock + input_qty
                    line_memo = f'+{input_qty} 습득 ({memo})' if memo else f'+{input_qty} 습득'
                elif mode == 'minus':
                    new_qty = max(0, cur_stock - input_qty)
                    line_memo = f'-{input_qty} 분실 ({memo})' if memo else f'-{input_qty} 분실'
                else:  # set
                    new_qty = input_qty
                    line_memo = memo
                tx_svc.create_adjustment(s, location_id=loc_id,
                                         option_canonical_sku=sku, new_qty=new_qty,
                                         memo=line_memo, created_by=created_by)
                ok += 1
            except (ValueError, KeyError):
                err += 1
        if ok:
            s.commit()
            flash(f"조정 완료 — {ok} 라인" + (f" / 실패 {err}" if err else ""), 'success')
        else:
            s.rollback()
            flash("조정 실패: 유효한 라인 없음", 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.adjust_list'))


# ============ 이동 (Task 2.4) ============

@bp.get('/move')
def move_list():
    s = SessionLocal()
    try:
        page = max(1, int(request.args.get('page', 1)))
        per_page = 30
        status = request.args.get('status', 'all')
        f = _list_filter(s, 'move', page, per_page, status)
        all_total = s.query(InventoryTx).filter(InventoryTx.tx_type == 'move').count()
        draft_total = s.query(InventoryPending).filter(InventoryPending.tx_type == 'move').count()
        return render_template(
            'inventory/move/list.html',
            active='move', items=f['items'], total=f['total'], page=page,
            total_pages=max(1, (f['total'] + per_page - 1) // per_page),
            loc_map=_loc_map(s), per_page=per_page, status=status, mode=f['mode'],
            counts={'all': all_total, 'draft': draft_total, 'completed': all_total},
        )
    finally:
        s.close()


@bp.get('/move/<int:tx_id>')
def move_detail(tx_id):
    s = SessionLocal()
    try:
        tx = s.query(InventoryTx).filter(InventoryTx.id == tx_id, InventoryTx.tx_type == 'move').first()
        if not tx:
            flash('이동 거래 없음', 'error')
            return redirect(url_for('inventory.move_list'))
        opt = s.query(Option).filter(Option.canonical_sku == tx.option_canonical_sku).first()
        loc_from = s.query(InventoryLocation).filter(InventoryLocation.id == tx.location_id).first()
        loc_to = s.query(InventoryLocation).filter(InventoryLocation.id == tx.location_to_id).first() if tx.location_to_id else None
        return render_template('inventory/move/detail.html',
                               active='move', tx=tx, opt=opt, loc_from=loc_from, loc_to=loc_to)
    finally:
        s.close()


@bp.get('/move/new')
def move_new():
    s = SessionLocal()
    try:
        from lemouton.inventory.locations import list_active
        locations = list_active(s)
        return render_template('inventory/move/form.html',
                               active='move', locations=locations,
                               opt_data=_opt_data_all(s))
    finally:
        s.close()


@bp.post('/move/create')
def move_create():
    """다중 라인 이동."""
    s = SessionLocal()
    try:
        skus = request.form.getlist('option_canonical_sku')
        qtys = request.form.getlist('qty')
        if not skus and request.form.get('option_canonical_sku'):
            skus = [request.form['option_canonical_sku']]
            qtys = [request.form.get('qty', '0')]
        from_id = int(request.form['from_location_id'])
        to_id = int(request.form['to_location_id'])
        memo = request.form.get('memo', '')
        created_by = request.form.get('created_by', '운영자')
        ok = err = 0
        for i, sku in enumerate(skus):
            sku = (sku or '').strip()
            if not sku:
                continue
            try:
                qty = int((qtys[i] if i < len(qtys) else '0') or 0)
                if qty <= 0:
                    continue
                tx_svc.create_move(s, from_location_id=from_id, to_location_id=to_id,
                                   option_canonical_sku=sku, qty=qty,
                                   memo=memo, created_by=created_by)
                ok += 1
            except (ValueError, KeyError):
                err += 1
        if ok:
            s.commit()
            flash(f"이동 완료 — {ok} 라인" + (f" / 실패 {err}" if err else ""), 'success')
        else:
            s.rollback()
            flash("이동 실패: 유효한 라인 없음", 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.move_list'))


# ============ 히스토리 (Task 2.5) ============

@bp.get('/history')
def history():
    """박스히어로식 카드형 히스토리 — 거래서 묶음 + From/To + 색상 + 아바타 + 기간 필터."""
    s = SessionLocal()
    try:
        from sqlalchemy import or_
        from shared.search import split_tokens, apply_and_filter
        from lemouton.inventory.models import InventoryLocation
        from datetime import datetime as _dt, timedelta as _td
        page = max(1, int(request.args.get('page', 1)))
        tx_type = request.args.get('type', '')
        q = (request.args.get('q') or '').strip()
        search_tokens = split_tokens(q)
        # ★ 기간 필터 (박스히어로 1:1 — 전체/오늘/이번주/이번달/직접)
        period = request.args.get('period', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        query = s.query(InventoryTx).filter(InventoryTx.status == 'completed')
        if tx_type and tx_type in ('in', 'out', 'adjust', 'move'):
            query = query.filter(InventoryTx.tx_type == tx_type)
        # 기간 적용
        now = _dt.utcnow()
        if period == 'today':
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            query = query.filter(InventoryTx.created_at >= start)
        elif period == 'week':
            start = now - _td(days=7)
            query = query.filter(InventoryTx.created_at >= start)
        elif period == 'month':
            start = now - _td(days=30)
            query = query.filter(InventoryTx.created_at >= start)
        elif period == 'custom' and (date_from or date_to):
            try:
                if date_from:
                    df = _dt.strptime(date_from, '%Y-%m-%d')
                    query = query.filter(InventoryTx.created_at >= df)
                if date_to:
                    dt2 = _dt.strptime(date_to, '%Y-%m-%d') + _td(days=1)
                    query = query.filter(InventoryTx.created_at < dt2)
            except ValueError:
                pass
        # ★ 박스히어로식 다중 키워드 AND 교집합
        query = apply_and_filter(
            query, search_tokens,
            InventoryTx.option_canonical_sku, InventoryTx.partner_label, InventoryTx.memo,
        )
        total = query.count()
        items = query.order_by(InventoryTx.created_at.desc()).offset((page-1)*50).limit(50).all()

        # 위치 dict (id → name) — From/To 표시용
        locs = {loc.id: loc.name for loc in s.query(InventoryLocation).all()}

        # ★ 거래서 그룹화 — 같은 (분 단위 시각, 사용자, 종류, 위치) → 1 묶음
        # 박스히어로의 "거래서 = 여러 품목 묶음" UX 모사
        from collections import OrderedDict
        groups = OrderedDict()
        for tx in items:
            tx_minute = tx.created_at.replace(second=0, microsecond=0) if tx.created_at else None
            key = (tx_minute, tx.created_by or '', tx.tx_type, tx.location_id, tx.location_to_id)
            if key not in groups:
                groups[key] = {
                    'created_at': tx.created_at,
                    'created_by': tx.created_by or '시스템',
                    'tx_type': tx.tx_type,
                    'location_id': tx.location_id,
                    'location_to_id': tx.location_to_id,
                    'location_name': locs.get(tx.location_id, '?'),
                    'location_to_name': locs.get(tx.location_to_id, None) if tx.location_to_id else None,
                    'partner_label': tx.partner_label,
                    'memo': tx.memo,
                    'items': [],
                    'qty_total': 0,
                    'tx_ids': [],
                }
            qty_signed = tx.qty if tx.tx_type in ('in', 'adjust') else -tx.qty if tx.tx_type == 'out' else tx.qty
            groups[key]['items'].append({'sku': tx.option_canonical_sku, 'qty': qty_signed, 'tx_id': tx.id})
            groups[key]['qty_total'] += qty_signed
            groups[key]['tx_ids'].append(tx.id)
        grouped_list = list(groups.values())

        return render_template(
            'inventory/ledger.html',
            active='history', items=items, total=total, page=page,
            tx_type=tx_type, q=q, search_tokens=search_tokens,
            grouped=grouped_list, locs=locs,
            period=period, date_from=date_from, date_to=date_to,
        )
    finally:
        s.close()


# ============ 엑셀 내보내기 (박스히어로 1:1) ============

@bp.get('/history/export.xlsx')
def history_export():
    """히스토리 엑셀 내보내기."""
    import openpyxl
    s = SessionLocal()
    try:
        items = s.query(InventoryTx).order_by(InventoryTx.created_at.desc()).limit(5000).all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '히스토리'
        ws.append(['거래일', '유형', 'SKU', '수량', '위치', '거래처/고객',
                   '단가매입', '단가판매', 'COGS_snapshot', '메모', '작성자'])
        for tx in items:
            ws.append([
                tx.created_at.strftime('%Y-%m-%d %H:%M') if tx.created_at else '',
                tx.tx_type, tx.option_canonical_sku, tx.qty,
                tx.location_id, tx.partner_label or '',
                tx.unit_purchase_price_at_tx or 0,
                tx.unit_sale_price or 0,
                tx.unit_purchase_price_at_tx or 0,
                tx.memo or '', tx.created_by or '',
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'history_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname)
    finally:
        s.close()


@bp.get('/reports/sales/export.xlsx')
def sales_export():
    """매출 분석 엑셀 (COGS snapshot 기반)."""
    import openpyxl
    s = SessionLocal()
    try:
        items = (s.query(InventoryTx).filter(InventoryTx.tx_type == 'out')
                 .order_by(InventoryTx.created_at.desc()).limit(5000).all())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '매출분석'
        ws.append(['출고일', 'SKU', '수량', '판매단가', '매출', 'COGS_snapshot',
                   '매입원가', '마진', '마진율'])
        for tx in items:
            rev = (tx.unit_sale_price or 0) * (tx.qty or 0)
            cost = (tx.unit_purchase_price_at_tx or 0) * (tx.qty or 0)
            margin = rev - cost
            rate = (margin / rev * 100) if rev else 0
            ws.append([
                tx.created_at.strftime('%Y-%m-%d') if tx.created_at else '',
                tx.option_canonical_sku, tx.qty, tx.unit_sale_price or 0,
                rev, tx.unit_purchase_price_at_tx or 0, cost, margin,
                f'{rate:.1f}%',
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'sales_{datetime.now().strftime("%Y%m%d")}.xlsx')
    finally:
        s.close()


@bp.get('/reports/inventory/export.xlsx')
def inventory_export():
    """재고 분석 엑셀 (옵션 매트릭스 형태)."""
    import openpyxl
    s = SessionLocal()
    try:
        options = s.query(Option).order_by(Option.model_code).all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '재고분석'
        ws.append(['SKU', '박스히어로SKU', '모델', '색상', '사이즈',
                   '재고', '평균매입가', '자체마진오버라이드', '외부마진오버라이드'])
        for o in options:
            ws.append([
                o.canonical_sku, o.boxhero_sku or '', o.model_code,
                o.color_display or o.color_code, o.size_display or o.size_code,
                o.boxhero_stock_total or 0, o.boxhero_avg_purchase_price or 0,
                f'{o.option_boxhero_margin_mode}={o.option_boxhero_margin_value}' if o.option_boxhero_margin_mode else '',
                f'{o.option_external_margin_mode}={o.option_external_margin_value}' if o.option_external_margin_mode else '',
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'inventory_{datetime.now().strftime("%Y%m%d")}.xlsx')
    finally:
        s.close()


# ============ 거래 수정/삭제 (박스히어로 1:1) ============

@bp.route('/history/<int:tx_id>/edit', methods=['GET', 'POST'])
def tx_edit(tx_id: int):
    """거래 수정 (입고서·출고서·조정서·이동서 통합).

    GET: 폼 표시 (현재 값 prefill)
    POST: 변경 적용 — 재고 자동 재계산 (사용자 확인 체크박스 필수)
    """
    s = SessionLocal()
    try:
        from lemouton.inventory.models import InventoryLocation
        from flask import abort
        tx = s.query(InventoryTx).filter_by(id=tx_id).first()
        if not tx or tx.status != 'completed':
            abort(404)

        locs = s.query(InventoryLocation).filter(InventoryLocation.deleted_at.is_(None))\
                .order_by(InventoryLocation.name).all()
        error = None

        if request.method == 'POST':
            confirm_recalc = (request.form.get('confirm_recalc') == 'on')
            if not confirm_recalc:
                error = '재고 자동 계산 확인 체크가 필요합니다.'
            else:
                try:
                    new_loc = int(request.form.get('location_id') or tx.location_id)
                    new_qty = int(request.form.get('qty') or tx.qty)
                    new_partner = (request.form.get('partner_label') or '').strip() or None
                    new_memo = (request.form.get('memo') or '').strip() or None
                    new_loc_to = request.form.get('location_to_id')
                    new_loc_to = int(new_loc_to) if new_loc_to else None

                    tx.location_id = new_loc
                    tx.qty = new_qty
                    tx.partner_label = new_partner
                    tx.memo = new_memo
                    if tx.tx_type == 'move':
                        tx.location_to_id = new_loc_to
                    s.commit()
                    return redirect(url_for('inventory.history'))
                except Exception as e:
                    error = f'수정 실패: {e}'

        return render_template('inventory/tx_edit.html',
                               active='history', tx=tx, locs=locs, error=error)
    finally:
        s.close()


@bp.post('/history/<int:tx_id>/delete')
def tx_delete(tx_id: int):
    """거래 soft delete — status='deleted' 로 변경, 재고 자동 재계산."""
    s = SessionLocal()
    try:
        from flask import abort
        tx = s.query(InventoryTx).filter_by(id=tx_id).first()
        if not tx:
            abort(404)
        tx.status = 'deleted'
        s.commit()
        return jsonify(ok=True)
    finally:
        s.close()


# ============ Phase 4: SKU 별 이력 API (모달용) ============

@bp.get('/api/sku/<path:sku>/history')
def api_sku_history(sku: str):
    """SKU + (선택)위치 의 최근 거래 list. JSON.

    Query: ?location_id=N (선택, 없으면 전체 위치)
    Response: {items: [{tx_id, tx_type, qty, date, author, partner, memo, loc_name, loc_to_name}]}
    """
    from flask import jsonify
    from lemouton.inventory.models import InventoryLocation
    loc_id = request.args.get('location_id')
    s = SessionLocal()
    try:
        q = s.query(InventoryTx).filter(
            InventoryTx.option_canonical_sku == sku,
            InventoryTx.status == 'completed',
        )
        if loc_id:
            try:
                lid = int(loc_id)
                q = q.filter(
                    (InventoryTx.location_id == lid) | (InventoryTx.location_to_id == lid)
                )
            except ValueError:
                pass
        txs = q.order_by(InventoryTx.created_at.desc()).limit(50).all()
        locs = {loc.id: loc.name for loc in s.query(InventoryLocation).all()}
        items = []
        for tx in txs:
            qty_signed = tx.qty if tx.tx_type in ('in', 'adjust') else (-tx.qty if tx.tx_type == 'out' else tx.qty)
            items.append({
                'tx_id': tx.id,
                'tx_type': tx.tx_type,
                'qty': qty_signed,
                'date': tx.created_at.strftime('%Y-%m-%d %H:%M') if tx.created_at else '',
                'author': tx.created_by or '시스템',
                'partner': tx.partner_label or '',
                'memo': tx.memo or '',
                'loc_name': locs.get(tx.location_id, '?'),
                'loc_to_name': locs.get(tx.location_to_id, None) if tx.location_to_id else None,
            })
        return jsonify(ok=True, items=items, total=len(items))
    finally:
        s.close()
