"""[I] /inventory/data/* — 데이터 마스터 6 페이지 라우트.

ai-workflow STEP 7 Sprint 1A Task 1.2~1.6 + 묶음제품 (v1.4)
"""
import os
import json
import shutil
from datetime import datetime
from pathlib import Path
from flask import render_template, request, redirect, url_for, flash, jsonify, send_from_directory, current_app
from sqlalchemy import text as sa_text

from shared.db import SessionLocal
from lemouton.inventory import locations as locations_svc
from lemouton.inventory import attributes as attrs_svc
from lemouton.inventory import partner as partner_svc
from lemouton.sourcing.models import Option, Model

from . import bp


# ============ 묶음제품 (Bundle) — JSON 파일 기반 (DB schema 변경 회피) ============
BUNDLE_FILE = Path(__file__).resolve().parents[3] / 'data' / 'bundles.json'


def _load_bundles() -> list:
    if BUNDLE_FILE.exists():
        try:
            return json.loads(BUNDLE_FILE.read_text(encoding='utf-8'))
        except Exception:
            return []
    return []


def _save_bundles(bundles: list) -> None:
    BUNDLE_FILE.parent.mkdir(parents=True, exist_ok=True)
    BUNDLE_FILE.write_text(json.dumps(bundles, ensure_ascii=False, indent=2), encoding='utf-8')


# ============ 제품 이미지 업로드 (박스히어로 1:1) ============
UPLOAD_DIR = Path(__file__).resolve().parents[3] / 'data' / 'product_images'

@bp.post('/data/items/<path:sku>/upload-image')
def data_item_upload_image(sku):
    """옵션별 이미지 업로드. 파일은 data/product_images/<sku>.<ext>로 저장."""
    file = request.files.get('image')
    if not file or not file.filename:
        flash('이미지 파일을 선택하세요', 'error')
        return redirect(url_for('inventory.data_items'))
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('jpg', 'jpeg', 'png', 'webp', 'gif'):
        flash('이미지 형식 jpg/png/webp/gif 만 가능', 'error')
        return redirect(url_for('inventory.data_items'))
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    safe = sku.replace('/', '_').replace(' ', '_')
    fname = f'{safe}.{ext}'
    file.save(str(UPLOAD_DIR / fname))
    s = SessionLocal()
    try:
        opt = s.query(Option).filter(Option.canonical_sku == sku).first()
        if opt:
            opt.image_url = f'/inventory/data/product-image/{fname}'
            s.commit()
            flash(f'이미지 업로드 완료 — {sku}', 'success')
    finally:
        s.close()
    return redirect(url_for('inventory.data_items'))


@bp.get('/data/product-image/<path:filename>')
def data_product_image(filename):
    """업로드된 이미지 서빙."""
    return send_from_directory(str(UPLOAD_DIR), filename)


# ============ [Phase 3-3 · A2] 다중 선택 → 모음전 일괄 등록 ============

@bp.post('/data/items/bulk-bundle-register')
def data_items_bulk_bundle_register():
    """선택된 옵션들을 모음전으로 일괄 등록.

    [2026-05-28] Phase 3-3 (A2 시안):
      - 같은 모델·색상에 사이즈만 다른 N개 옵션 → 모음전 매트릭스로 묶기
      - 옵션의 model_code 변경 + Model upsert + BundleOptionStep 자동 추론

    body (JSON):
      {
        "skus": ["SKU-XXX", ...],            # 선택된 옵션 SKU 들
        "bundle_name": "메이트",              # 신규 모음전 이름
        "bundle_brand": "르무통",
        "bundle_category": "스니커즈",
        "bundle_article_no": "FV5420-002",   # 선택
      }
    Returns:
      {ok: True, new_model_code, options_moved, steps_inferred}
    """
    from flask import jsonify
    from lemouton.sourcing.master import upsert_model
    from lemouton.sourcing.option_service import save_step_design
    from lemouton.inventory.boxhero_import import _derive_model_code
    from shared.sku_format import clean_article_no
    import json as _json

    payload = request.get_json(silent=True) or {}
    skus = payload.get('skus') or []
    bundle_name = (payload.get('bundle_name') or '').strip()
    bundle_brand = (payload.get('bundle_brand') or '').strip() or '미상'
    bundle_category = (payload.get('bundle_category') or '').strip()
    bundle_article_no = clean_article_no(payload.get('bundle_article_no'))

    if not isinstance(skus, list) or not skus:
        return jsonify({'ok': False, 'error': 'SKU 목록이 비어있어요.'}), 400
    if not bundle_name:
        return jsonify({'ok': False, 'error': '모음전 이름은 필수예요.'}), 400

    s = SessionLocal()
    try:
        # 1. 선택된 옵션 조회
        opts = s.query(Option).filter(Option.canonical_sku.in_(skus)).all()
        if not opts:
            return jsonify({'ok': False, 'error': '옵션을 찾을 수 없어요.'}), 404

        # 2. 자동 단계 추론 — 색상 unique + 사이즈 unique
        colors = sorted({(o.color_display or o.color_code or '').strip() for o in opts if o.color_display or o.color_code})
        sizes = sorted({(o.size_display or o.size_code or '').strip() for o in opts if o.size_display or o.size_code})
        steps = []
        if colors:
            steps.append({'axis_name': '색상', 'values': list(colors)})
        if sizes:
            steps.append({'axis_name': '사이즈', 'values': list(sizes)})

        # 3. 새 model_code 생성 + Model upsert
        # 임시 SKU 로 model_code 도출 (옵션의 첫 SKU 기준)
        ref_sku = opts[0].canonical_sku
        new_model_code = _derive_model_code(bundle_brand,
                                             bundle_article_no if bundle_article_no != '-' else bundle_name,
                                             ref_sku)
        upsert_model(
            s,
            model_code=new_model_code,
            model_name_raw=bundle_name[:255],
            brand=bundle_brand[:100],
        )
        m_obj = s.query(Model).filter_by(model_code=new_model_code).first()
        if m_obj:
            if not (m_obj.model_name_display or '').strip():
                m_obj.model_name_display = bundle_name[:255]
            if bundle_brand and not (m_obj.brand or '').strip():
                m_obj.brand = bundle_brand[:100]
            if bundle_category and not (getattr(m_obj, 'category', None) or '').strip():
                m_obj.category = bundle_category[:100]
            if bundle_article_no != '-' and not (getattr(m_obj, 'article_no', None) or '').strip():
                m_obj.article_no = bundle_article_no

        # 4. 옵션의 model_code 를 새 모음전으로 변경
        for opt in opts:
            opt.model_code = new_model_code

        # 5. 단계 설계 저장
        save_step_design(s, new_model_code, steps)

        s.commit()
        # 모음전 상세 페이지 URL (있으면)
        try:
            bundle_url = url_for('bundles.bundle_edit', code=new_model_code)
        except Exception:
            bundle_url = None
        return jsonify({
            'ok': True,
            'new_model_code': new_model_code,
            'created': len(opts),         # JS 와 호환 (options_moved 와 동일)
            'options_moved': len(opts),
            'steps_inferred': steps,
            'bundle_url': bundle_url,
        })
    except Exception as e:
        s.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        s.close()


@bp.post('/data/items/backfill-article-no')
def data_items_backfill_article_no():
    """박스히어로 model_name → Model.article_no 일괄 백필.

    JSON body = {SKU: article_no, ...}.
    각 SKU 의 Model.article_no 가 비어있으면 채움 (기존 값 보존).
    `Option.canonical_sku` 또는 `Option.boxhero_sku` 둘 다 매칭 시도.
    """
    from flask import jsonify
    payload = request.get_json(silent=True) or {}
    mapping = payload.get('mapping') if isinstance(payload, dict) else None
    if not isinstance(mapping, dict) or not mapping:
        return jsonify({'ok': False, 'error': 'mapping body required'}), 400

    s = SessionLocal()
    try:
        # boxhero_sku ↔ canonical_sku 양쪽 다 매칭
        from sqlalchemy import or_
        skus = list(mapping.keys())
        opts = s.query(Option).filter(or_(
            Option.canonical_sku.in_(skus),
            Option.boxhero_sku.in_(skus),
        )).all()
        # sku → article 매핑 (canonical 과 boxhero 둘 다 인덱싱)
        # 동일 SKU 가 둘 다인 경우도 처리
        model_codes_seen: set[str] = set()
        updated = 0
        skipped_existing = 0
        skipped_no_model = 0
        for opt in opts:
            article = (mapping.get(opt.canonical_sku) or mapping.get(opt.boxhero_sku) or '').strip()
            if not article:
                continue
            if not opt.model_code or opt.model_code in model_codes_seen:
                continue
            m = s.query(Model).filter(Model.model_code == opt.model_code).first()
            if not m:
                skipped_no_model += 1
                continue
            existing = (getattr(m, 'article_no', None) or '').strip()
            if existing:
                skipped_existing += 1
                model_codes_seen.add(opt.model_code)
                continue
            m.article_no = article[:64]
            updated += 1
            model_codes_seen.add(opt.model_code)
        s.commit()
        return jsonify({
            'ok': True,
            'updated': updated,
            'skipped_existing': skipped_existing,
            'skipped_no_model': skipped_no_model,
            'mapping_size': len(mapping),
            'opts_matched': len(opts),
        })
    except Exception as e:
        s.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        s.close()


@bp.post('/data/items/auto-clean-colors')
def data_items_auto_clean_colors():
    """LCP 알고리즘으로 모든 옵션의 color_display + Model.model_name_display 일괄 정리.

    같은 model_code 그룹의 color_display 들의 LCP = 모델명.
    각 옵션의 color_display 에서 그 prefix strip + Model.model_name_display 에 LCP 저장.
    """
    from collections import defaultdict
    s = SessionLocal()
    try:
        options = s.query(Option).all()
        color_by_model: dict[str, list[str]] = defaultdict(list)
        for o in options:
            raw_c = (o.color_display or o.color_code or '').strip()
            if raw_c and o.model_code:
                color_by_model[o.model_code].append(raw_c)

        def _lcp(strs):
            if len(strs) < 2:
                return ''
            ss = sorted(strs)
            first, last = ss[0], ss[-1]
            i = 0
            while i < len(first) and i < len(last) and first[i] == last[i]:
                i += 1
            cp = first[:i]
            while cp and not cp[-1].isspace():
                cp = cp[:-1]
            return cp.strip()

        cleaned_count = 0
        model_updated = 0
        opts_by_model: dict[str, list] = defaultdict(list)
        for o in options:
            if o.model_code:
                opts_by_model[o.model_code].append(o)

        for mc, colors in color_by_model.items():
            cp = _lcp(colors)
            if not (cp and len(cp) >= 2):
                continue
            m = s.query(Model).filter_by(model_code=mc).first()
            if m and not (m.model_name_display or '').strip():
                m.model_name_display = cp[:255]
                model_updated += 1
            for o in opts_by_model.get(mc, []):
                raw = (o.color_display or '').strip()
                if raw.startswith(cp):
                    new = raw[len(cp):].strip()
                    if new and new != raw:
                        o.color_display = new[:64]
                        cleaned_count += 1

        s.commit()
        flash(f'✅ 색상 자동 정리 완료 — 옵션 {cleaned_count}개, 모델 {model_updated}개', 'success')
    except Exception as e:
        s.rollback()
        flash(f'❌ 자동 정리 실패: {e}', 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.home'))


@bp.get('/data/items/<path:sku>/info')
def data_item_info(sku):
    """[1부-2] 박스히어로 스타일 모달용 — 옵션 + 모델 통합 정보 JSON."""
    from flask import jsonify
    from lemouton.sourcing.models import Model as ModelM
    s = SessionLocal()
    try:
        opt = s.query(Option).filter(Option.canonical_sku == sku).first()
        if not opt:
            return jsonify({'ok': False, 'error': 'not found'}), 404
        model = s.query(ModelM).filter(ModelM.model_code == opt.model_code).first() if opt.model_code else None
        from shared.inventory_stock import get_stock_by_sku
        stock = get_stock_by_sku(s, sku)
        return jsonify({
            'ok': True,
            'sku': opt.canonical_sku,
            'barcode': opt.barcode or '',
            'boxhero_sku': opt.boxhero_sku or '',
            'color_display': opt.color_display or '',
            'size_display': opt.size_display or '',
            'avg_purchase_price': opt.boxhero_avg_purchase_price or 0,
            'image_url': getattr(opt, 'image_url', '') or '',
            'stock': stock,
            'model_code': opt.model_code or '',
            'model_name': (model.model_name_display or model.model_name_raw) if model else '',
            'brand': (model.brand if model else '') or '',
            'category': (model.category if model else '') or '',
            'article_no': (model.article_no if model else '') or '',
        })
    finally:
        s.close()


@bp.post('/data/items/<path:sku>/update')
def data_item_update(sku):
    """박스히어로 1:1 인라인 popover 편집 — 제품명·박스히어로 SKU·컬러·사이즈·이미지 동시 저장."""
    from lemouton.sourcing.models import Model as ModelM
    s = SessionLocal()
    try:
        opt = s.query(Option).filter(Option.canonical_sku == sku).first()
        if not opt:
            flash(f'옵션을 찾을 수 없습니다: {sku}', 'error')
            return redirect(url_for('inventory.data_items'))

        new_color = (request.form.get('color_display') or '').strip()
        new_size = (request.form.get('size_display') or '').strip()
        new_bh = (request.form.get('boxhero_sku') or '').strip()
        if hasattr(opt, 'color_display'):
            opt.color_display = new_color or opt.color_display
        if hasattr(opt, 'size_display'):
            opt.size_display = new_size or opt.size_display
        if new_bh:
            opt.boxhero_sku = new_bh

        # [2026-05-29 1부-2] 평균매입가 수정 + JSON 응답 옵션
        if 'avg_purchase_price' in request.form:
            from shared.sku_format import clean_avg_price
            opt.boxhero_avg_purchase_price = clean_avg_price(request.form.get('avg_purchase_price'))

        if opt.model_code:
            model = s.query(ModelM).filter(ModelM.model_code == opt.model_code).first()
            if model:
                new_name = (request.form.get('model_name') or '').strip()
                new_brand = (request.form.get('brand') or '').strip()
                new_article = (request.form.get('article_no') or '').strip()
                new_category = request.form.get('category')
                if new_name:
                    model.model_name_display = new_name
                if new_brand:
                    model.brand = new_brand
                # 품번은 빈 문자열도 허용 (= 명시적 삭제 의도)
                if 'article_no' in request.form:
                    model.article_no = new_article[:64] if new_article else None
                # [1부-2] 카테고리 — clean_category fallback 적용
                if new_category is not None:
                    from shared.sku_format import clean_category
                    model.category = clean_category(new_category)

        file = request.files.get('image')
        if file and file.filename:
            ext = file.filename.rsplit('.', 1)[-1].lower()
            if ext in ('jpg', 'jpeg', 'png', 'webp', 'gif'):
                UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
                safe = sku.replace('/', '_').replace(' ', '_')
                fname = f'{safe}.{ext}'
                file.save(str(UPLOAD_DIR / fname))
                if hasattr(opt, 'image_url'):
                    opt.image_url = f'/inventory/data/product-image/{fname}'

        s.commit()
        # [1부-2] AJAX 요청은 JSON 응답 (모달에서 페이지 이동 없이 처리)
        if request.headers.get('X-Requested-With') == 'fetch' or request.headers.get('Accept') == 'application/json':
            from flask import jsonify
            return jsonify({'ok': True, 'sku': sku, 'message': f'{sku} 저장됨'})
        flash(f'{sku} 저장됨', 'success')
    except Exception as e:
        s.rollback()
        if request.headers.get('X-Requested-With') == 'fetch':
            from flask import jsonify
            return jsonify({'ok': False, 'error': str(e)}), 500
        flash(f'저장 실패: {e}', 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.data_items'))


# ============ 위치 (Q1) ============

@bp.get('/data/locations')
def data_locations():
    s = SessionLocal()
    try:
        locs = locations_svc.list_active(s)
        return render_template(
            'inventory/data/locations.html',
            active='data-locations',
            locations=locs,
        )
    finally:
        s.close()


@bp.post('/data/locations/create')
def data_locations_create():
    name = request.form.get('name', '').strip()
    is_default = request.form.get('is_default') == 'on'
    s = SessionLocal()
    try:
        try:
            loc = locations_svc.create(s, name, is_default)
            s.commit()
            flash(f"위치 '{loc.name}' 추가됨", 'success')
        except ValueError as e:
            s.rollback()
            flash(str(e), 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.data_locations'))


@bp.post('/data/locations/<int:loc_id>/update')
def data_locations_update(loc_id):
    s = SessionLocal()
    try:
        try:
            locations_svc.update(
                s, loc_id,
                name=request.form.get('name'),
                sort_order=int(request.form['sort_order']) if request.form.get('sort_order') else None,
                is_default=(request.form.get('is_default') == 'on') if 'is_default' in request.form else None,
            )
            s.commit()
            flash("수정됨", 'success')
        except ValueError as e:
            s.rollback()
            flash(str(e), 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.data_locations'))


@bp.post('/data/locations/<int:loc_id>/delete')
def data_locations_delete(loc_id):
    s = SessionLocal()
    try:
        try:
            locations_svc.delete(s, loc_id)
            s.commit()
            flash("삭제됨", 'success')
        except ValueError as e:
            s.rollback()
            flash(str(e), 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.data_locations'))


@bp.post('/data/locations/seed-defaults')
def data_locations_seed():
    """박스히어로 기본 3 위치 시드 (그로스/기본 위치/판매불가)."""
    s = SessionLocal()
    try:
        created = locations_svc.seed_defaults(s)
        s.commit()
        flash(f"기본 위치 {len(created)}개 추가됨", 'success')
    finally:
        s.close()
    return redirect(url_for('inventory.data_locations'))


# ============ 제품 (Task 1.3) — 모음전 Option = 박스히어로 SKU 1:1 ============

@bp.get('/data/items')
def data_items():
    """제품 마스터 — 옵션 표 형식.

    형식 분기 — `?format=json` 또는 `Accept: application/json` 헤더 시 JSON 응답.
    JSON 모드는 칩 UI 의 AJAX 라이브 검색용 (새로고침 없이 표·KPI 갱신).
    """
    from lemouton.sourcing.models import Option, Model
    from lemouton.inventory.models import InventoryLocation
    from shared.search import split_tokens, apply_and_filter
    from shared.inventory_stock import get_stock_batch, get_stock_by_location_batch
    from sqlalchemy import func
    from sqlalchemy.orm import contains_eager

    page = max(1, int(request.args.get('page', 1)))
    # ★ 엑셀식 정렬·필터가 전체 데이터 대상이 되도록 한 페이지에 전체 로드 (기본 2000)
    page_size = min(5000, int(request.args.get('page_size', 2000)))
    q = (request.args.get('q') or '').strip()
    brand = (request.args.get('brand') or '').strip()
    in_stock_only = request.args.get('in_stock_only') == '1'
    search_tokens = split_tokens(q)
    want_json = (
        request.args.get('format') == 'json'
        or 'application/json' in (request.headers.get('Accept') or '')
    )

    s = SessionLocal()
    try:
        # contains_eager → 템플릿 o.model.* 접근 시 추가 SELECT 0회 (N+1 제거)
        # [2026-05-28] is_active=True 만 표시 (사용자 OFF 비활성 옵션은 DB 보존하되 화면에서 숨김)
        query = (
            s.query(Option)
            .join(Model, Option.model_code == Model.model_code)
            .options(contains_eager(Option.model))
            .filter(Option.is_active == True)  # noqa: E712
        )
        # 다중 키워드 AND 교집합 — 토큰별 OR (SKU·바코드·브랜드·제품명·모델명·컬러·사이즈)
        query = apply_and_filter(
            query, search_tokens,
            Option.canonical_sku, Option.boxhero_sku, Option.barcode,
            Model.brand, Model.model_name_display, Model.model_name_raw, Model.model_code,
            Option.color_display, Option.color_code,
            Option.size_display, Option.size_code,
        )
        if brand:
            query = query.filter(Model.brand == brand)

        all_skus_for_stock = [r[0] for r in query.with_entities(Option.canonical_sku).all()]
        stock_map = get_stock_batch(s, all_skus_for_stock)
        if in_stock_only:
            in_skus = {sk for sk, st in stock_map.items() if st > 0}
            query = query.filter(Option.canonical_sku.in_(in_skus) if in_skus else False)

        total = query.count()
        # [2026-05-27] 정렬: 브랜드 > 카테고리 > 모델명 > 색상 > 사이즈 (사용자 룰)
        items = (
            query.order_by(Model.brand, Model.category, Model.model_name_display,
                           Option.color_display, Option.size_display)
            .offset((page - 1) * page_size).limit(page_size).all()
        )

        # [2026-05-28] 활성 옵션만 카운트 (비활성 31건 숨김 룰)
        all_options_count = s.query(func.count(Option.canonical_sku)).filter(Option.is_active == True).scalar() or 0  # noqa: E712
        in_stock_count = sum(1 for st in stock_map.values() if st > 0)
        total_qty = sum(stock_map.values())
        kpi = {
            'all_options': all_options_count,
            'in_stock': in_stock_count,
            'total_qty': int(total_qty),
        }

        # ★ 검색 결과 요약 (시안 A — 요약 배너) — 검색어 있을 때만
        _summary_rows = query.with_entities(Option.boxhero_avg_purchase_price, Model.brand).all()
        _prices = [r[0] for r in _summary_rows if r[0] and r[0] > 0]
        _avg_price = round(sum(_prices) / len(_prices)) if _prices else 0
        _brands = sorted({r[1] for r in _summary_rows if r[1]})
        search_summary = {
            'active': bool(q or brand or in_stock_only),
            'q': q,
            'option_count': total,
            'avg_price': _avg_price,
            'priced_count': len(_prices),
            'total_stock': int(total_qty),
            'in_stock_count': in_stock_count,
            'brands': _brands,
        }

        # 색상·제품명 정리 — shared.product_display 헬퍼 (전 시스템 통일)
        from shared.product_display import compute_display_maps
        cleaned_color, display_pname = compute_display_maps(items)

        # 위치별 재고 — 사용자 spec: ... / 총재고 / 위치별 재고 N
        locs = (
            s.query(InventoryLocation)
            .filter(InventoryLocation.deleted_at.is_(None))
            .order_by(InventoryLocation.sort_order, InventoryLocation.id)
            .all()
        )
        # 위치별 재고 — 한 쿼리로 N 위치 × M SKU pivot (구 for loc 루프 = 2N 쿼리)
        page_skus = [o.canonical_sku for o in items]
        per_loc_stock = get_stock_by_location_batch(s, page_skus)

        # ★ ⑤ 역참조 — 제품별 사용처(모음전·옵션) batch 조회 (N+1 회피)
        # OptionProductLink 를 product_canonical_sku 로 한 번에 조회.
        usage_map: dict[str, int] = {}
        try:
            from lemouton.inventory.models import OptionProductLink
            if page_skus:
                _link_rows = (
                    s.query(OptionProductLink.product_canonical_sku)
                    .filter(OptionProductLink.product_canonical_sku.in_(page_skus))
                    .all()
                )
                for (psku,) in _link_rows:
                    usage_map[psku] = usage_map.get(psku, 0) + 1
        except Exception:
            usage_map = {}

        if want_json:
            from flask import jsonify
            def _size(o):
                return (o.size_display or o.size_code or '').strip() or 'FREE'
            # [2026-05-27] SKU 표시 룰: canonical_sku 가 옛 sku (한글) 형식이면 boxhero_sku 사용
            def _display_sku(o):
                cs = o.canonical_sku
                if cs and not cs.startswith('SKU-') and o.boxhero_sku:
                    return o.boxhero_sku
                return cs
            rows = [
                {
                    'sku': _display_sku(o),
                    'bh': o.boxhero_sku or '',
                    'barcode': o.barcode or '',
                    'name': display_pname.get(o.canonical_sku, ''),
                    'name_raw': (o.model.model_name_display or o.model.model_name_raw or '') if o.model else '',
                    'article_no': (getattr(o.model, 'article_no', None) or '') if o.model else '',
                    'model_code': (o.model.model_code or '') if o.model else '',
                    'brand': (o.model.brand or '') if o.model else '',
                    'color': cleaned_color.get(o.canonical_sku, 'ONE Color'),
                    'color_raw': o.color_display or o.color_code or '',
                    'size': _size(o),
                    'size_raw': o.size_display or o.size_code or '',
                    'avg': int(o.boxhero_avg_purchase_price or 0),
                    'stock': int(stock_map.get(o.canonical_sku, 0)),
                    'loc_stock': {str(loc.id): int(per_loc_stock.get(o.canonical_sku, {}).get(loc.id, 0)) for loc in locs},
                    'usage': int(usage_map.get(o.canonical_sku, 0)),  # ★ ⑤ 역참조 사용처 개수
                }
                for o in items
            ]
            return jsonify({
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size,
                'items': rows,
                'kpi': kpi,
                'locs': [{'id': loc.id, 'name': loc.name} for loc in locs],
                'search_summary': search_summary,
            })

        brands = [b for (b,) in s.query(Model.brand).distinct().filter(Model.brand.isnot(None)).all()]
        return render_template(
            'inventory/data/items.html',
            active='data-items',
            items=items,
            total=total, page=page, page_size=page_size, total_pages=(total + page_size - 1) // page_size,
            q=q, brand=brand, in_stock_only=in_stock_only,
            search_tokens=search_tokens,
            brands=sorted(brands),
            stock_map=stock_map,
            cleaned_color=cleaned_color,  # ★ LCP strip 색상
            display_pname=display_pname,  # ★ brand+모델명 (색상 X, brand 중복 X)
            kpi=kpi,
            locs=locs,                    # ★ 위치별 재고 컬럼 헤더용
            per_loc_stock=per_loc_stock,  # ★ {sku: {loc_id: stock}}
            search_summary=search_summary,  # ★ 검색 결과 요약 배너
            usage_map=usage_map,          # ★ ⑤ {sku: 사용처(모음전·옵션) 개수}
        )
    finally:
        s.close()


# ============ 신규 제품 1건 추가 (박스히어로 1:1 모달 폼) ============

@bp.post('/data/items/create')
def data_items_create():
    """신규 옵션 1건 생성 — 박스히어로 '제품 추가' 폼 1:1.

    필드: SKU(자동/직접), 제품명, 바코드(자동/직접), 브랜드, 품번, 색상, 사이즈,
          카테고리, 평균매입가, 메모.
    - Model 은 upsert (model_code = derive(brand, article_no or model_name, sku))
    - Option 은 신규 INSERT — canonical_sku 중복 시 자동 재생성 시도
    """
    # [2026-05-28] Phase 1-4 — shared.sku_format 통일
    from shared.sku_format import gen_sku as _shared_gen_sku, gen_barcode as _shared_gen_barcode

    from lemouton.sourcing.master import upsert_model
    from lemouton.inventory.boxhero_import import _derive_model_code, _clean_article_no
    from lemouton.inventory.inbound import create_inbound
    from lemouton.inventory.models import InventoryLocation

    f = request.form
    canonical_sku = (f.get('canonical_sku') or '').strip()
    model_name = (f.get('model_name') or '').strip()
    barcode = (f.get('barcode') or '').strip()
    brand = (f.get('brand') or '').strip() or '미상'
    article_no_in = (f.get('article_no') or '').strip()
    color = (f.get('color') or '').strip() or 'ONE'
    size = (f.get('size') or '').strip() or 'FREE'
    category = (f.get('category') or '').strip()
    avg_price_raw = (f.get('avg_purchase_price') or '').strip()
    memo = (f.get('memo') or '').strip()

    if not model_name:
        flash('모델명은 필수입니다.', 'error')
        return redirect(url_for('inventory.data_items') + '#new')

    # [2026-05-28] Phase 1-4 — shared.sku_format 모듈 사용
    _gen_sku = _shared_gen_sku
    _gen_barcode = _shared_gen_barcode

    s = SessionLocal()
    try:
        # SKU 중복 회피 — 비었으면 생성, 입력값이 이미 있으면 에러
        if not canonical_sku:
            for _ in range(30):
                cand = _gen_sku()
                if not s.query(Option).filter_by(canonical_sku=cand).first():
                    canonical_sku = cand
                    break
            else:
                flash('SKU 자동 생성 실패 — 다시 시도해 주세요.', 'error')
                return redirect(url_for('inventory.data_items') + '#new')
        else:
            if s.query(Option).filter_by(canonical_sku=canonical_sku).first():
                flash(f'SKU 중복 — "{canonical_sku}" 는 이미 사용 중입니다.', 'error')
                return redirect(url_for('inventory.data_items') + '#new')

        if not barcode:
            barcode = _gen_barcode()

        try:
            avg_price = int(avg_price_raw.replace(',', '').replace(' ', '')) if avg_price_raw else 0
        except ValueError:
            avg_price = 0

        # [2026-05-28] Phase 2-2 — 사용자 룰: 자동 모음전 등록 X.
        #   "register_as_bundle" 체크박스 시에만 정상 model 생성 → 모음전 list 에 노출.
        #   기본 (체크 X) → "단독_SKU-XXX" prefix 모델로 분리 (모음전 list 에서 제외).
        register_as_bundle = (f.get('register_as_bundle') or '').lower() in ('on', 'true', '1')

        if register_as_bundle:
            # 정상 모음전 등록 — Model 자동 생성·갱신
            model_code = _derive_model_code(brand, article_no_in or model_name, canonical_sku)
        else:
            # 단독 옵션 — "단독_{canonical_sku}" 모델로 분리 (모음전 X)
            model_code = f'단독_{canonical_sku}'

        upsert_model(
            s,
            model_code=model_code,
            model_name_raw=model_name[:255],
            brand=brand[:100],
        )
        m_obj = s.query(Model).filter_by(model_code=model_code).first()
        if m_obj:
            if model_name and not (m_obj.model_name_display or '').strip():
                m_obj.model_name_display = model_name[:255]
            if article_no_in and not (getattr(m_obj, 'article_no', None) or '').strip():
                m_obj.article_no = _clean_article_no(article_no_in)[:64]
            if category and not (getattr(m_obj, 'category', None) or '').strip():
                m_obj.category = category[:100]
            if memo and not (getattr(m_obj, 'note', None) or '').strip():
                m_obj.note = memo

        opt = Option(
            canonical_sku=canonical_sku,
            model_code=model_code,
            color_code=color[:32],
            color_display=color[:64],
            size_code=size[:32],
            size_display=size[:64],
            boxhero_sku=canonical_sku,
            barcode=barcode[:64],
            # 매입가 — 위치 입력이 없으면 직접 set (입고가 없으면 moving avg 계산 안 됨)
            boxhero_avg_purchase_price=avg_price,
            boxhero_stock_total=0,
        )
        s.add(opt)
        s.flush()  # Option 확정 후에 create_inbound 가 조회할 수 있도록

        # 📍 위치별 초기 재고 — stock_loc_<id> 폼 필드를 모두 받아 InventoryTx('in') 로 등록
        active_locs = (
            s.query(InventoryLocation)
            .filter(InventoryLocation.deleted_at.is_(None))
            .all()
        )
        initial_stock_total = 0
        initial_stock_detail: list[str] = []
        for loc in active_locs:
            raw = (f.get(f'stock_loc_{loc.id}') or '').strip()
            if not raw:
                continue
            try:
                qty = int(raw.replace(',', ''))
            except ValueError:
                continue
            if qty <= 0:
                continue
            create_inbound(
                s,
                location_id=loc.id,
                option_canonical_sku=canonical_sku,
                qty=qty,
                unit_purchase_price=avg_price,
                memo='신규 제품 등록 초기 재고',
                created_by='제품 추가 폼',
            )
            initial_stock_total += qty
            initial_stock_detail.append(f'{loc.name} {qty}')

        s.commit()

        msg = f'✅ 제품 추가 완료 — {canonical_sku} ({model_name} / {color} / {size})'
        if initial_stock_total > 0:
            msg += f' · 초기 재고 {initial_stock_total}개 ({", ".join(initial_stock_detail)})'
        flash(msg, 'success')
    except Exception as e:
        s.rollback()
        flash(f'제품 추가 실패: {e}', 'error')
        return redirect(url_for('inventory.data_items') + '#new')
    finally:
        s.close()

    return redirect(url_for('inventory.data_items'))


# ============ 옵션 매트릭스 일괄 생성 (색상 N × 사이즈 M) ============

@bp.post('/data/items/bulk_create')
def data_items_bulk_create():
    """옵션 매트릭스 일괄 생성 — 1회 POST 로 N개 SKU 생성.

    프론트 폼: bulk_cells_json = [{color, size, qty}, ...]  체크된 셀만 포함.
    공통 필드: model_name, brand, article_no, category, avg_purchase_price, memo
    위치 분배: qty > 0 인 셀은 모두 "기본 위치" 1곳으로 입고 (사용자 결정 옵션 (1)).
    무결성: SKU/바코드 자동 부여, 동일 (model_code, color, size) 중복 거부.
            중간 1건이라도 실패하면 전체 롤백 (부분 성공 금지).
    """
    import random
    import string
    import json as _json

    from lemouton.sourcing.master import upsert_model
    from lemouton.inventory.boxhero_import import _derive_model_code, _clean_article_no
    from lemouton.inventory.inbound import create_inbound
    from lemouton.inventory.models import InventoryLocation

    f = request.form
    model_name = (f.get('model_name') or '').strip()
    brand = (f.get('brand') or '').strip() or '미상'
    article_no_in = (f.get('article_no') or '').strip()
    category = (f.get('category') or '').strip()
    avg_price_raw = (f.get('avg_purchase_price') or '').strip()
    memo = (f.get('memo') or '').strip()
    cells_json = (f.get('bulk_cells_json') or '').strip()

    if not model_name:
        flash('모델명은 필수입니다.', 'error')
        return redirect(url_for('inventory.data_items') + '#new')

    try:
        cells = _json.loads(cells_json) if cells_json else []
    except Exception:
        flash('옵션 데이터 파싱 실패 — 다시 시도해 주세요.', 'error')
        return redirect(url_for('inventory.data_items') + '#new')

    if not isinstance(cells, list) or not cells:
        flash('최소 1개 옵션을 체크하세요.', 'error')
        return redirect(url_for('inventory.data_items') + '#new')

    # 셀 정규화 + 중복 (color, size) 검사
    norm_cells: list[dict] = []
    seen_combos: set[tuple[str, str]] = set()
    for c in cells:
        color = (str(c.get('color', '')).strip() or 'ONE')[:64]
        size = (str(c.get('size', '')).strip() or 'FREE')[:64]
        try:
            qty = int(c.get('qty') or 0)
        except (TypeError, ValueError):
            qty = 0
        if qty < 0:
            qty = 0
        key = (color, size)
        if key in seen_combos:
            flash(f'중복 옵션 — ({color} / {size}) 가 2번 이상 입력되었습니다.', 'error')
            return redirect(url_for('inventory.data_items') + '#new')
        seen_combos.add(key)
        norm_cells.append({'color': color, 'size': size, 'qty': qty})

    try:
        avg_price = int(avg_price_raw.replace(',', '').replace(' ', '')) if avg_price_raw else 0
    except ValueError:
        avg_price = 0

    def _gen_sku() -> str:
        chars = string.ascii_uppercase + string.digits
        return 'SKU-' + ''.join(random.choices(chars, k=8))

    def _gen_barcode() -> str:
        digits = '200' + ''.join(random.choices(string.digits, k=9))
        chk = sum(int(d) * (3 if i % 2 else 1) for i, d in enumerate(digits))
        return digits + str((10 - chk % 10) % 10)

    s = SessionLocal()
    try:
        # 기본 위치 결정 — is_default=True 우선, 없으면 가장 오래된 활성 위치
        default_loc = (
            s.query(InventoryLocation)
            .filter(InventoryLocation.deleted_at.is_(None))
            .filter(InventoryLocation.is_default.is_(True))
            .first()
        )
        if not default_loc:
            default_loc = (
                s.query(InventoryLocation)
                .filter(InventoryLocation.deleted_at.is_(None))
                .order_by(InventoryLocation.id.asc())
                .first()
            )
        # qty > 0 셀이 하나라도 있는데 활성 위치가 없으면 막음
        any_qty = any(c['qty'] > 0 for c in norm_cells)
        if any_qty and not default_loc:
            flash('재고를 입력했지만 활성 위치가 없습니다 — 먼저 위치 관리에서 등록하세요.', 'error')
            return redirect(url_for('inventory.data_items') + '#new')

        # Model upsert (1회) — 모든 셀이 같은 모델 공유
        model_code = _derive_model_code(brand, article_no_in or model_name, model_name)
        upsert_model(
            s,
            model_code=model_code,
            model_name_raw=model_name[:255],
            brand=brand[:100],
        )
        m_obj = s.query(Model).filter_by(model_code=model_code).first()
        if m_obj:
            if model_name and not (m_obj.model_name_display or '').strip():
                m_obj.model_name_display = model_name[:255]
            if article_no_in and not (getattr(m_obj, 'article_no', None) or '').strip():
                m_obj.article_no = _clean_article_no(article_no_in)[:64]
            if category and not (getattr(m_obj, 'category', None) or '').strip():
                m_obj.category = category[:100]
            if memo and not (getattr(m_obj, 'note', None) or '').strip():
                m_obj.note = memo

        # DB 사전 중복 검사 — 동일 (model_code, color, size) 활성 옵션 존재 시 거부.
        # is_active=False (사용자 OFF) 옵션은 UI 에서 숨겨져 있으므로 일괄 생성 시 무시.
        existing_pairs: set[tuple[str, str]] = set()
        dup_in_db = (
            s.query(Option.color_code, Option.size_code)
            .filter(Option.model_code == model_code)
            .filter(Option.is_active == True)  # noqa: E712
            .all()
        )
        for cc, sc in dup_in_db:
            existing_pairs.add(((cc or 'ONE')[:64], (sc or 'FREE')[:64]))

        clashes = []
        for c in norm_cells:
            if (c['color'][:32], c['size'][:32]) in existing_pairs:
                clashes.append(f"({c['color']} / {c['size']})")
        if clashes:
            flash('이미 등록된 조합이 있어 일괄 생성을 취소했습니다: ' + ', '.join(clashes[:5])
                  + (f' 외 {len(clashes) - 5}건' if len(clashes) > 5 else ''), 'error')
            s.rollback()
            return redirect(url_for('inventory.data_items') + '#new')

        # SKU 일괄 생성 — 각 셀마다 SKU/바코드 자동 부여
        created_skus: list[str] = []
        total_inbound = 0
        for c in norm_cells:
            # SKU 중복 회피
            canonical_sku = ''
            for _ in range(30):
                cand = _gen_sku()
                if not s.query(Option).filter_by(canonical_sku=cand).first() and cand not in created_skus:
                    canonical_sku = cand
                    break
            if not canonical_sku:
                raise RuntimeError('SKU 자동 생성 실패 — 다시 시도해 주세요')

            opt = Option(
                canonical_sku=canonical_sku,
                model_code=model_code,
                color_code=c['color'][:32],
                color_display=c['color'][:64],
                size_code=c['size'][:32],
                size_display=c['size'][:64],
                boxhero_sku=canonical_sku,
                barcode=_gen_barcode()[:64],
                boxhero_avg_purchase_price=avg_price,
                boxhero_stock_total=0,
            )
            s.add(opt)
            s.flush()
            created_skus.append(canonical_sku)

            if c['qty'] > 0 and default_loc:
                create_inbound(
                    s,
                    location_id=default_loc.id,
                    option_canonical_sku=canonical_sku,
                    qty=c['qty'],
                    unit_purchase_price=avg_price,
                    memo='매트릭스 일괄 생성 초기 재고',
                    created_by='제품 추가 폼 (일괄)',
                )
                total_inbound += c['qty']

        s.commit()

        loc_label = f' · 기본 위치 [{default_loc.name}] 입고 {total_inbound}개' if total_inbound > 0 else ''
        flash(f'✅ {len(created_skus)}개 SKU 일괄 추가 완료 — {model_name}{loc_label}', 'success')
    except Exception as e:
        s.rollback()
        flash(f'일괄 생성 실패 (전체 롤백): {e}', 'error')
        return redirect(url_for('inventory.data_items') + '#new')
    finally:
        s.close()

    return redirect(url_for('inventory.data_items'))


# ============ 제품 마스터 전체 삭제 (TEST 리셋) ============

DATA_DIR = Path(__file__).resolve().parents[3] / 'data'
# BACKUP_ROOT: 로컬은 OneDrive 폴더 5단계 위, 컨테이너(Fly.io 등) 는 깊이 부족 → env 또는 안전 폴백
import os as _os
_parents = Path(__file__).resolve().parents
if _os.environ.get("BACKUP_ROOT"):
    BACKUP_ROOT = Path(_os.environ["BACKUP_ROOT"])
elif len(_parents) > 5:
    BACKUP_ROOT = _parents[5] / '백업'
else:
    # 컨테이너 환경 — 임시 디렉토리 사용 (Supabase 가 별도 백업 처리)
    BACKUP_ROOT = Path('/tmp/모음전_백업') if _os.name != 'nt' else _parents[len(_parents) - 1] / 'tmp_backup'

# wipe 대상 — FK 의존 순서로 (자식 → 부모)
# Model/Option 을 참조하는 모든 테이블. 누락 시 FK 위반 → 백업으로 복원 가능.
_WIPE_TABLES_IN_ORDER = [
    'option_source_urls',           # FK options (CASCADE)
    'option_price_config',          # FK options (CASCADE)
    'option_source_links',          # FK options
    'option_account_registrations', # FK options
    'option_benefit_overrides',     # FK options (가정)
    'price_track_history',          # FK options
    'model_source_links',           # FK models
    'bundle_account_registrations', # FK models
    'bundle_source_urls',           # FK models
    'combo_sets',                   # FK models
    # 재고/거래 — canonical_sku 기반이지만 FK 없을 수 있음 (안전상 비움)
    'inventory_txs',
    'inventory_pending',
    'inventory_counts',
    'inventory_count_sheet_items',
    'inventory_count_sheets',
    'inventory_safety_stock',
    'inventory_products',
    'item_attribute_values',
    'purchase_orders',
    'sales_orders',
    'return_orders',
    # 디스커버리 큐
    'discovery_queue',
    # 마지막: options → models
    'options',
    'models',
]


def _backup_data_dir() -> Path:
    """data 폴더를 백업/data_YYYYMMDD_HHMMSS/ 로 복사. 잠긴 파일은 robocopy 로 skip."""
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = BACKUP_ROOT / f'data_{ts}_pre_wipe'
    BACKUP_ROOT.mkdir(parents=True, exist_ok=True)
    dst.mkdir(parents=True, exist_ok=True)

    # robocopy 로 잠긴 파일 skip + 브라우저 캐시 폴더 제외
    import subprocess
    args = [
        'robocopy', str(DATA_DIR), str(dst),
        '/E', '/R:0', '/W:0', '/MT:8', '/NFL', '/NDL',
        '/XD', 'Sessions', 'Network', 'Safe Browsing Network',
        'GPUCache', 'ShaderCache', 'GraphiteDawnCache',
    ]
    try:
        proc = subprocess.run(args, capture_output=True, timeout=300)
        # robocopy exit 0~7 = 정상, 8+ = 실패
        if proc.returncode >= 8:
            raise RuntimeError(f'robocopy exit={proc.returncode}: {proc.stderr.decode(errors="replace")[:500]}')
    except FileNotFoundError:
        # robocopy 없는 환경 (테스트) → 그냥 shutil
        shutil.copytree(str(DATA_DIR), str(dst), dirs_exist_ok=True)
    return dst


@bp.post('/data/items/wipe')
def data_items_wipe():
    """제품 마스터 전체 삭제 — TEST 리셋용. 자동 백업 후 진행.

    안전장치:
    1. POST 만 허용 (CSRF token 검증은 Flask-WTF 없으므로 confirm 키워드로 대체)
    2. 자동으로 data 폴더 통째 백업 (백업/data_YYYYMMDD_HHMMSS_pre_wipe/)
    3. 사용자가 'DELETE ALL' 정확히 입력해야 진행
    """
    confirm = (request.form.get('confirm') or '').strip()
    if confirm != 'DELETE ALL':
        flash("확인 키워드 불일치 — 'DELETE ALL' 입력 시에만 삭제 진행됩니다.", 'error')
        return redirect(url_for('inventory.data_items'))

    # 1. 자동 백업
    try:
        backup_path = _backup_data_dir()
    except Exception as e:
        flash(f'자동 백업 실패 — 삭제 중단: {e}', 'error')
        return redirect(url_for('inventory.data_items'))

    # 2. DB wipe — dialect 별 분기 (SQLite: PRAGMA + DELETE, Postgres: TRUNCATE CASCADE)
    s = SessionLocal()
    deleted_counts = {}
    dialect = s.bind.dialect.name  # 'sqlite' | 'postgresql'
    try:
        # 카운트 (before)
        before_models = s.execute(sa_text('SELECT COUNT(*) FROM models')).scalar() or 0
        before_options = s.execute(sa_text('SELECT COUNT(*) FROM options')).scalar() or 0

        # 테이블 존재 확인용 쿼리 (dialect 별)
        def _table_exists(tbl: str) -> bool:
            if dialect == 'sqlite':
                return s.execute(sa_text(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name=:n"
                ), {'n': tbl}).first() is not None
            # Postgres
            return s.execute(sa_text(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_schema = current_schema() AND table_name = :n"
            ), {'n': tbl}).first() is not None

        if dialect == 'sqlite':
            # FK 잠시 OFF (자식→부모 의존 순서 DELETE 안전망)
            s.execute(sa_text('PRAGMA foreign_keys = OFF'))
            for tbl in _WIPE_TABLES_IN_ORDER:
                try:
                    if _table_exists(tbl):
                        res = s.execute(sa_text(f'DELETE FROM {tbl}'))
                        deleted_counts[tbl] = res.rowcount or 0
                except Exception as e:
                    deleted_counts[tbl] = f'ERR: {e}'
            s.execute(sa_text('PRAGMA foreign_keys = ON'))
        else:
            # Postgres: 존재하는 테이블만 모아 TRUNCATE ... RESTART IDENTITY CASCADE 1발 처리
            tbls_present = [t for t in _WIPE_TABLES_IN_ORDER if _table_exists(t)]
            if tbls_present:
                tbl_list = ', '.join(f'"{t}"' for t in tbls_present)
                s.execute(sa_text(f'TRUNCATE TABLE {tbl_list} RESTART IDENTITY CASCADE'))
                for t in tbls_present:
                    deleted_counts[t] = '(truncated)'

        s.commit()

        after_models = s.execute(sa_text('SELECT COUNT(*) FROM models')).scalar() or 0
        after_options = s.execute(sa_text('SELECT COUNT(*) FROM options')).scalar() or 0

        flash(
            f"✅ 제품 마스터 전체 삭제 완료\n"
            f"  - 모델: {before_models} → {after_models}\n"
            f"  - 옵션: {before_options} → {after_options}\n"
            f"  - 자동 백업: {backup_path}\n"
            f"  → 이제 [📥 박스히어로 엑셀 업로드] 에서 재로드 가능합니다.",
            'success'
        )
    except Exception as e:
        s.rollback()
        flash(f'삭제 중 오류 (롤백됨, 백업은 남아있음 {backup_path}): {e}', 'error')
    finally:
        s.close()

    return redirect(url_for('inventory.data_items'))


# ============ 속성 (Task 1.4) ============

@bp.get('/data/attributes')
def data_attributes():
    s = SessionLocal()
    try:
        attrs = attrs_svc.list_active(s)
        return render_template(
            'inventory/data/attributes.html',
            active='data-attrs',
            attrs=attrs,
            valid_types=attrs_svc.VALID_TYPES,
        )
    finally:
        s.close()


@bp.post('/data/attributes/create')
def data_attributes_create():
    s = SessionLocal()
    try:
        try:
            attrs_svc.create(s, request.form.get('name', ''), request.form.get('type', 'text'))
            s.commit()
            flash("속성 추가됨", 'success')
        except ValueError as e:
            s.rollback()
            flash(str(e), 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.data_attributes'))


@bp.post('/data/attributes/<int:attr_id>/delete')
def data_attributes_delete(attr_id):
    s = SessionLocal()
    try:
        try:
            attrs_svc.delete(s, attr_id)
            s.commit()
            flash("삭제됨", 'success')
        except ValueError as e:
            s.rollback()
            flash(str(e), 'error')
    finally:
        s.close()
    return redirect(url_for('inventory.data_attributes'))


# ============ 묶음제품 (Bundle) CRUD ============

@bp.get('/data/bundles')
def data_bundles():
    bundles = _load_bundles()
    return render_template('inventory/data/bundles.html',
                           active='data-bundles', bundles=bundles)


def _option_price_map(s) -> dict:
    """SKU → {model_code, color, size, avg_purchase, sale_price} 맵.

    옵션이 모델 다르면 가격 전혀 다를 수 있어 — Bundle 폼에서 참조 가격 표시용.
    """
    opts = s.query(Option).order_by(Option.canonical_sku).limit(2000).all()
    return {
        o.canonical_sku: {
            'sku': o.canonical_sku,
            'model': o.model_code or '',
            'color': o.color_display or o.color_code or '',
            'size': o.size_display or o.size_code or '',
            'avg_purchase': int(o.boxhero_avg_purchase_price or 0),
            'price': int(getattr(o, 'price_default', 0) or 0),
        }
        for o in opts
    }


@bp.get('/data/bundles/new')
def data_bundle_new():
    s = SessionLocal()
    try:
        opt_map = _option_price_map(s)
    finally:
        s.close()
    return render_template('inventory/data/bundle_form.html',
                           active='data-bundles', bundle=None,
                           available_skus=list(opt_map.keys()),
                           opt_price_map=opt_map)


@bp.get('/data/bundles/<bundle_id>/edit')
def data_bundle_edit(bundle_id):
    bundles = _load_bundles()
    b = next((b for b in bundles if str(b.get('id')) == str(bundle_id)), None)
    if not b:
        flash('묶음을 찾을 수 없습니다.', 'error')
        return redirect(url_for('inventory.data_bundles'))
    s = SessionLocal()
    try:
        opt_map = _option_price_map(s)
    finally:
        s.close()
    return render_template('inventory/data/bundle_form.html',
                           active='data-bundles', bundle=b,
                           available_skus=list(opt_map.keys()),
                           opt_price_map=opt_map)


@bp.post('/data/bundles/save', defaults={'bundle_id': None})
@bp.post('/data/bundles/<bundle_id>/save')
def data_bundle_save(bundle_id):
    bundles = _load_bundles()
    sku = (request.form.get('sku') or '').strip()
    name = (request.form.get('name') or '').strip()
    if not sku or not name:
        flash('묶음 SKU·이름은 필수입니다.', 'error')
        return redirect(url_for('inventory.data_bundles'))

    comp_skus = request.form.getlist('comp_sku')
    comp_qtys = request.form.getlist('comp_qty')
    components = []
    for csk, cq in zip(comp_skus, comp_qtys):
        csk = (csk or '').strip()
        try:
            cq_i = int(cq or 1)
        except ValueError:
            cq_i = 1
        if csk and cq_i > 0:
            components.append({'sku': csk, 'qty': cq_i})

    try:
        price = int(request.form.get('price') or 0)
    except ValueError:
        price = 0

    if bundle_id:
        existing = next((b for b in bundles if str(b.get('id')) == str(bundle_id)), None)
        if existing:
            existing.update({'sku': sku, 'name': name, 'components': components,
                             'price': price, 'memo': request.form.get('memo') or '',
                             'updated_at': datetime.now().isoformat()})
    else:
        # SKU duplicate check
        if any(b.get('sku') == sku for b in bundles):
            flash(f'묶음 SKU 중복: {sku}', 'error')
            return redirect(url_for('inventory.data_bundle_new'))
        new_id = max([b.get('id', 0) for b in bundles] + [0]) + 1
        bundles.append({
            'id': new_id, 'sku': sku, 'name': name,
            'components': components, 'price': price,
            'memo': request.form.get('memo') or '',
            'created_at': datetime.now().isoformat(),
        })
    _save_bundles(bundles)
    flash('묶음 저장됨', 'success')
    return redirect(url_for('inventory.data_bundles'))


@bp.post('/data/bundles/<bundle_id>/delete')
def data_bundle_delete(bundle_id):
    bundles = _load_bundles()
    bundles = [b for b in bundles if str(b.get('id')) != str(bundle_id)]
    _save_bundles(bundles)
    flash('묶음 삭제됨', 'success')
    return redirect(url_for('inventory.data_bundles'))


# ============ 거래처 (Task 1.5) — ADR-003 텍스트만 ============

@bp.get('/data/partners')
def data_partners():
    s = SessionLocal()
    try:
        recent = partner_svc.recent_labels(s, limit=100)
        return render_template(
            'inventory/data/partners.html',
            active='data-partners',
            recent=recent,
        )
    finally:
        s.close()


# ============ 가격 템플릿 (Task 1.6) — 기존 PriceTemplate 활용 ============

@bp.get('/data/price-templates')
def data_price_templates():
    from lemouton.templates.models import PriceTemplate
    s = SessionLocal()
    try:
        templates = s.query(PriceTemplate).order_by(PriceTemplate.id).all()
        return render_template(
            'inventory/data/price_templates.html',
            active='data-pt',
            templates=templates,
        )
    finally:
        s.close()


@bp.post('/data/price-templates/<int:tpl_id>/update-margin')
def data_price_templates_update_margin(tpl_id):
    """박스히어로 마진 4컬럼 수정 (R2 핵심)."""
    from lemouton.templates.models import PriceTemplate
    s = SessionLocal()
    try:
        tpl = s.query(PriceTemplate).filter(PriceTemplate.id == tpl_id).first()
        if not tpl:
            flash("템플릿 없음", 'error')
            return redirect(url_for('inventory.data_price_templates'))
        tpl.boxhero_margin_mode_self = request.form.get('mode_self', 'rate')
        tpl.boxhero_margin_value_self = int(request.form.get('value_self', 2500) or 2500)
        tpl.boxhero_margin_mode_external = request.form.get('mode_external', 'rate')
        tpl.boxhero_margin_value_external = int(request.form.get('value_external', 2000) or 2000)
        s.commit()
        flash(f"'{tpl.name}' 사입 마진 업데이트됨", 'success')
    finally:
        s.close()
    return redirect(url_for('inventory.data_price_templates'))


@bp.route('/data/items/export.xlsx', methods=['GET', 'POST'])
def data_items_export():
    """[2026-05-25 D-6 v2] 사용자 명시 양식 — 평균매입가 추가.

    헤더 (10 base + N 위치):
      A SKU / B 바코드 / C 품번 / D 브랜드 / E 카테고리 / F 모델명 /
      G 색상 / H 사이즈 / I 평균매입가 / J 총재고 / K 위치1 / L 위치2 ...

    품번 NULL/한글 → '-' (사용자 룰).
    색상·사이즈 빈값 → '-'.

    필터 반영: filtered=1 + skus[] (POST) 전달 시 그 옵션만 export. 미전달 시 전체.
    """
    from io import BytesIO
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from sqlalchemy.orm import joinedload
    from lemouton.sourcing.models import Option, Model
    from lemouton.inventory.models import InventoryLocation
    from shared.inventory_stock import get_stock_batch

    s = SessionLocal()
    try:
        # 화면에서 필터된 SKU 목록 — filtered=1 이면 그 옵션만 export (없으면 전체)
        is_filtered = request.values.get('filtered') == '1'
        # [2026-05-28] is_active=True 만 export (사용자 OFF 비활성은 숨김)
        q = (s.query(Option).options(joinedload(Option.model))
             .filter(Option.is_active == True))  # noqa: E712
        if is_filtered:
            q = q.filter(Option.canonical_sku.in_(request.values.getlist('skus')))
        # [2026-05-27] 정렬: 브랜드 > 카테고리 > 모델명 > 색상 > 사이즈 (사용자 룰)
        options = (
            q.join(Model, Option.model_code == Model.model_code)
             .order_by(Model.brand, Model.category, Model.model_name_display,
                       Option.color_display, Option.size_display).all()
        )
        all_skus = [o.canonical_sku for o in options]

        total_stock_map = get_stock_batch(s, all_skus)
        locs = (
            s.query(InventoryLocation)
            .filter(InventoryLocation.deleted_at.is_(None))
            .order_by(InventoryLocation.sort_order, InventoryLocation.id)
            .all()
        )
        per_loc_stock = {}
        for loc in locs:
            loc_map = get_stock_batch(s, all_skus, location_id=loc.id)
            for sku, st in loc_map.items():
                per_loc_stock.setdefault(sku, {})[loc.id] = st

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '재고관리'

        # [2026-05-25 D-6 v2] 사용자 양식 — 10 base + N 위치 (평균매입가 추가)
        headers = ['SKU', '바코드', '품번', '브랜드', '카테고리', '모델명',
                   '색상', '사이즈', '평균매입가', '총재고']
        for loc in locs:
            headers.append(loc.name)
        ws.append(headers)

        for o in options:
            m = o.model
            barcode = o.barcode or ''
            brand = (m.brand or '') if m else ''
            article = (getattr(m, 'article_no', None) or '-') if m else '-'
            category = (getattr(m, 'category', None) or '') if m else ''
            mname = ((getattr(m, 'model_name_display', None) or
                      getattr(m, 'model_name_raw', None)) if m else '') or ''
            color = o.color_display or o.color_code or '-'
            size = o.size_display or o.size_code or '-'
            avg = int(o.boxhero_avg_purchase_price or 0)
            total = int(total_stock_map.get(o.canonical_sku, 0))
            # [2026-05-27] SKU 표시 룰: canonical_sku 가 옛 sku 형식이면 boxhero_sku 사용
            sku_display = o.canonical_sku
            if sku_display and not sku_display.startswith('SKU-') and o.boxhero_sku:
                sku_display = o.boxhero_sku
            row = [sku_display, barcode, article, brand, category, mname,
                   color, size, avg, total]
            for loc in locs:
                row.append(int(per_loc_stock.get(o.canonical_sku, {}).get(loc.id, 0)))
            ws.append(row)

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4F67FF', end_color='4F67FF', fill_type='solid')
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'

        # [D-6 v2] 컬럼 너비 — 10 base (SKU/바코드/품번/브랜드/카테고리/모델명/색상/사이즈/평균매입가/총재고) + N 위치
        widths = [18, 16, 16, 12, 18, 28, 18, 10, 12, 8] + [12] * len(locs)
        for i, w in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        ts = datetime.now().strftime('%Y%m%d-%H%M%S')
        from flask import send_file
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'재고관리_{"필터_" if is_filtered else ""}{ts}.xlsx',
        )
    finally:
        s.close()
