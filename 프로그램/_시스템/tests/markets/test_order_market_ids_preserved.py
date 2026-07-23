# -*- coding: utf-8 -*-
"""주문 파서가 마켓 상품·옵션 식별자를 행에 보존하는가 (M4 가격 전후 표시의 열쇠).

M4(주문 시점 전후 가격)는 주문 행을 우리 옵션(SKU)에 이어야 값을 낸다. 그 연결고리가
마켓 식별자다 — 없으면 전 행이 '확인 불가'가 된다.

이 테스트가 지키는 것:
  ① 마켓별로 응답에 있는 식별자가 `_pd_` 키로 행에 실린다
  ② 응답에 그 필드가 없으면 **조용히 None 이 아니라 빈값** → 하류가 '확인 불가'로 간다
  ③ 기존 열(상품명·옵션·단가·정산 등)이 한 글자도 안 바뀐다(회귀)
  ④ 엑셀 내보내기에 `_pd_` 가 새어나가지 않는다

전부 대역(fixture)·모킹 — 라이브 마켓에 요청하지 않는다.
"""
import datetime as dt

import pytest

from lemouton.markets import order_export as oe

KST = dt.timezone(dt.timedelta(hours=9))
SINCE = dt.datetime(2026, 7, 5, tzinfo=KST)
UNTIL = dt.datetime(2026, 7, 5, 23, tzinfo=KST)

_E11_ITERS = ("iter_orders", "iter_delivered", "iter_completed", "iter_preparing",
              "iter_shipping", "iter_cancel", "iter_canceled", "iter_return",
              "iter_exchange")


# ─────────────────────────────────────────────────────────────────────────────
#  스마트스토어 — productId(채널 상품 번호) · originalProductId(원상품 번호)
# ─────────────────────────────────────────────────────────────────────────────

class _SSClient:
    """SmartStoreClient 대역. productOrder 에 담을 필드를 테스트가 지정한다."""

    def __init__(self, po_extra):
        self._po_extra = po_extra

    def request(self, method, path, query="", body=None):
        if "last-changed-statuses" in path:
            return {"data": {"lastChangeStatuses": [{"productOrderId": "P1"}]}}
        if path.endswith("/product-orders/query"):
            po = {"productOrderId": "P1", "productName": "코트",
                  "productOption": "블랙/95", "quantity": 1, "unitPrice": 189000,
                  "shippingAddress": {"name": "수령자A", "tel1": "01011112222",
                                      "zipCode": "13105", "baseAddress": "서울",
                                      "detailedAddress": "101동"}}
            po.update(self._po_extra)
            return {"data": [{"order": {"orderDate": "2026-07-05T09:00:00",
                                        "ordererName": "구매자A",
                                        "ordererTel": "01000000000"},
                              "productOrder": po}]}
        if "pay-settle/settle/case" in path:
            return {"elements": [{"productOrderId": "P1",
                                  "settleExpectAmount": 169155}],
                    "pagination": {"totalPages": 1}}
        return {"data": {}}


def _ss_rows(po_extra):
    return oe.smartstore_order_rows(SINCE, UNTIL, client=_SSClient(po_extra))


def test_smartstore_keeps_both_product_ids():
    """채널 상품 번호·원상품 번호를 둘 다 보존한다.

    판매처 연동은 둘 중 무엇으로도 등록될 수 있어(market_fetch._fetch_smartstore 가
    채널·원상품 둘 다 받는다) 한쪽만 담으면 연결이 통째 실패한다.
    """
    r = _ss_rows({"productId": "77", "originalProductId": "99"})[0]
    assert r["_pd_market_product_id"] == "77"
    assert r["_pd_market_product_id_alt"] == "99"


def test_smartstore_missing_ids_stay_empty_not_wrong():
    """응답에 상품번호가 없으면 빈값 — 없는 값을 지어내지 않는다."""
    r = _ss_rows({})[0]
    assert r["_pd_market_product_id"] == ""
    assert r["_pd_market_product_id_alt"] == ""


def test_smartstore_does_not_guess_option_id():
    """옵션 단위 id 는 담지 않는다 — optionCode·optionId 가 우리 옵션 식별자와
    같다는 근거가 없어(문서상 뜻 불명·deprecated 예정) 이으면 엉뚱한 옵션 가격이 뜬다."""
    r = _ss_rows({"productId": "77", "optionCode": "OPT-XYZ",
                  "optionManageCode": "MC-1", "optionId": "5"})[0]
    assert "_pd_market_option_id" not in r


def test_smartstore_existing_columns_unchanged():
    """회귀 — 기존 매핑·정산 조인 결과가 그대로다."""
    r = _ss_rows({"productId": "77"})[0]
    assert r["상품명"] == "코트" and r["옵션"] == "블랙/95"
    assert r["수령자"] == "수령자A" and r["구매자"] == "구매자A"
    assert r["단가"] == 189000
    assert r["정산예정금액"] == 169155        # 정산 조인 유지
    assert r["_settle_source"] == "real"
    assert r["판매처"] == "스마트스토어"
    assert r["오픈마켓주문번호"] == "P1"


# ─────────────────────────────────────────────────────────────────────────────
#  11번가 — prdStckNo(주문상품옵션코드) · prdNo(11번가상품번호)
# ─────────────────────────────────────────────────────────────────────────────

def _e11_rows(monkeypatch, *, active=(), canceled=()):
    import shared.platforms.eleven11.orders as e11
    for nm in _E11_ITERS:
        monkeypatch.setattr(e11, nm, lambda *a, **k: iter([]))
    monkeypatch.setattr(e11, "iter_orders", lambda *a, **k: iter(list(active)))
    monkeypatch.setattr(e11, "iter_canceled", lambda *a, **k: iter(list(canceled)))
    return oe.eleven11_order_rows(SINCE, UNTIL, client=object(),
                                  include_settlement=False)


def test_eleven11_keeps_option_and_product_id(monkeypatch):
    """prdStckNo = 주문상품옵션코드 = 판매처 연동의 옵션 식별자(재고번호)."""
    od = {"ordNo": "E1", "ordPrdSeq": "1", "prdNm": "코트",
          "slctPrdOptNm": "블랙/95", "ordQty": "1", "selPrc": "189000",
          "prdNo": "PRD-9", "prdStckNo": "STK-77", "ordDt": "2026-07-05 09:00:00"}
    r = _e11_rows(monkeypatch, active=[od])[0]
    assert r["_pd_market_option_id"] == "STK-77"
    assert r["_pd_market_product_id"] == "PRD-9"


def test_eleven11_without_option_code_is_empty(monkeypatch):
    """목록 응답이 prdStckNo 를 안 주면 빈값 — 하류에서 '확인 불가'로 간다."""
    od = {"ordNo": "E2", "ordPrdSeq": "1", "prdNm": "코트",
          "slctPrdOptNm": "블랙/95", "prdNo": "PRD-9",
          "ordDt": "2026-07-05 09:00:00"}
    r = _e11_rows(monkeypatch, active=[od])[0]
    assert r["_pd_market_option_id"] == ""
    assert r["_pd_market_product_id"] == "PRD-9"


def test_eleven11_claim_row_is_product_level_only(monkeypatch):
    """클레임 목록은 문서상 옵션코드를 주지 않는다 → 상품 단위만 보존."""
    cl = {"ordNo": "E3", "ordPrdSeq": "1", "slctPrdOptNm": "블랙/95",
          "prdNo": "PRD-9", "ordPrdStat": "701"}
    r = [x for x in _e11_rows(monkeypatch, canceled=[cl])
         if x["오픈마켓주문번호"] == "E3"][0]
    assert r["_pd_market_product_id"] == "PRD-9"
    assert "_pd_market_option_id" not in r
    assert r["_kind"] == "change"          # 회귀 — 클레임 태그 유지
    assert r["주문일"] == ""                # 회귀 — 합성 주문일 금지


def test_eleven11_existing_columns_unchanged(monkeypatch):
    """회귀 — 기존 매핑(단가·배송비·송장·정산예정)이 그대로다."""
    od = {"ordNo": "E9", "ordPrdSeq": "1", "prdNm": "코트",
          "slctPrdOptNm": "블랙/95", "ordQty": "2", "selPrc": "189000",
          "dlvCst": "3000", "invcNo": "123456", "stlPlnAmt": "169155",
          "prdNo": "PRD-9", "prdStckNo": "STK-77", "ordDt": "2026-07-05 09:00:00"}
    r = _e11_rows(monkeypatch, active=[od])[0]
    assert r["상품명"] == "코트" and r["옵션"] == "블랙/95"
    assert r["수량"] == "2" and r["단가"] == "189000"
    assert r["배송비"] == "3000" and r["송장입력"] == "123456"
    assert r["정산예정금액"] == 169155 - 3000 and r["_settle_source"] == "real"   # M열=stlPlnAmt−배송비(2026-07-23 샵마인 규약)
    assert r["주문일"] == "2026-07-05 09:00:00"


# ─────────────────────────────────────────────────────────────────────────────
#  옥션·G마켓(ESM) — SiteGoodsNo(사이트 상품번호)
# ─────────────────────────────────────────────────────────────────────────────

def _esm_rows(monkeypatch, market, od):
    import shared.platforms.esm.orders as esm
    import shared.platforms.esm.settlements as st
    monkeypatch.setattr(esm, "iter_orders", lambda *a, **k: iter([od]))
    monkeypatch.setattr(st, "settle_price_map", lambda *a, **k: {})
    return oe.esm_order_rows(market, SINCE, UNTIL, client=None)


@pytest.mark.parametrize("market,label", [("auction", "옥션"), ("gmarket", "G마켓")])
def test_esm_keeps_site_goods_no(monkeypatch, market, label):
    od = {"OrderNo": "A1", "GoodsName": "코트", "SalePrice": "189000",
          "ContrAmount": 1, "SiteGoodsNo": "SG-55", "GoodsNo": None,
          "OrderDate": "2026-07-05 09:00:00",
          "ItemOptionSelectList": [{"ItemOptionValue": "색상:블랙",
                                    "ItemOptionCode": "IOC-1"}]}
    r = _esm_rows(monkeypatch, market, od)[0]
    assert r["_pd_market_product_id"] == "SG-55"
    # 옵션 코드는 우리 옵션 식별자(판매자옵션코드)와 같다는 근거가 없어 담지 않는다.
    assert "_pd_market_option_id" not in r
    # 회귀 — 기존 열 그대로
    assert r["판매처"] == label and r["상품명"] == "코트"
    assert r["단가"] == "189000" and r["옵션"] == "색상:블랙 IOC-1"


def test_esm_missing_site_goods_no_is_empty(monkeypatch):
    od = {"OrderNo": "A2", "GoodsName": "코트", "OrderDate": "2026-07-05 09:00:00"}
    r = _esm_rows(monkeypatch, "auction", od)[0]
    assert r["_pd_market_product_id"] == ""


# ─────────────────────────────────────────────────────────────────────────────
#  누출 방지 — 엑셀·열 목록
# ─────────────────────────────────────────────────────────────────────────────

def test_internal_ids_never_reach_excel():
    """`_pd_` 는 내부 표시 — 엑셀 열 화이트리스트(ALL_COLUMNS)에 없어야 하고,
    행에 들어있어도 파일에 찍히면 안 된다."""
    import io as _io
    import openpyxl

    assert not [c for c in oe.ALL_COLUMNS if c.startswith("_")]
    rows = [{"주문일": "2026-07-05 09:00:00", "판매처": "스마트스토어",
             "상품명": "코트", "단가": 189000,
             "_pd_market_product_id": "77", "_pd_market_product_id_alt": "99",
             "_pd_market_option_id": "STK-77"}]
    wb = openpyxl.load_workbook(_io.BytesIO(oe.rows_to_xlsx(rows)))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert not [h for h in header if h and str(h).startswith("_")]
    body = [str(c.value) for c in ws[2]]
    for leaked in ("77", "99", "STK-77"):
        assert leaked not in body
