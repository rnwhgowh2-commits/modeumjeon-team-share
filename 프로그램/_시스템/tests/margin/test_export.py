# -*- coding: utf-8 -*-
"""export.to_xlsx — 탭별 시트."""
import io

import openpyxl

from lemouton.margin import export as E


def _payload():
    return {
        "matched": [{"마켓주문번호": "1001", "마켓": "쿠팡", "상품명": "코트",
                     "순마진": 20000, "마진율": 25.0}],
        "unmatched_buy": [{"마켓주문번호": "2002", "마켓명": "쿠팡"}],
        "unmatched_sell": [{"마켓주문번호": "3003", "쇼핑몰": "06.쿠팡"}],
        "summary": {"총매출": 80000, "총순마진": 20000},
        "market": [{"마켓": "쿠팡", "매출": 80000, "순마진": 20000}],
        "daily": [{"일자": "2026-07-04", "매출": 80000, "순마진": 20000}],
        "monthly": [{"월": "2026-07", "매출": 80000, "순마진": 20000}],
        "brand": [{"브랜드": "나이키", "매출": 80000, "순마진": 20000}],
        "priceRange": [{"금액대": "5~10만", "매출": 80000, "순마진": 20000}],
        "product": [{"상품코드": "12345", "상품명": "코트", "매출": 80000}],
    }


def _sheets(data: bytes):
    return openpyxl.load_workbook(io.BytesIO(data)).sheetnames


def test_tab_all_has_every_sheet():
    names = _sheets(E.to_xlsx(_payload(), tab="all"))
    for s in ("전체매칭", "마켓X_매입O", "마켓O_매입X", "요약",
              "마켓별", "일별", "월별", "브랜드별", "금액대별"):
        assert s in names, s


def test_tab_daily_only():
    names = _sheets(E.to_xlsx(_payload(), tab="daily"))
    assert names == ["일별"]


def test_detail_filtered_respects_rows_and_column_order():
    rows = [{"마켓": "쿠팡", "순마진": 20000, "상품명": "코트"}]
    data = E.to_xlsx(_payload(), tab="detail_filtered",
                     rows=rows, column_order=["상품명", "마켓", "순마진"])
    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    assert [c.value for c in ws[1]] == ["상품명", "마켓", "순마진"]
    assert [c.value for c in ws[2]] == ["코트", "쿠팡", 20000]


def test_empty_payload_still_produces_workbook():
    data = E.to_xlsx({"matched": []}, tab="all")
    assert _sheets(data)   # 빈 워크북은 openpyxl 이 못 읽는다


def test_real_payload_roundtrips():
    """실데이터 166행이 실제로 엑셀로 나가는지 (openpyxl 시트명 31자 제한 등).

    2026-09: sell_source 의 구 통합주문관리 엑셀 변환 함수가 삭제됐다(전 마켓 API
    연동 완료) — 이 재현 경로는 원본 데이터가 있는 그 PC 에서도 더 이상 동작하지
    않는다(조용한 실패 금지 — 명시적으로 실패한다).
    """
    import pathlib

    import pytest

    d = pathlib.Path(r"C:\dev\대량등록 마진계산기\데이터\260704")
    if not d.exists():
        pytest.skip("원본 엑셀 없음")
    pytest.skip(
        "구 통합주문관리 엑셀 변환 함수가 삭제돼(2026-09) 이 재현 경로는 더 이상 동작하지 "
        "않습니다 — 전 마켓 API 연동 완료로 그 입력 포맷 자체가 은퇴함."
    )
