# -*- coding: utf-8 -*-
"""엑셀 내보내기 — 「주문 관리」 상태를 **맨 앞 열**로 얹는다(사장님 확정 2026-08-06).

🔴 지켜야 할 것: **기존 열 순서·이름이 하나도 안 바뀐다.** 엑셀을 쓰는 다른 흐름
   (양식 설정 UI · 프리셋 · GET 레거시 내보내기)이 `ALL_COLUMNS` 를 그대로 본다.
   그래서 `ALL_COLUMNS` 에는 넣지 않고 `rows_to_xlsx(lead_columns=...)` 로 앞에만 붙인다.
"""
import io
import pathlib

import openpyxl
import pytest
from flask import Flask

from webapp.routes import orders as om
from lemouton.markets import order_export as oe


def _client():
    app = Flask(__name__, template_folder="webapp/templates",
                root_path=pathlib.Path(om.__file__).parents[2].as_posix())
    app.register_blueprint(om.bp)
    return app.test_client()


def _sheet(data):
    return openpyxl.load_workbook(io.BytesIO(data)).active


ROWS = [
    {"수령자": "홍길동", "주소": "서울 강남구", "상품명": "신발A", "주문 관리": "배송완료"},
    {"수령자": "김철수", "주소": "서울 송파구", "상품명": "신발B", "주문 관리": "결제완료 (기본)"},
    {"수령자": "박영희", "주소": "서울 마포구", "상품명": "신발C", "주문 관리": ""},
]
COLS = ["수령자", "주소", "상품명"]


def test_기존_열_목록은_그대로다():
    """🔴 ALL_COLUMNS 를 건드리면 양식 UI·프리셋·레거시 내보내기가 통째로 밀린다."""
    assert "주문 관리" not in oe.ALL_COLUMNS
    assert "주문 관리" not in oe.DEFAULT_COLUMNS
    assert oe.resolve_columns(["주문 관리", "수령자"]) == ["수령자"]   # 화이트리스트 유지


def test_lead_columns_는_맨_앞에만_붙는다():
    ws = _sheet(oe.rows_to_xlsx(ROWS, columns=COLS, lead_columns=["주문 관리"]))
    assert [c.value for c in ws[1]] == ["주문 관리"] + COLS
    assert ws.cell(row=2, column=1).value == "배송완료"
    assert ws.cell(row=3, column=1).value == "결제완료 (기본)"
    # 안 고른 줄 = 빈칸(지어내지 않음). openpyxl 은 빈 문자열을 None 으로 되읽는다.
    assert ws.cell(row=4, column=1).value in (None, "")
    # 기존 열은 자리·값 그대로 한 칸씩 뒤로만 밀린다
    assert ws.cell(row=2, column=2).value == "홍길동"
    assert ws.cell(row=2, column=4).value == "신발A"


def test_lead_columns_없으면_예전과_완전히_같다():
    """옛 호출부(마진계산기·레거시 GET)는 인자를 안 준다 — 결과가 달라지면 안 된다."""
    before = _sheet(oe.rows_to_xlsx(ROWS, columns=COLS))
    assert [c.value for c in before[1]] == COLS
    assert before.max_column == len(COLS)


def test_중복_열을_만들지_않는다():
    ws = _sheet(oe.rows_to_xlsx(ROWS, columns=["수령자"], lead_columns=["수령자"]))
    assert [c.value for c in ws[1]] == ["수령자"]


def test_route_가_화이트리스트_밖의_열은_무시한다():
    """클라이언트가 아무 이름이나 보내도 열이 되지 않는다(엑셀 쓰는 다른 흐름 보호)."""
    r = _client().post("/orders/export.xlsx",
                       json={"rows": ROWS, "cols": COLS,
                             "lead_cols": ["주문 관리", "몰래끼운열", "수령자전화번호"]})
    assert r.status_code == 200
    ws = _sheet(r.data)
    assert [c.value for c in ws[1]] == ["주문 관리"] + COLS


def test_route_가_lead_cols_없으면_열을_안_붙인다():
    r = _client().post("/orders/export.xlsx", json={"rows": ROWS, "cols": COLS})
    assert r.status_code == 200
    assert [c.value for c in _sheet(r.data)[1]] == COLS


@pytest.mark.parametrize("bad", [None, "주문 관리", 5, {}])
def test_route_가_이상한_lead_cols_에_안_터진다(bad):
    r = _client().post("/orders/export.xlsx",
                       json={"rows": ROWS, "cols": COLS, "lead_cols": bad})
    assert r.status_code == 200
