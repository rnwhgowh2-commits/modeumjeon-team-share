"""테스트용 우리 양식 xlsx 생성 — 12 컬럼 (9 base + 3 위치) × 10 SKU.

사용:
  cd 프로그램/_시스템
  python scripts/_make_test_xlsx.py

출력: 바탕화면 의 '테스트_재고관리_양식.xlsx'

데이터 패턴:
  - 2 브랜드 × 2 모델 × 색상/사이즈 다양
  - LCP 가 의도된 모델명을 정확히 도출하도록 같은 model_code 안에 색상 prefix 공유
  - 위치별 재고는 의미있는 분포
"""
from __future__ import annotations

import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = Path(r"C:\Users\seung\OneDrive\바탕 화면\테스트_재고관리_양식.xlsx")

HEADERS = [
    'SKU', '바코드', '브랜드', '제품명', '품번',
    '색상', '사이즈', '평균매입가', '총재고',
    '기본 위치 재고', '그로스 재고', '판매불가 재고',
]

ROWS = [
    # 아디다스 슈퍼스타 — 3 색상, 다양한 사이즈
    ('TEST-AD-001', '8800010001', '아디다스', '아디다스 슈퍼스타', 'SS-FW', '화이트', '270', 50000, 6, 3, 2, 1),
    ('TEST-AD-002', '8800010002', '아디다스', '아디다스 슈퍼스타', 'SS-FW', '화이트', '275', 50000, 4, 2, 2, 0),
    ('TEST-AD-003', '8800010003', '아디다스', '아디다스 슈퍼스타', 'SS-FW', '블랙', '270', 50000, 3, 2, 1, 0),
    ('TEST-AD-004', '8800010004', '아디다스', '아디다스 슈퍼스타', 'SS-FW', '블랙', '275', 50000, 5, 3, 2, 0),
    ('TEST-AD-005', '8800010005', '아디다스', '아디다스 슈퍼스타', 'SS-FW', '레드', '280', 50000, 2, 1, 1, 0),

    # 컨버스 척테일러 — 2 색상
    ('TEST-CV-001', '8800020001', '컨버스', '컨버스 척테일러', 'CT-LO', '화이트', '250', 35000, 4, 2, 2, 0),
    ('TEST-CV-002', '8800020002', '컨버스', '컨버스 척테일러', 'CT-LO', '블랙', '260', 35000, 3, 1, 2, 0),
    ('TEST-CV-003', '8800020003', '컨버스', '컨버스 척테일러', 'CT-LO', '블랙', '265', 35000, 5, 3, 2, 0),

    # 나이키 코르테즈 — 단일 색상 옵션 (LCP 불가 case)
    ('TEST-NK-001', '8800030001', '나이키', '나이키 코르테즈', 'CTZ-CL', '클래식', '275', 80000, 2, 1, 1, 0),
    ('TEST-NK-002', '8800030002', '나이키', '나이키 코르테즈', 'CTZ-CL', '클래식', '280', 80000, 3, 1, 2, 0),
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '재고관리'
    ws.append(HEADERS)
    for row in ROWS:
        ws.append(row)

    # 헤더 스타일 (파란 배경 + 흰 글씨)
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F67FF', end_color='4F67FF', fill_type='solid')
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    widths = [16, 16, 14, 30, 12, 12, 10, 14, 10, 14, 12, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f'✓ saved: {OUT}')
    print(f'  rows: {len(ROWS)} + 1 header')
    print(f'  cols: {len(HEADERS)}')
    print(f'  brands: 아디다스(5) / 컨버스(3) / 나이키(2)')


if __name__ == "__main__":
    main()
