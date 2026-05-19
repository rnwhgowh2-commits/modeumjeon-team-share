"""엑셀 양식 round-trip 시연.

흐름:
  1. 임시 SQLite (DB1) 에 박스히어로 xlsx import → 옵션·재고 채움
  2. DB1 의 데이터로 우리 양식 xlsx export (8 base + 동적 위치별)
  3. 새 임시 SQLite (DB2) 에 우리 양식 xlsx import → 같은 결과 재현 확인
  4. DB1 ↔ DB2 통계 비교

사용:
  cd 프로그램/_시스템
  python scripts/_demo_xlsx_roundtrip.py "<source xlsx>"

(인자 생략 시 사용자의 Items_Export 파일 사용)
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent  # _시스템
sys.path.insert(0, str(ROOT))

DEFAULT_SRC = r"C:\Users\seung\OneDrive\바탕 화면\Items_Export_99LAB_2026-05-17T23-11-20.xlsx"
SRC_XLSX = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC


def setup_db():
    """임시 SQLite + 모든 모델 등록 + init_db."""
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp.close()
    os.environ['DATABASE_URL'] = f"sqlite:///{Path(tmp.name).as_posix()}"
    # 캐시된 engine 무력화를 위해 shared.db 다시 import
    import importlib
    import shared.db
    importlib.reload(shared.db)
    from shared.db import init_db, SessionLocal
    for _mod in [
        "lemouton.sourcing.models", "lemouton.sourcing.models_pricing",
        "lemouton.pricing.settings", "lemouton.uploader.models",
        "lemouton.templates.models", "lemouton.inventory.models",
        "lemouton.sources.models", "lemouton.sourcing.models_v2",
        "lemouton.multitenancy.models", "lemouton.audit.models",
    ]:
        try:
            __import__(_mod)
        except ImportError:
            pass
    try:
        import webapp.auth.models  # noqa: F401
    except ImportError:
        pass
    init_db()
    return tmp.name, SessionLocal


def import_and_stats(xlsx, label):
    """xlsx 1회 import 후 통계 반환."""
    db_path, SL = setup_db()
    from lemouton.inventory.boxhero_import import import_xlsx, verify_after_import
    from lemouton.sourcing.boxhero_xlsx import detect_format
    fmt = detect_format(xlsx)
    print(f"\n=== {label} (xlsx: {Path(xlsx).name})")
    print(f"  format 감지: {fmt}")
    s = SL()
    try:
        result = import_xlsx(xlsx, s)
        s.commit()
        v = verify_after_import(s)
        print(f"  records: {result['records_count']}")
        print(f"  auto_created_options: {result.get('auto_created_options', 0)}")
        print(f"  stock_updated: {result['stock_updated']}")
        print(f"  with_stock: {v['with_stock']}")
        print(f"  total_stock: {v['total_stock']}")
        return db_path, SL, v
    finally:
        s.close()


def export_internal(SL, out_path):
    """SL 의 데이터로 우리 양식 xlsx export."""
    from io import BytesIO
    import openpyxl
    from sqlalchemy.orm import joinedload
    from lemouton.sourcing.models import Option
    from lemouton.inventory.models import InventoryLocation
    from shared.inventory_stock import get_stock_batch

    s = SL()
    try:
        options = (
            s.query(Option)
            .options(joinedload(Option.model))
            .order_by(Option.model_code, Option.sort_order, Option.canonical_sku)
            .all()
        )
        all_skus = [o.canonical_sku for o in options]
        total_stock_map = get_stock_batch(s, all_skus)
        locs = (
            s.query(InventoryLocation)
            .filter(InventoryLocation.deleted_at.is_(None))
            .order_by(InventoryLocation.sort_order, InventoryLocation.id)
            .all()
        )
        per_loc_stock = {}
        for loc in locs:
            for sku, st in get_stock_batch(s, all_skus, location_id=loc.id).items():
                per_loc_stock.setdefault(sku, {})[loc.id] = st

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '재고관리'
        headers = ['SKU', '바코드', '브랜드', '제품명', '품번', '색상', '사이즈', '평균매입가', '총재고']
        for loc in locs:
            headers.append(f'{loc.name} 재고')
        ws.append(headers)

        for o in options:
            barcode = o.barcode or o.boxhero_sku or ''
            brand = (o.model.brand or '') if o.model else ''
            pname = ((o.model.model_name_display or o.model.model_name_raw) if o.model else o.canonical_sku) or ''
            article = (getattr(o.model, 'article_no', None) or '') if o.model else ''
            color = (o.color_display or o.color_code or 'one')
            size = (o.size_display or o.size_code or 'free')
            if color == pname or (len(color) > 12 and pname.startswith(color[:8])):
                color = 'one'
            avg = int(o.boxhero_avg_purchase_price or 0)
            total = int(total_stock_map.get(o.canonical_sku, 0))
            row = [o.canonical_sku, barcode, brand, pname, article, color, size, avg, total]
            for loc in locs:
                row.append(int(per_loc_stock.get(o.canonical_sku, {}).get(loc.id, 0)))
            ws.append(row)

        wb.save(out_path)
        print(f"  → {out_path} 저장 (rows={len(options)+1}, cols={len(headers)})")
    finally:
        s.close()


def inspect_xlsx(xlsx_path):
    """xlsx 의 헤더 + 1행 보기."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    print(f"  sheet: {ws.title}, rows: {len(rows)}")
    print(f"  헤더: {list(rows[0])}")
    if len(rows) > 1:
        print(f"  row1: {list(rows[1])[:8]}...")
    wb.close()


def main():
    if not Path(SRC_XLSX).exists():
        print(f"❌ source xlsx 없음: {SRC_XLSX}")
        return 2

    # Step 1+2: 박스히어로 xlsx → DB1 import → 우리 양식 xlsx export
    db1, SL1, v1 = import_and_stats(SRC_XLSX, "Step 1) 박스히어로 xlsx → DB1 import")
    out_xlsx = tempfile.NamedTemporaryFile(suffix='_재고관리_export.xlsx', delete=False)
    out_xlsx.close()
    print(f"\n=== Step 2) DB1 → 우리 양식 xlsx export")
    export_internal(SL1, out_xlsx.name)
    print(f"  검증:")
    inspect_xlsx(out_xlsx.name)

    # Step 3: 우리 양식 xlsx → DB2 import
    db2, SL2, v2 = import_and_stats(out_xlsx.name, "Step 3) 우리 양식 xlsx → DB2 import")

    # Step 4: DB1 ↔ DB2 비교
    print(f"\n=== Step 4) DB1 ↔ DB2 통계 비교")
    print(f"  {'항목':<24} {'DB1 (박스히어로)':>18} {'DB2 (우리 양식)':>18}")
    for k in ('mapped_count', 'with_stock', 'total_stock'):
        a, b = v1.get(k, 0), v2.get(k, 0)
        match = "✓" if a == b else "✗"
        print(f"  {k:<24} {a:>18} {b:>18}  {match}")

    print(f"\nDB1: {db1}\nDB2: {db2}\nexport xlsx: {out_xlsx.name}")
    return 0 if (v1['total_stock'] == v2['total_stock'] and v1['with_stock'] == v2['with_stock']) else 1


if __name__ == "__main__":
    sys.exit(main())
