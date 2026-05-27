"""재고관리 엑셀 (박스히어로 810 + 스카이블루 신규).

행 구성:
- 박스히어로 엑셀의 모든 SKU 810건 (우리 DB 메타와 조인)
- 우리 DB 의 르무통_메이트 스카이블루 옵션 8건 (박스히어로엔 없음)
- 다른 비활성 옵션 (르무통_버디 step3, 아이보리 등) 제외

컬럼: SKU, 바코드, 품번, 브랜드, 카테고리, 모델명, 색상, 사이즈,
      평균매입가, 총재고, 기본 위치, 그로스, TEST
"""
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

BX_PATH = Path(r'C:\Users\seung\Downloads\Items_Export_99LAB_2026-05-27T21-41-37.xlsx')


def main():
    # 박스히어로
    bx_wb = openpyxl.load_workbook(BX_PATH, data_only=True)
    bx_ws = bx_wb['BoxHero']
    bx_hdr = [c.value for c in bx_ws[1]]
    bx_rows = [dict(zip(bx_hdr, r)) for r in bx_ws.iter_rows(min_row=2, values_only=True)]
    bx_by_sku = {r['SKU']: r for r in bx_rows if r.get('SKU')}
    print(f'박스히어로 SKU: {len(bx_rows)}')

    from sqlalchemy.orm import joinedload
    from sqlalchemy import text
    from shared.db import SessionLocal
    from lemouton.sourcing.models import Option
    from lemouton.inventory.models import InventoryLocation
    from shared.inventory_stock import get_stock_batch

    s = SessionLocal()
    try:
        # 1. 박스히어로 SKU 매칭되는 DB 옵션 (boxhero_sku 기준)
        bx_skus = list(bx_by_sku.keys())
        bx_matched_opts = (s.query(Option).options(joinedload(Option.model))
                           .filter(Option.boxhero_sku.in_(bx_skus)).all())
        bx_matched_by_sku = {}
        for o in bx_matched_opts:
            # 활성 우선, 옛 sku canonical 우선
            existing = bx_matched_by_sku.get(o.boxhero_sku)
            if existing is None:
                bx_matched_by_sku[o.boxhero_sku] = o
                continue
            score_o = (1 if o.is_active else 0, 1 if not o.canonical_sku.startswith('SKU-') else 0)
            score_e = (1 if existing.is_active else 0, 1 if not existing.canonical_sku.startswith('SKU-') else 0)
            if score_o > score_e:
                bx_matched_by_sku[o.boxhero_sku] = o
        print(f'박스히어로 SKU ↔ DB 매칭: {len(bx_matched_by_sku)}')

        # 2. 스카이블루 옵션 (르무통_메이트)
        sky_opts = (s.query(Option).options(joinedload(Option.model))
                    .filter(Option.model_code == '르무통_메이트',
                            Option.color_code == '스카이블루').all())
        print(f'스카이블루 옵션: {len(sky_opts)}')

        # 모든 옵션 모음 (재고 batch 위해)
        all_opts = list(bx_matched_by_sku.values()) + sky_opts
        all_skus = [o.canonical_sku for o in all_opts]

        total_stock = get_stock_batch(s, all_skus)
        locs = (s.query(InventoryLocation)
                .filter(InventoryLocation.deleted_at.is_(None))
                .order_by(InventoryLocation.sort_order, InventoryLocation.id).all())
        per_loc = {}
        for loc in locs:
            loc_map = get_stock_batch(s, all_skus, location_id=loc.id)
            for sku, st in loc_map.items():
                per_loc.setdefault(sku, {})[loc.id] = st

        # Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '재고관리'
        headers = ['SKU', '바코드', '품번', '브랜드', '카테고리', '모델명',
                   '색상', '사이즈', '평균매입가', '총재고']
        for loc in locs:
            headers.append(loc.name)
        ws.append(headers)

        def make_row(o, sku_for_display, bx=None):
            m = o.model if o else None
            barcode = (o.barcode if o else '') or (str(bx.get('바코드')) if bx else '') or ''
            article = (getattr(m, 'article_no', None) or '-') if m else '-'
            brand = (m.brand or '') if m else (bx.get('브랜드') or '') if bx else ''
            category = (m.category or '') if m else (bx.get('카테고리') or '') if bx else ''
            mname = ((getattr(m, 'model_name_display', None) or
                      getattr(m, 'model_name_raw', None)) if m else '') or ''
            color = (o.color_display or o.color_code or '-') if o else '-'
            size = (o.size_display or o.size_code or '-') if o else (str(bx.get('사이즈')) if bx and bx.get('사이즈') else '-')
            avg = int(o.boxhero_avg_purchase_price or 0) if o else 0
            total = int(total_stock.get(o.canonical_sku, 0)) if o else 0
            row = [sku_for_display, barcode, article, brand, category, mname,
                   color, size, avg, total]
            for loc in locs:
                row.append(int(per_loc.get(o.canonical_sku if o else None, {}).get(loc.id, 0)))
            return row

        # 정렬: 브랜드 > 카테고리 > 모델명 > 색상 > 사이즈 (사용자 룰)
        def sort_key(item):
            sku, o, bx = item
            m = o.model if o else None
            brand = (m.brand if m else None) or (bx.get('브랜드') if bx else '') or ''
            cat = (m.category if m else None) or (bx.get('카테고리') if bx else '') or ''
            mname = ((m.model_name_display or m.model_name_raw) if m else '') or ''
            color = (o.color_display or o.color_code) if o else ''
            color = color or ''
            try:
                size_n = int(o.size_display or o.size_code) if o else 0
                size_key = (0, size_n)  # 숫자 사이즈 우선
            except (ValueError, TypeError):
                size_key = (1, (o.size_display or o.size_code or '') if o else '')
            return (brand, cat, mname, color, size_key)

        # 모든 항목 모음
        all_items = []
        for bx_sku in bx_by_sku.keys():
            bx = bx_by_sku[bx_sku]
            o = bx_matched_by_sku.get(bx_sku)
            all_items.append((bx_sku, o, bx))
        for o in sky_opts:
            all_items.append((o.boxhero_sku or o.canonical_sku, o, None))

        all_items.sort(key=sort_key)

        for sku, o, bx in all_items:
            ws.append(make_row(o, sku, bx))

        # 헤더 스타일
        hd_font = Font(bold=True, color='FFFFFF', size=11)
        hd_fill = PatternFill('solid', fgColor='4F67FF')
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = hd_font; c.fill = hd_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        widths = [18, 16, 16, 12, 18, 28, 18, 10, 12, 8] + [12] * len(locs)
        for i, w in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

        ts = datetime.now().strftime('%Y%m%d-%H%M%S')
        out = Path(r'C:\Users\seung\Downloads') / f'재고관리_박스히어로810_스카이블루_{ts}.xlsx'
        wb.save(out)
        print(f'\n저장: {out}')
        print(f'전체 행: {len(bx_by_sku) + len(sky_opts)}')
    finally:
        s.close()


if __name__ == '__main__':
    main()
