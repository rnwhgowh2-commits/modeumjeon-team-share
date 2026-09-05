# -*- coding: utf-8 -*-
"""르무통 메이트 소싱처별 크롤 데이터(최신 스냅샷) → 비교 엑셀."""
from __future__ import annotations
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
from pathlib import Path
import os
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
import config  # noqa
from shared.db import SessionLocal
from lemouton.sourcing.models import Model, Option
import lemouton.sources.models  # noqa
from lemouton.templates.models import PriceTrackHistory

SRC_ORDER = [("lemouton", "르무통"), ("musinsa", "무신사"), ("ssf", "SSF"), ("lotte", "롯데온")]
COLOR_ORDER = ['그레이','다크네이비','라이트블루','블랙','스카이블루','아이보리','오렌지','올리브그린','크림핑크']

mate = '르무통_메이트'
s = SessionLocal()
opts = {o.canonical_sku: o for o in s.query(Option).filter_by(model_code=mate).all()}
skus = set(opts)
# 최신 스냅샷만
latest = {}
for r in s.query(PriceTrackHistory).all():
    if r.canonical_sku not in skus:
        continue
    k = (r.canonical_sku, r.source)
    if k not in latest or (r.captured_at or 0) > (latest[k].captured_at or 0):
        latest[k] = r
s.close()

# 옵션(색상,사이즈)별 정리
def ckey(c): return COLOR_ORDER.index(c) if c in COLOR_ORDER else 99
def skey(z):
    d = ''.join(ch for ch in (z or '') if ch.isdigit())
    return int(d) if d else 0

rows = []
for sku, o in opts.items():
    color, size = o.color_code, o.size_code
    cell = {"색상": color, "사이즈": size}
    prices_avail = []
    for key, label in SRC_ORDER:
        r = latest.get((sku, key))
        if r:
            cell[f"{label}_가격"] = r.price
            cell[f"{label}_재고"] = r.stock
            if r.price and (r.stock is None or r.stock != 0):
                prices_avail.append((r.price, label))
        else:
            cell[f"{label}_가격"] = None
            cell[f"{label}_재고"] = None
    if prices_avail:
        lo = min(prices_avail)
        cell["최저가"] = lo[0]
        cell["최저소싱처"] = lo[1]
    else:
        cell["최저가"] = None
        cell["최저소싱처"] = "전 소싱처 품절"
    rows.append((ckey(color), skey(size), cell))

rows.sort(key=lambda x: (x[0], x[1]))
data = [r[2] for r in rows]
print(f"옵션 {len(data)}행 구성")

# ── 엑셀 ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb = Workbook(); ws = wb.active; ws.title = "메이트_소싱처비교"
A = "Arial"
hf = Font(name=A, bold=True, color="FFFFFF"); hfill = PatternFill("solid", start_color="2E5A88")
cf = Font(name=A); ctr = Alignment(horizontal="center")
thin = Side(style="thin", color="DDDDDD"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
lowfill = PatternFill("solid", start_color="E8F8EE")

ws["A1"] = "르무통 메이트 — 소싱처별 가격/재고 비교 (크롤 최신 스냅샷)"
ws["A1"].font = Font(name=A, bold=True, size=13)
ws.merge_cells("A1:L1")
ws["A2"] = "재고 999 = 있음(수량미상), 0 = 품절. 최저가 = 재고 있는 소싱처 중 최저."
ws["A2"].font = Font(name=A, size=10, color="888888"); ws.merge_cells("A2:L2")

headers = ["색상","사이즈",
           "르무통_가격","르무통_재고","무신사_가격","무신사_재고",
           "SSF_가격","SSF_재고","롯데온_가격","롯데온_재고","최저가","최저소싱처"]
hr = 4
for ci, h in enumerate(headers, 1):
    c = ws.cell(hr, ci, h); c.font = hf; c.fill = hfill; c.alignment = ctr; c.border = bd

price_cols = {"르무통_가격","무신사_가격","SSF_가격","롯데온_가격","최저가"}
for ri, d in enumerate(data, hr + 1):
    for ci, h in enumerate(headers, 1):
        v = d.get(h)
        c = ws.cell(ri, ci, v); c.font = cf; c.border = bd
        if h not in ("색상","최저소싱처"):
            c.alignment = ctr
        if h in price_cols and isinstance(v, (int, float)):
            c.number_format = "#,##0"
    # 최저 소싱처 가격 셀 초록 강조
    low_src = d.get("최저소싱처")
    low_map = {"르무통":"르무통_가격","무신사":"무신사_가격","SSF":"SSF_가격","롯데온":"롯데온_가격"}
    if low_src in low_map:
        idx = headers.index(low_map[low_src]) + 1
        ws.cell(ri, idx).fill = lowfill
        ws.cell(ri, idx).font = Font(name=A, bold=True, color="0F7A3D")

widths = [12,8,11,9,11,9,9,9,11,9,11,14]
from openpyxl.utils import get_column_letter
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

out = Path(os.path.expanduser("~")) / "Desktop" / "르무통_메이트_소싱처별_가격재고.xlsx"
wb.save(out)
print(str(out))
