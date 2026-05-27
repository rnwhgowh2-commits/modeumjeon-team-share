"""재고관리 DB → 박스히어로 양식 엑셀 v4 — 색상 정리 강화."""
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

BX_PATH = Path(r'C:\Users\seung\Downloads\Items_Export_99LAB_2026-05-27T20-43-39.xlsx')


def norm(s):
    if not s:
        return ''
    s = s.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    s = s.replace(':', '').replace('/', '').lower()
    return s


def clean_color(raw_color, model_name_display, model_name_raw, brand, bx_product_name):
    """다단계 색상 정리.

    1. raw 가 model_name_display 또는 raw 로 시작 → strip
    2. 정규화 매칭 ("에어 포스 1" ↔ "에어포스1")
    3. 박스히어로 제품명 폴백 — 브랜드·모델명 빼서 색상 추출
    """
    if not raw_color:
        return 'ONE Color'
    raw = raw_color.strip()

    for mn in [model_name_display, model_name_raw]:
        if not mn:
            continue
        if raw.startswith(mn):
            r = raw[len(mn):].strip()
            if r:
                return r
        # 정규화 매칭
        rn, mn_n = norm(raw), norm(mn)
        if rn.startswith(mn_n) and mn_n:
            consumed = 0
            cut_idx = len(raw)
            for i, ch in enumerate(raw):
                if consumed >= len(mn_n):
                    cut_idx = i
                    break
                if ch.lower() not in ' -():/':
                    consumed += 1
            r = raw[cut_idx:].strip().lstrip('-:/ ').strip()
            if r:
                return r

    # 박스히어로 제품명 폴백
    if bx_product_name:
        s = bx_product_name.strip()
        if brand and brand in s:
            s = s.replace(brand, '', 1).strip()
        for mn in [model_name_display, model_name_raw]:
            if mn and mn in s:
                s = s.replace(mn, '', 1).strip()
                break
        s = s.strip(' -:/').strip()
        if s and len(s) < len(raw):
            return s
    return raw


def main():
    # 박스히어로 엑셀
    bx_wb = openpyxl.load_workbook(BX_PATH, data_only=True)
    bx_ws = bx_wb['BoxHero']
    bx_headers = [c.value for c in bx_ws[1]]
    bx_rows = [dict(zip(bx_headers, row)) for row in bx_ws.iter_rows(min_row=2, values_only=True)]
    bx_by_sku = {r['SKU']: r for r in bx_rows if r.get('SKU')}

    from sqlalchemy.orm import joinedload
    from sqlalchemy import text
    from shared.db import SessionLocal
    from lemouton.sourcing.models import Option

    s = SessionLocal()
    items = (s.query(Option).options(joinedload(Option.model))
             .order_by(Option.model_code, Option.color_code, Option.size_code).all())

    usage = {}
    for r in s.execute(text(
            'SELECT product_canonical_sku, COUNT(*) FROM option_product_links GROUP BY product_canonical_sku'
    )).fetchall():
        usage[r[0]] = r[1]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '재고관리_DB'
    headers = ['SKU', '바코드', '브랜드', '제품명', '품번', '색상', '사이즈',
               '평균매입가', '총재고', '기본 위치', '그로스', '모음전 적용', '활성']
    hd_fill = PatternFill('solid', fgColor='4F81BD')
    hd_font = Font(color='FFFFFF', bold=True, size=11)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hd_fill
        c.font = hd_font
        c.alignment = center
        c.border = border
    widths = [16, 16, 8, 26, 14, 16, 8, 11, 9, 11, 9, 11, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    off_fill = PatternFill('solid', fgColor='F0F0F0')
    no_sku_fill = PatternFill('solid', fgColor='FFF2CC')

    for r_idx, o in enumerate(items, start=2):
        bs = o.boxhero_sku or ''
        bc = o.barcode or ''
        bx = bx_by_sku.get(bs) if bs else None
        qty_main = (bx and bx.get('수량(기본 위치)')) or 0
        qty_gros = (bx and bx.get('수량(그로스)')) or 0
        qty_total = o.boxhero_stock_total if o.boxhero_stock_total is not None else ((bx and bx.get('수량')) or 0)
        brand = (o.model.brand if o.model else '') or ''
        mnd = (o.model.model_name_display if o.model else '') or ''
        mnr = (o.model.model_name_raw if o.model else '') or ''
        bxname = bx['제품명'] if bx else ''
        raw_color = o.color_display or o.color_code or ''
        cleaned = clean_color(raw_color, mnd, mnr, brand, bxname)

        # 제품명 — model_name_display 우선, 브랜드·색상 strip
        pn = mnd or mnr or ''
        if brand and pn.startswith(brand):
            pn = pn[len(brand):].strip()
        if cleaned and cleaned != 'ONE Color' and pn.endswith(cleaned):
            pn = pn[:-len(cleaned)].strip()

        article_no = ''
        if o.model and o.model.article_no and o.model.article_no != '-':
            article_no = o.model.article_no
        elif '_' in o.model_code:
            article_no = o.model_code.split('_', 1)[1]
        else:
            article_no = o.model_code

        row_data = [
            bs, bc, brand, pn, article_no, cleaned,
            (o.size_display or o.size_code or 'FREE'),
            o.boxhero_avg_purchase_price or 0,
            qty_total, qty_main, qty_gros,
            usage.get(o.canonical_sku, 0),
            'O' if o.is_active else 'X',
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if c_idx in (1, 2, 3, 5, 6, 7, 12, 13):
                cell.alignment = center
        if not o.is_active:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = off_fill
        elif not bs:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = no_sku_fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # 요약 시트
    ws2 = wb.create_sheet('요약')
    ws2['A1'] = '항목'
    ws2['B1'] = '값'
    for c in ws2[1]:
        c.fill = hd_fill
        c.font = hd_font
        c.alignment = center
    summary = [
        ('생성일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('전체 옵션', len(items)),
        ('SKU 보유', sum(1 for o in items if o.boxhero_sku)),
        ('is_active=True', sum(1 for o in items if o.is_active)),
        ('is_active=False (OFF)', sum(1 for o in items if not o.is_active)),
    ]
    for i, (k, v) in enumerate(summary, start=2):
        ws2.cell(row=i, column=1, value=k).border = border
        ws2.cell(row=i, column=2, value=v).border = border
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 22

    out_path = Path(r'C:\Users\seung\Downloads') / f'재고관리_DB_Export_v4_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(out_path)
    print(f'저장: {out_path}')

    # 검증
    print()
    print('=== 검증 ===')
    test_models = ['나이키_FZ0627-010', '리츠플리_롱스커트', '리츠플리_미디원피스', '리츠플리_미디가디건',
                   '리츠플리_미스트_베이직_반목', '잔스포츠_슈퍼브레이크', '마스마룰즈_데일리백팩', '빔즈_지갑']
    for tm in test_models:
        samples = [o for o in items if o.model_code == tm][:3]
        if samples:
            print(f'\n[{tm}]')
            for o in samples:
                brand = (o.model.brand if o.model else '') or ''
                mnd = (o.model.model_name_display if o.model else '') or ''
                mnr = (o.model.model_name_raw if o.model else '') or ''
                bx = bx_by_sku.get(o.boxhero_sku or '', {})
                cleaned = clean_color(o.color_code or '', mnd, mnr, brand, bx.get('제품명') or '')
                print(f'  raw="{(o.color_code or "")[:35]:<35}" → cleaned="{cleaned}"')
    s.close()


if __name__ == '__main__':
    main()
