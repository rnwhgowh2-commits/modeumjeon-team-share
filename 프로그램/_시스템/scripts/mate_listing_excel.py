# -*- coding: utf-8 -*-
"""메이트 — listing(모음전/단품)별 소싱처 가격/재고 엑셀 (+URL 열, 클릭 검증용).

SourceProduct/SourceOption + OptionSourceLink 에서 읽음.
bundle_source_urls.label 로 listing 이름 + 단품 색상 판별(오염 링크 정리).
"""
from __future__ import annotations
import sys
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
from pathlib import Path
import os
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
import config  # noqa
from shared.db import SessionLocal
from lemouton.sourcing.models import Model, Option, BundleSourceUrl
from lemouton.sources.models import SourceProduct, SourceOption, OptionSourceLink

MATE = "르무통_메이트"
COLOR_ORDER = ['그레이','다크네이비','라이트블루','블랙','스카이블루','아이보리','오렌지','올리브그린','크림핑크']
SRC_LABEL = {'lemouton':'르무통','musinsa':'무신사','ssf':'SSF','lotteon':'롯데온','ssg':'SSG','ss_lemouton':'스스르무통'}

def _ns(t): return "".join((t or "").split()).lower()
def _label_color(label):
    if not label or "_" not in label: return None
    tail = label.split("_",1)[1]
    for c in COLOR_ORDER:
        if _ns(c)==_ns(tail): return c
    return None

s = SessionLocal()
opts = {o.canonical_sku:(o.color_code,o.size_code) for o in s.query(Option).filter_by(model_code=MATE).all()}
skus = set(opts)
url_label = {r.url: r.label for r in s.query(BundleSourceUrl).filter_by(model_code=MATE).all()}
sp_map = {sp.id:(sp.site, url_label.get(sp.url) or (sp.product_name or sp.url), sp.url) for sp in s.query(SourceProduct).all()}
so_map = {so.id:(so.source_product_id, so.current_price, so.current_stock) for so in s.query(SourceOption).all()}

rows = []  # (color,size,site,listing,price,stock,url)
for ln in s.query(OptionSourceLink).all():
    if ln.canonical_sku not in skus: continue
    so = so_map.get(ln.source_option_id)
    if not so: continue
    sp_id, price, stock = so
    sp = sp_map.get(sp_id)
    if not sp: continue
    site, listing, url = sp
    color, size = opts[ln.canonical_sku]
    intended = _label_color(listing)
    if intended and color != intended:
        continue  # 단품 listing 인데 다른 색 → 오염 링크 제거
    rows.append((color, size, site, listing, price, stock, url))
s.close()

# dedup (color,size,site,listing) → 1
seen=set(); uniq=[]
for r in rows:
    k=(r[0],r[1],r[2],r[3])
    if k in seen: continue
    seen.add(k); uniq.append(r)
rows=uniq

def ck(c): return COLOR_ORDER.index(c) if c in COLOR_ORDER else 99
def sk(z):
    d="".join(ch for ch in (z or "") if ch.isdigit()); return int(d) if d else 0
rows.sort(key=lambda r:(ck(r[0]), sk(r[1]), r[2], r[3]))
print("상세 행:", len(rows))

from collections import defaultdict
best=defaultdict(lambda:(None,None,None))
for color,size,site,listing,price,stock,url in rows:
    if price and stock!=0:
        cur=best[(color,size)]
        if cur[0] is None or price<cur[0]:
            best[(color,size)]=(price, SRC_LABEL.get(site,site), listing)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb=Workbook()
A="Arial"
hf=Font(name=A,bold=True,color="FFFFFF"); hfill=PatternFill("solid",start_color="2E5A88")
cf=Font(name=A); ctr=Alignment(horizontal="center")
thin=Side(style="thin",color="DDDDDD"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
lowfill=PatternFill("solid",start_color="E8F8EE")
linkfont=Font(name=A,color="1155CC",underline="single")

# Sheet1 최저가요약
ws=wb.active; ws.title="최저가요약"
ws["A1"]="르무통 메이트 — 옵션별 최저 소싱가 (listing 통합)"; ws["A1"].font=Font(name=A,bold=True,size=13); ws.merge_cells("A1:E1")
for ci,h in enumerate(["색상","사이즈","최저가","최저 소싱처","최저 listing"],1):
    c=ws.cell(3,ci,h); c.font=hf; c.fill=hfill; c.alignment=ctr; c.border=bd
keys=sorted(best.keys(), key=lambda k:(ck(k[0]),sk(k[1])))
r=4
for color,size in keys:
    p,site,listing=best.get((color,size),(None,None,None))
    ws.cell(r,1,color).font=cf; ws.cell(r,2,size).font=cf; ws.cell(r,2).alignment=ctr
    pc=ws.cell(r,3,p); pc.font=Font(name=A,bold=True,color="0F7A3D"); pc.alignment=ctr; pc.number_format="#,##0"
    ws.cell(r,4, site or "").font=cf; ws.cell(r,5, listing or "").font=cf
    for ci in range(1,6): ws.cell(r,ci).border=bd
    r+=1
for col,w in zip("ABCDE",[12,8,12,12,22]): ws.column_dimensions[col].width=w
ws.freeze_panes="A4"

# Sheet2 listing별상세 (+URL)
ws2=wb.create_sheet("listing별상세")
ws2["A1"]="소싱처·listing 별 가격/재고 + URL (클릭해서 실제값 대조)"; ws2["A1"].font=Font(name=A,bold=True,size=13); ws2.merge_cells("A1:G1")
for ci,h in enumerate(["색상","사이즈","소싱처","listing","가격","재고","URL (클릭)"],1):
    c=ws2.cell(3,ci,h); c.font=hf; c.fill=hfill; c.alignment=ctr; c.border=bd
r=4
for color,size,site,listing,price,stock,url in rows:
    ws2.cell(r,1,color).font=cf
    ws2.cell(r,2,size).font=cf; ws2.cell(r,2).alignment=ctr
    ws2.cell(r,3,SRC_LABEL.get(site,site)).font=cf
    ws2.cell(r,4,listing).font=cf
    pc=ws2.cell(r,5,price); pc.font=cf; pc.alignment=ctr; pc.number_format="#,##0"
    sc=ws2.cell(r,6,stock); sc.font=cf; sc.alignment=ctr
    uc=ws2.cell(r,7,url)
    if url:
        uc.hyperlink=url; uc.font=linkfont
    else:
        uc.font=cf
    b=best.get((color,size))
    if b and price==b[0] and price and stock!=0:
        pc.fill=lowfill; pc.font=Font(name=A,bold=True,color="0F7A3D")
    for ci in range(1,8): ws2.cell(r,ci).border=bd
    r+=1
for col,w in zip("ABCDEFG",[11,7,9,20,10,8,70]): ws2.column_dimensions[col].width=w
ws2.freeze_panes="A4"

out=Path(os.path.expanduser("~"))/"Desktop"/"르무통_메이트_소싱처_정답비교.xlsx"
wb.save(out)
print("옵션수:", len(keys))
print(str(out))
