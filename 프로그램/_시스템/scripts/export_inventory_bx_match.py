"""재고관리 엑셀 (박스히어로 기준, 우리 양식).

룰:
- 행 = 박스히어로 엑셀 그대로 (810건)
- 컬럼 = 우리 프로그램 양식 12개 (SKU·바코드·품번·브랜드·카테고리·모델명·색상·사이즈·평균매입가·총재고·기본 위치·그로스)
- 메타 (품번·브랜드·카테고리·모델명·색상·사이즈·평균매입가) = 우리 DB 우선, 없으면 박스히어로
- 재고 (총재고·기본 위치·그로스) = 박스히어로 그대로
"""
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def main():
    # 박스히어로 (정답 행 기준)
    bx_path = Path(r'C:\Users\seung\Downloads\Items_Export_99LAB_2026-05-27T21-41-37.xlsx')
    bx_wb = openpyxl.load_workbook(bx_path, data_only=True)
    bx_ws = bx_wb['BoxHero']
    bx_hdr = [c.value for c in bx_ws[1]]
    bx_rows = [dict(zip(bx_hdr, r)) for r in bx_ws.iter_rows(min_row=2, values_only=True)]
    print(f'박스히어로 행: {len(bx_rows)}')

    # 우리 DB 의 옵션 (boxhero_sku 기준 매핑)
    from sqlalchemy.orm import joinedload
    from shared.db import SessionLocal
    from lemouton.sourcing.models import Option

    s = SessionLocal()
    try:
        opts = (s.query(Option).options(joinedload(Option.model))
                .filter(Option.boxhero_sku.isnot(None))
                .filter(Option.boxhero_sku != '')
                .all())
        # boxhero_sku → Option (활성 우선, 옛 sku 우선)
        db_by_bs = {}
        for o in opts:
            existing = db_by_bs.get(o.boxhero_sku)
            if existing is None:
                db_by_bs[o.boxhero_sku] = o
                continue
            # 우선순위: is_active=True > 옛 sku 형식 > SKU-XXX
            score_o = (1 if o.is_active else 0, 1 if not o.canonical_sku.startswith('SKU-') else 0)
            score_e = (1 if existing.is_active else 0, 1 if not existing.canonical_sku.startswith('SKU-') else 0)
            if score_o > score_e:
                db_by_bs[o.boxhero_sku] = o
        print(f'DB boxhero_sku 매핑: {len(db_by_bs)}')

        # Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '재고관리'

        headers = ['SKU', '바코드', '품번', '브랜드', '카테고리', '모델명',
                   '색상', '사이즈', '평균매입가', '총재고', '기본 위치', '그로스']
        ws.append(headers)

        matched = 0
        unmatched = []
        for bx in bx_rows:
            sku = bx.get('SKU') or ''
            barcode = str(bx.get('바코드') or '')
            o = db_by_bs.get(sku)
            if o:
                matched += 1
                m = o.model
                article = (getattr(m, 'article_no', None) or '-') if m else '-'
                brand = (m.brand or '') if m else (bx.get('브랜드') or '')
                category = (m.category or '') if m else (bx.get('카테고리') or '')
                mname = ((getattr(m, 'model_name_display', None) or
                          getattr(m, 'model_name_raw', None)) if m else '') or ''
                color = o.color_display or o.color_code or '-'
                size = o.size_display or o.size_code or (str(bx.get('사이즈')) if bx.get('사이즈') else '-')
                avg = int(o.boxhero_avg_purchase_price or 0)
            else:
                unmatched.append(sku)
                article = '-'
                brand = bx.get('브랜드') or ''
                category = bx.get('카테고리') or ''
                mname = bx.get('제품명') or ''
                color = '-'
                size = str(bx.get('사이즈')) if bx.get('사이즈') else '-'
                avg = int(bx.get('구매가') or 0)

            total = int(bx.get('수량') or 0)
            qty_main = int(bx.get('수량(기본 위치)') or 0)
            qty_gros = int(bx.get('수량(그로스)') or 0)

            ws.append([
                sku, barcode, article, brand, category, mname,
                color, size, avg, total, qty_main, qty_gros,
            ])

        # 헤더 스타일
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4F67FF', end_color='4F67FF', fill_type='solid')
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        widths = [18, 16, 16, 12, 18, 28, 18, 10, 12, 8, 11, 9]
        for i, w in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

        ts = datetime.now().strftime('%Y%m%d-%H%M%S')
        out_path = Path(r'C:\Users\seung\Downloads') / f'재고관리_박스히어로기준_{ts}.xlsx'
        wb.save(out_path)
        print(f'저장: {out_path}')
        print(f'행: {len(bx_rows)}')
        print(f'DB 매칭됨: {matched}')
        print(f'DB 매칭 안됨 (박스히어로 자체 정보 사용): {len(unmatched)}')
        for u in unmatched[:5]:
            print(f'  {u}')
    finally:
        s.close()


if __name__ == '__main__':
    main()
