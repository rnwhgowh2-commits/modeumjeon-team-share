# -*- coding: utf-8 -*-
"""르무통 메이트 '올리브그린' 사이즈별 크롤 → 엑셀(.xlsx) 저장. (읽기 전용 크롤)

실행: python -m scripts.crawl_olive_to_excel
"""
from __future__ import annotations
import os
os.environ["WATCH_CRAWL"] = "1"  # 보이는 브라우저

import sys
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
from pathlib import Path
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import config  # noqa: F401
from shared.db import SessionLocal
from lemouton.sourcing.models import Model

TARGET_COLOR_HINT = "올리브"  # '올리브그린' 부분일치


def find_lemouton_url():
    s = SessionLocal()
    try:
        for m in s.query(Model).all():
            u = getattr(m, "url_lemouton", None)
            if u and ("product_no=130" in u or "메이트" in (m.model_code or "")):
                return m.model_code, u
        # fallback: 첫 lemouton URL
        for m in s.query(Model).all():
            u = getattr(m, "url_lemouton", None)
            if u:
                return m.model_code, u
        return None, None
    finally:
        s.close()


def main() -> int:
    code, url = find_lemouton_url()
    print(f"모음전: {code}")
    print(f"르무통 URL: {url}")
    if not url:
        print("르무통 URL 없음")
        return 1

    from lemouton.sourcing.crawlers.lemouton import LemoutonCrawler
    crawler = LemoutonCrawler(prefer_playwright=True)
    print("크롤 시작 (브라우저 창이 뜹니다)...")
    result = crawler.fetch(url)
    opts = getattr(result, "options", []) or []
    pname = getattr(result, "product_name_raw", "?")
    print(f"상품명: {pname} / 전체 옵션 {len(opts)}개")

    colors = sorted({(o.get("color_text") or "").strip() for o in opts})
    print(f"색상 목록: {colors}")

    olive = [o for o in opts
             if TARGET_COLOR_HINT in (o.get("color_text") or "")]
    if not olive:
        print(f"⚠️ '{TARGET_COLOR_HINT}' 색상 매칭 0건 — 위 색상 목록 중 정확한 이름 알려주세요.")
        return 2

    # 사이즈 숫자 기준 정렬
    def _size_key(o):
        sz = "".join(ch for ch in (o.get("size_text") or "") if ch.isdigit())
        return int(sz) if sz else 0
    olive.sort(key=_size_key)

    print(f"\n올리브그린 {len(olive)}건:")
    for o in olive:
        print(f"  {o.get('color_text')}/{o.get('size_text')}  "
              f"가격={o.get('price')}  재고={o.get('stock')}")

    # ── 엑셀 저장 ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "올리브그린"
    arial = "Arial"
    hdr_font = Font(name=arial, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="2E5A88")
    cell_font = Font(name=arial)
    center = Alignment(horizontal="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = f"{pname} — 올리브그린 (크롤: 르무통 공홈)"
    ws["A1"] = title
    ws["A1"].font = Font(name=arial, bold=True, size=13)
    ws.merge_cells("A1:D1")

    headers = ["색상", "사이즈", "가격(원)", "재고"]
    ws.append([])  # row2 빈줄
    ws.append(headers)  # row3
    hdr_row = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=col)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border

    first_data = hdr_row + 1
    for o in olive:
        ws.append([
            o.get("color_text"),
            o.get("size_text"),
            o.get("price"),
            o.get("stock"),
        ])
    last_data = first_data + len(olive) - 1

    for r in range(first_data, last_data + 1):
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.font = cell_font; cell.border = border
            if col in (2, 3, 4):
                cell.alignment = center
            if col == 3:
                cell.number_format = "#,##0"

    # 요약 (formula)
    sumr = last_data + 2
    ws.cell(row=sumr, column=1, value="요약").font = Font(name=arial, bold=True)
    ws.cell(row=sumr+1, column=1, value="옵션 수").font = cell_font
    ws.cell(row=sumr+1, column=2, value=f"=COUNTA(A{first_data}:A{last_data})")
    ws.cell(row=sumr+2, column=1, value="최저가").font = cell_font
    ws.cell(row=sumr+2, column=2, value=f"=MIN(C{first_data}:C{last_data})").number_format = "#,##0"
    ws.cell(row=sumr+3, column=1, value="최고가").font = cell_font
    ws.cell(row=sumr+3, column=2, value=f"=MAX(C{first_data}:C{last_data})").number_format = "#,##0"
    ws.cell(row=sumr+4, column=1, value="재고 합계").font = cell_font
    ws.cell(row=sumr+4, column=2, value=f"=SUM(D{first_data}:D{last_data})")

    widths = {"A": 16, "B": 12, "C": 12, "D": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = Path(os.path.expanduser("~")) / "Desktop" / "르무통_메이트_올리브그린.xlsx"
    try:
        wb.save(out)
    except Exception:
        out = _ROOT.parent.parent / "르무통_메이트_올리브그린.xlsx"
        wb.save(out)
    print(f"\n✅ 엑셀 저장: {out}")
    print(str(out))  # 마지막 줄 = 경로
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
