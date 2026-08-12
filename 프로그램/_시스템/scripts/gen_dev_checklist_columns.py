# -*- coding: utf-8 -*-
"""사장님 엑셀 「■ 마켓별 상품등록 정보」 Sheet2 → 판매처 열 정의 JSON.

엑셀은 사장님 바탕화면에 있어 서버가 못 읽는다. **한 번 뽑아 저장소에 넣는다.**
엑셀이 바뀌면 이 스크립트를 다시 돌리고 결과를 커밋한다.

  python scripts/gen_dev_checklist_columns.py "C:/Users/seung/OneDrive/바탕 화면/■ 마켓별 상품등록 정보.xlsx"

🔴 열 순서·묶음은 엑셀 그대로 둔다. 우리가 재배열하면 사장님이 보던 표와 어긋난다.
"""
import json
import pathlib
import sys

import openpyxl

# 엑셀 열(B~Z) → 가공정책 항목 키. None = 아직 담을 칸이 프로그램에 없음.
ITEM_BY_COL = {
    2: "name", 3: "brand", 4: "category", 5: "price", 6: "price",
    7: "images", 8: "images", 9: "options", 10: "detail",
    11: "shipping", 12: "shipping", 13: "price_compare",
    14: "listing", 15: "listing", 16: "origin", 17: "listing", 18: "listing",
    19: "kc", 20: "notice", 21: "tags",
    22: "ids", 23: "ids", 24: None, 25: "shipping", 26: None,
}
# 화면에서만 다르게 부르는 이름. 엑셀 원문은 그대로 두고 표시만 바꾼다.
#   26(Z)열 — 엑셀에선 U~Y 묶음과 이름이 똑같이 「기타」라 머리글에 「기타」가 두 번 찍힌다.
#   Z열은 위너 가격·로켓그로스·스마일캐시처럼 **그 마켓에만 있는 것**이라 성격이 다르다.
#   사장님 확정 2026-08-13 — 「마켓 고유」로 부른다.
NAME_OVERRIDE = {26: "마켓 고유"}

ROWS = [(3, "lotteon"), (4, "coupang"), (5, "smartstore"),
        (6, "eleven11"), (7, "auction"), (8, "gmarket")]


def _split_rule(text):
    """「이름 ▶ 사장님 기준」이 한 칸에 붙어 있다. 둘로 나눈다."""
    t = (text or "").replace("\r", "")
    i = t.find("▶")
    return (t.strip(), "") if i < 0 else (t[:i].strip(), t[i + 1:].strip())


def build(xlsx_path):
    ws = openpyxl.load_workbook(xlsx_path, data_only=True)["Sheet2"]
    merged = {}
    for rng in ws.merged_cells.ranges:
        top_left = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = top_left

    def cell(r, c):
        v = merged[(r, c)] if (r, c) in merged else ws.cell(r, c).value
        return "" if v is None else str(v).strip()

    columns = []
    for col in range(2, 27):
        group, _ = _split_rule(cell(1, col))
        name, rule = _split_rule(cell(2, col) or cell(1, col))
        columns.append({
            "col": col,
            "group": NAME_OVERRIDE.get(col, group),
            "name": NAME_OVERRIDE.get(col, name or group),
            "rule": rule or _split_rule(cell(1, col))[1],
            "item": ITEM_BY_COL[col],
            "specs": {mk: cell(row, col) for row, mk in ROWS},
        })
    return {"source": "■ 마켓별 상품등록 정보.xlsx / Sheet2", "columns": columns}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit('사용법: python scripts/gen_dev_checklist_columns.py "<엑셀 경로>"')
    out = pathlib.Path(__file__).parents[1] / "webapp" / "data" / "dev_checklist_columns.json"
    data = build(sys.argv[1])
    out.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"열 {len(data['columns'])}개 → {out}")
