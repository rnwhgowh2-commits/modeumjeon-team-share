"""재고관리 DB → 엑셀 v5 — 박스히어로 제품명 기반 자동 분리 (제품명 + 색상).

사용자 룰:
- 품번: 영숫자+하이픈+언더스코어만 (한글·SKU 형식 ❌)
- 제품명: 박스히어로 제품명에서 brand·색상 빼고 (예: "슈퍼브레이크 플러스")
- 색상: 박스히어로 제품명의 마지막 단어 (예: "그레이")
"""
import re
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

BX_PATH = Path(r'C:\Users\seung\Downloads\Items_Export_99LAB_2026-05-27T20-43-39.xlsx')

# 영숫자 + 하이픈 + 언더스코어 — 한글 ❌
ARTICLE_NO_RE = re.compile(r'^[A-Za-z0-9_\-]+$')


def is_valid_article_no(s):
    return bool(s) and bool(ARTICLE_NO_RE.match(s)) and not s.startswith('SKU-')


def parse_bx_product(bx_name, brand):
    """박스히어로 제품명 → (제품명, 색상).

    룰: 마지막 1 단어 = 색상. 나머지 = 제품명. brand 는 제거.
    """
    if not bx_name:
        return ('', '')
    s = bx_name.strip()
    # "(W)" 보존 — 가장 앞 토큰
    prefix = ''
    if s.startswith('(W)'):
        prefix = '(W) '
        s = s[3:].strip()
    # brand 제거 (첫 등장만)
    if brand and brand in s:
        s = s.replace(brand, '', 1).strip()
    # 트레일링 하이픈·콜론 제거
    s = s.rstrip(' -:/').strip()
    # 토큰 분리
    tokens = s.split()
    if len(tokens) == 0:
        return (prefix.strip(), '')
    if len(tokens) == 1:
        # 단일 토큰 → 색상 (제품명 빈칸)
        return (prefix.strip(), tokens[0])
    # 마지막 1 단어 = 색상, 나머지 = 제품명
    color = tokens[-1]
    product = (prefix + ' '.join(tokens[:-1])).strip()
    return (product, color)


def main():
    bx_wb = openpyxl.load_workbook(BX_PATH, data_only=True)
    bx_ws = bx_wb['BoxHero']
    bx_headers = [c.value for c in bx_ws[1]]
    bx_rows = [dict(zip(bx_headers, row)) for row in bx_ws.iter_rows(min_row=2, values_only=True)]
    bx_by_sku = {r['SKU']: r for r in bx_rows if r.get('SKU')}

    from sqlalchemy.orm import joinedload
    from sqlalchemy import text
    from shared.db import SessionLocal
    from lemouton.sourcing.models import Option

    s = SessionLocal()
    items = (s.query(Option).options(joinedload(Option.model))
             .order_by(Option.model_code, Option.color_code, Option.size_code).all())

    usage = {}
    for r in s.execute(text(
            'SELECT product_canonical_sku, COUNT(*) FROM option_product_links GROUP BY product_canonical_sku'
    )).fetchall():
        usage[r[0]] = r[1]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '재고관리_DB'
    headers = ['SKU', '바코드', '브랜드', '제품명', '품번', '색상', '사이즈',
               '평균매입가', '총재고', '기본 위치', '그로스', '모음전 적용', '활성']
    hd_fill = PatternFill('solid', fgColor='4F81BD')
    hd_font = Font(color='FFFFFF', bold=True, size=11)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hd_fill
        c.font = hd_font
        c.alignment = center
        c.border = border
    widths = [16, 16, 8, 24, 14, 14, 8, 11, 9, 11, 9, 11, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    off_fill = PatternFill('solid', fgColor='F0F0F0')
    no_sku_fill = PatternFill('solid', fgColor='FFF2CC')
    dirty_fill = PatternFill('solid', fgColor='FFC7CE')  # 박스히어로 자체 더러운 데이터

    for r_idx, o in enumerate(items, start=2):
        bs = o.boxhero_sku or ''
        bc = o.barcode or ''
        bx = bx_by_sku.get(bs) if bs else None
        qty_main = (bx and bx.get('수량(기본 위치)')) or 0
        qty_gros = (bx and bx.get('수량(그로스)')) or 0
        qty_total = o.boxhero_stock_total if o.boxhero_stock_total is not None else ((bx and bx.get('수량')) or 0)
        brand = (o.model.brand if o.model else '') or ''

        # 박스히어로 제품명에서 자동 분리 (제품명·색상)
        if bx and bx.get('제품명'):
            product_name, color = parse_bx_product(bx['제품명'], brand)
        else:
            # 박스히어로 없으면 DB fallback
            product_name = (o.model.model_name_display if o.model else '') or ''
            color = (o.color_display or o.color_code or '').strip()
            # 색상에서 제품명 시작 부분 제거
            if product_name and color.startswith(product_name):
                color = color[len(product_name):].strip()

        # 품번 — 영숫자+하이픈+_ 만, 한글/SKU 형식 ❌
        article_no = ''
        if o.model and o.model.article_no and is_valid_article_no(o.model.article_no):
            article_no = o.model.article_no
        elif '_' in o.model_code:
            candidate = o.model_code.split('_', 1)[1]
            if is_valid_article_no(candidate):
                article_no = candidate

        # 박스히어로 자체 더러운 데이터 감지 — 모델명이 박스히어로 제품명과 큰 차이
        is_dirty = False
        if bx and bx.get('모델명') and bx.get('제품명') and brand:
            bx_model = str(bx.get('모델명'))
            bx_pname = str(bx.get('제품명'))
            # 같은 모델인데 표기가 너무 다른 케이스 — model_name_raw 또는 display 와 매칭 안 되면
            mnd = (o.model.model_name_display if o.model else '')
            mnr = (o.model.model_name_raw if o.model else '')
            # 박스히어로 제품명에 brand 가 없거나 model 매칭 안 되는 경우
            if brand not in bx_pname:
                is_dirty = True

        row_data = [
            bs, bc, brand, product_name, article_no, color,
            (o.size_display or o.size_code or 'FREE'),
            o.boxhero_avg_purchase_price or 0,
            qty_total, qty_main, qty_gros,
            usage.get(o.canonical_sku, 0),
            'O' if o.is_active else 'X',
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if c_idx in (1, 2, 3, 5, 6, 7, 12, 13):
                cell.alignment = center
        if not o.is_active:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = off_fill
        elif is_dirty:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = dirty_fill
        elif not bs:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = no_sku_fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # 요약
    ws2 = wb.create_sheet('요약')
    ws2['A1'] = '항목'
    ws2['B1'] = '값'
    for c in ws2[1]:
        c.fill = hd_fill
        c.font = hd_font
        c.alignment = center
    summary = [
        ('생성일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('전체 옵션', len(items)),
        ('SKU 보유', sum(1 for o in items if o.boxhero_sku)),
        ('is_active=True', sum(1 for o in items if o.is_active)),
        ('is_active=False (OFF)', sum(1 for o in items if not o.is_active)),
    ]
    for i, (k, v) in enumerate(summary, start=2):
        ws2.cell(row=i, column=1, value=k).border = border
        ws2.cell(row=i, column=2, value=v).border = border
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 22

    # 범례
    ws3 = wb.create_sheet('범례')
    ws3['A1'] = '색상'
    ws3['B1'] = '의미'
    for c in ws3[1]:
        c.fill = hd_fill
        c.font = hd_font
        c.alignment = center
    legend = [
        ('흰색', '활성 + SKU 보유 (정상)'),
        ('노랑', '활성이지만 박스히어로 미등록 (사용자 추가)'),
        ('빨강', '박스히어로 자체 데이터 이상 (브랜드·모델명 미입력)'),
        ('회색', '비활성 — 사용자 OFF'),
    ]
    fills = [PatternFill(), no_sku_fill, dirty_fill, off_fill]
    for i, ((name, desc), f) in enumerate(zip(legend, fills), start=2):
        c1 = ws3.cell(row=i, column=1, value=name)
        c1.fill = f
        c1.border = border
        ws3.cell(row=i, column=2, value=desc).border = border
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 60

    out_path = Path(r'C:\Users\seung\Downloads') / f'재고관리_DB_Export_v5_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(out_path)
    print(f'저장: {out_path}')

    # 검증
    print()
    print('=== 검증: 사용자가 지적한 케이스 ===')
    test = [
        ('잔스포츠_슈퍼브레이크', '슈퍼브레이크 / 슈퍼브레이크 플러스 분리'),
        ('나이키_FV5420-002', '(W) 코르테즈 텍스타일 / 블랙'),
        ('나이키_FZ0627-010', '박스히어로 자체 더러움'),
        ('마스마룰즈_데일리백팩', '데일리 백팩 / 색상'),
        ('빔즈_빔즈_음식_키링', '음식 키링 / 음식명'),
        ('빔즈_지갑', '지갑 / 색상'),
        ('리츠플리_롱스커트', '롱스커트 / 색상'),
        ('리츠플리_미디원피스', '미디원피스 / 색상'),
    ]
    for tm, desc in test:
        samples = [o for o in items if o.model_code == tm][:3]
        if not samples:
            print(f'\n[{tm}] — DB 없음')
            continue
        print(f'\n[{tm}] {desc}')
        for o in samples:
            brand = (o.model.brand if o.model else '') or ''
            bx = bx_by_sku.get(o.boxhero_sku or '', {})
            pn, col = parse_bx_product(bx.get('제품명') or '', brand) if bx else ('', '')
            print(f'  bx="{(bx.get("제품명") or "")[:35]:<35}" → 제품명="{pn:<15}" 색상="{col}"')
    s.close()


if __name__ == '__main__':
    main()
