"""재고관리 DB → 엑셀 (사용자 export 양식 그대로).

컬럼: SKU, 바코드, 품번, 브랜드, 카테고리, 모델명, 색상, 사이즈,
      평균매입가, 총재고, 기본 위치, 그로스
"""
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def main():
    # 박스히어로 엑셀 — 기본 위치/그로스 수량 가져오기
    bx_path = Path(r'C:\Users\seung\Downloads\Items_Export_99LAB_2026-05-27T20-43-39.xlsx')
    bx_wb = openpyxl.load_workbook(bx_path, data_only=True)
    bx_ws = bx_wb['BoxHero']
    bx_hdr = [c.value for c in bx_ws[1]]
    bx_rows = [dict(zip(bx_hdr, r)) for r in bx_ws.iter_rows(min_row=2, values_only=True)]
    bx_by_sku = {r['SKU']: r for r in bx_rows if r.get('SKU')}

    from sqlalchemy.orm import joinedload
    from shared.db import SessionLocal
    from lemouton.sourcing.models import Option

    s = SessionLocal()
    # boxhero_sku 보유 옵션만 (사용자 화면 export 와 동일)
    items = (s.query(Option).options(joinedload(Option.model))
             .filter(Option.boxhero_sku.isnot(None))
             .filter(Option.boxhero_sku != '')
             .order_by(Option.model_code, Option.color_code, Option.size_code).all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '재고관리'

    headers = ['SKU', '바코드', '품번', '브랜드', '카테고리', '모델명', '색상',
               '사이즈', '평균매입가', '총재고', '기본 위치', '그로스']

    hd_fill = PatternFill('solid', fgColor='4F81BD')
    hd_font = Font(color='FFFFFF', bold=True, size=11)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hd_fill
        c.font = hd_font
        c.alignment = center
        c.border = border

    widths = [16, 16, 16, 10, 12, 24, 28, 8, 11, 9, 11, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    off_fill = PatternFill('solid', fgColor='F0F0F0')

    for r_idx, o in enumerate(items, start=2):
        bs = o.boxhero_sku or ''
        bc = o.barcode or ''
        bx = bx_by_sku.get(bs) if bs else None
        qty_main = (bx and bx.get('수량(기본 위치)')) or 0
        qty_gros = (bx and bx.get('수량(그로스)')) or 0
        qty_total = o.boxhero_stock_total if o.boxhero_stock_total is not None else ((bx and bx.get('수량')) or 0)

        # 품번 — model.article_no (사용자 화면과 동일: 비어있으면 '-')
        article_no = '-'
        if o.model and o.model.article_no:
            an = o.model.article_no
            if an and not an.startswith('SKU-') and an != '-':
                article_no = an

        brand = (o.model.brand if o.model else '') or ''
        category = (o.model.category if o.model else '') or ''
        model_name = (o.model.model_name_display if o.model else '') or ''

        # 색상은 DB raw 그대로 (color_display 우선, 없으면 color_code)
        color = (o.color_display or o.color_code or '').strip()
        size = (o.size_display or o.size_code or '').strip() or 'FREE'

        row_data = [
            bs, bc, article_no, brand, category, model_name, color, size,
            o.boxhero_avg_purchase_price or 0,
            qty_total, qty_main, qty_gros,
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if c_idx in (1, 2, 3, 4, 5, 8):
                cell.alignment = center
        if not o.is_active:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = off_fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    out_path = Path(r'C:\Users\seung\Downloads') / f'재고관리_DB_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(out_path)
    print(f'저장: {out_path}')
    print(f'행: {len(items)}')

    s.close()


if __name__ == '__main__':
    main()
