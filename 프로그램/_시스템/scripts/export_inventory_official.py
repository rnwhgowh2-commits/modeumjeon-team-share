"""재고관리 DB → 엑셀 (data.py 의 /data/items/export.xlsx 라우트 1:1 복제).

사용자 화면 export 와 100% 동일한 양식·데이터.
"""
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def main():
    from sqlalchemy.orm import joinedload
    from shared.db import SessionLocal
    from lemouton.sourcing.models import Option, Model
    from lemouton.inventory.models import InventoryLocation
    from shared.inventory_stock import get_stock_batch

    s = SessionLocal()
    try:
        # Query — 정답 라우트와 동일
        q = s.query(Option).options(joinedload(Option.model))
        # 정렬: 브랜드 > 카테고리 > 모델명 > 색상 > 사이즈 (사용자 룰)
        options = (
            q.join(Model, Option.model_code == Model.model_code)
             .order_by(Model.brand, Model.category, Model.model_name_display,
                       Option.color_display, Option.size_display).all()
        )
        all_skus = [o.canonical_sku for o in options]
        print(f'옵션 수: {len(options)}')

        # 재고 (inventory_txs 집계)
        total_stock_map = get_stock_batch(s, all_skus)

        # 인벤토리 위치들
        locs = (
            s.query(InventoryLocation)
            .filter(InventoryLocation.deleted_at.is_(None))
            .order_by(InventoryLocation.sort_order, InventoryLocation.id)
            .all()
        )
        per_loc_stock = {}
        for loc in locs:
            loc_map = get_stock_batch(s, all_skus, location_id=loc.id)
            for sku, st in loc_map.items():
                per_loc_stock.setdefault(sku, {})[loc.id] = st

        # Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '재고관리'

        headers = ['SKU', '바코드', '품번', '브랜드', '카테고리', '모델명',
                   '색상', '사이즈', '평균매입가', '총재고']
        for loc in locs:
            headers.append(loc.name)
        ws.append(headers)

        for o in options:
            m = o.model
            barcode = o.barcode or ''
            brand = (m.brand or '') if m else ''
            article = (getattr(m, 'article_no', None) or '-') if m else '-'
            category = (getattr(m, 'category', None) or '') if m else ''
            mname = ((getattr(m, 'model_name_display', None) or
                      getattr(m, 'model_name_raw', None)) if m else '') or ''
            color = o.color_display or o.color_code or '-'
            size = o.size_display or o.size_code or '-'
            avg = int(o.boxhero_avg_purchase_price or 0)
            total = int(total_stock_map.get(o.canonical_sku, 0))
            # SKU 컬럼: SKU-XXX 형식만 허용 — 옛 sku (한글) 면 boxhero_sku, 없으면 '-'
            sku_display = o.canonical_sku
            if sku_display and not sku_display.startswith('SKU-'):
                sku_display = o.boxhero_sku or '-'
            row = [sku_display, barcode, article, brand, category, mname,
                   color, size, avg, total]
            for loc in locs:
                row.append(int(per_loc_stock.get(o.canonical_sku, {}).get(loc.id, 0)))
            ws.append(row)

        # 헤더 스타일
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4F67FF', end_color='4F67FF', fill_type='solid')
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'

        widths = [18, 16, 16, 12, 18, 28, 18, 10, 12, 8] + [12] * len(locs)
        for i, w in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

        ts = datetime.now().strftime('%Y%m%d-%H%M%S')
        out_path = Path(r'C:\Users\seung\Downloads') / f'재고관리_{ts}.xlsx'
        wb.save(out_path)
        print(f'저장: {out_path}')
        print(f'위치 컬럼: {[loc.name for loc in locs]}')
    finally:
        s.close()


if __name__ == '__main__':
    main()
