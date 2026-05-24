"""엑셀 import 파서.

지원 양식 2종 (헤더로 자동 감지):

A) 박스히어로 19컬럼 (SKU, 바코드, 제품명, 구매가, 판매가, 카테고리,
   브랜드, 모델명, 사이즈, 메모, 안전재고x4, 생성일, 수량x4)
   색상 별도 컬럼 X — 제품명에서 추출.

B) 우리 양식 9 base + 동적 위치별 컬럼
   (SKU, 바코드, 브랜드, 제품명, 품번, 색상, 사이즈, 평균매입가, 총재고, {위치명1} 재고, ...)
   색상 컬럼 직접 사용 ('one' → 색상 없음 처리). 품번 = 박스히어로 model_name 컬럼.
"""
from typing import Iterator
import openpyxl


COL_INDEX = {
    "sku": 0,
    "barcode": 1,
    "name": 2,
    "purchase_price": 3,
    "sale_price": 4,
    "category": 5,
    "brand": 6,
    "model_name": 7,
    "size": 8,
    "memo": 9,
    "stock_safety_total": 10,
    "stock_safety_default": 11,
    "stock_safety_disabled": 12,
    "stock_safety_overall": 13,
    "created_at": 14,
    "quantity": 15,
    "quantity_gross": 16,
    "quantity_default": 17,
    "quantity_disabled": 18,
}


def _extract_color(name: str, brand: str | None, model_name: str | None) -> str:
    """제품명에서 브랜드+모델명을 떼고 나머지를 색상 텍스트로."""
    text = (name or "").strip()
    if brand and text.startswith(brand):
        text = text[len(brand):].strip()
    if model_name and text.startswith(model_name):
        text = text[len(model_name):].strip()
    return text


def parse_boxhero_xlsx(xlsx_path: str) -> Iterator[dict]:
    """박스히어로 엑셀 → 정규화된 dict 레코드 yield."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return

    for row in rows[1:]:
        if not row or row[COL_INDEX["sku"]] is None:
            continue

        sku = str(row[COL_INDEX["sku"]]).strip()
        name = str(row[COL_INDEX["name"]] or "").strip()
        brand = str(row[COL_INDEX["brand"]] or "").strip() or None
        model_name = str(row[COL_INDEX["model_name"]] or "").strip() or None
        size = row[COL_INDEX["size"]]
        size_str = str(size).strip() if size is not None else ""
        quantity = row[COL_INDEX["quantity"]] or 0
        purchase_price = row[COL_INDEX["purchase_price"]] or 0
        color_text = _extract_color(name, brand, model_name)

        # [2026-05-25 D-4] F 카테고리 추가 — models.category 매핑용
        category = str(row[COL_INDEX["category"]] or "").strip() or None
        yield {
            "sku": sku,
            "barcode": str(row[COL_INDEX["barcode"]] or "").strip(),
            "name": name,
            "brand": brand,
            "model_name": model_name,
            "size": size_str,
            "color_text": color_text,
            "quantity": int(quantity),
            "purchase_price": int(purchase_price),
            "category": category,
        }


# ─── 우리 양식 (10 base + 동적 위치별) ───
# [2026-05-25 D-6 v2] 사용자 양식: SKU/바코드/품번/브랜드/카테고리/모델명/색상/사이즈/평균매입가/총재고 + N 위치
INTERNAL_BASE_HEADERS = ['SKU', '바코드', '품번', '브랜드', '카테고리', '모델명',
                        '색상', '사이즈', '평균매입가', '총재고']
INTERNAL_BASE_COL_COUNT = len(INTERNAL_BASE_HEADERS)  # 10


def detect_format(xlsx_path: str) -> str:
    """첫 헤더 행으로 양식 자동 감지. return 'internal' 또는 'boxhero'."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c or '').strip() for c in row]
            break
    finally:
        wb.close()

    # [D-6 v2] 신 우리 양식: SKU/바코드/품번/브랜드/카테고리/모델명/색상/사이즈/평균매입가/총재고
    # 첫 6컬럼만 매칭으로 신·구 v2 판별 (평균매입가 위치는 read 시 헤더 인덱스로)
    if (len(headers) >= 6 and
        headers[:6] == ['SKU', '바코드', '품번', '브랜드', '카테고리', '모델명']):
        return 'internal'
    # 옛 우리 양식 (v1·v2) 호환 — SKU/바코드/브랜드/제품명
    if len(headers) >= 4 and headers[:4] == ['SKU', '바코드', '브랜드', '제품명']:
        return 'internal'
    if len(headers) >= 7 and headers[0] == 'SKU' and headers[2] == '제품명' and headers[6] == '브랜드':
        return 'boxhero'
    # 폴백 — 박스히어로 (기존 동작 유지)
    return 'boxhero'


def parse_internal_xlsx(xlsx_path: str) -> Iterator[dict]:
    """우리 양식 xlsx → 정규화된 dict yield (박스히어로 format 호환).

    9 base + N 위치별 (v2). 5번째 컬럼이 '품번' 이면 v2, 그 외면 v1 호환 모드 (4컬럼 == ['SKU',...,'제품명'] 까지 매칭하고
    5번째가 '색상' 이면 v1 = 8 base).

    추가 필드: 'article_no' (품번), 'per_loc_stock' = {위치명: 재고}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return
    headers = [str(c or '').strip() for c in rows[0]]
    # 양식 판별
    # 신 양식 (D-6): SKU/바코드/품번/브랜드/카테고리/모델명/색상/사이즈/총재고 + N 위치
    is_new_fmt = (len(headers) >= 6 and
                  headers[:6] == ['SKU', '바코드', '품번', '브랜드', '카테고리', '모델명'])
    # 옛 v2 (품번 5번째)
    has_article_old = (not is_new_fmt and len(headers) >= 5 and headers[4] == '품번')
    # [D-6 v2] 신 양식 base = 10 (평균매입가 추가), 단 옛 v3 호환 (평균매입가 없는 9 base) 도 인식
    # 헤더 9번째 칸이 '평균매입가' 면 10 base, 아니면 9 base (v3 = 평균매입가 없는 양식)
    if is_new_fmt:
        if len(headers) >= 10 and headers[8] == '평균매입가':
            base_count = 10
        else:
            base_count = 9
    elif has_article_old:
        base_count = 9
    else:
        base_count = 8
    # 위치별 재고 컬럼 — base 이후
    # 신 양식: 위치명 그대로 (예: '기본 위치', '그로스')
    # 옛 양식: '{위치명} 재고' 패턴
    loc_cols: dict[str, int] = {}
    for i, h in enumerate(headers[base_count:], start=base_count):
        if is_new_fmt:
            if h: loc_cols[h] = i
        else:
            if h.endswith(' 재고') and h != '총재고':
                loc_cols[h[:-len(' 재고')]] = i

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        sku = str(row[0]).strip()
        if not sku:
            continue
        barcode = str(row[1] or '').strip()
        category = None
        if is_new_fmt:
            # A SKU / B 바코드 / C 품번 / D 브랜드 / E 카테고리 / F 모델명 / G 색상 / H 사이즈
            # base=10: I 평균매입가 / J 총재고
            # base=9 (v3 호환): I 총재고
            article_no = str(row[2] or '').strip() or None
            brand = (str(row[3] or '').strip() or None)
            category = str(row[4] or '').strip() or None
            pname = str(row[5] or '').strip()
            color = str(row[6] or '').strip()
            size = str(row[7] or '').strip()
            if base_count == 10:
                try: avg = int(row[8] or 0)
                except (ValueError, TypeError): avg = 0
                total_idx = 9
            else:
                avg = 0
                total_idx = 8
        elif has_article_old:
            brand = (str(row[2] or '').strip() or None)
            pname = str(row[3] or '').strip()
            article_no = str(row[4] or '').strip() or None
            color = str(row[5] or '').strip()
            size = str(row[6] or '').strip()
            avg_idx, total_idx = 7, 8
            try: avg = int(row[avg_idx] or 0)
            except (ValueError, TypeError): avg = 0
        else:
            brand = (str(row[2] or '').strip() or None)
            pname = str(row[3] or '').strip()
            article_no = None
            color = str(row[4] or '').strip()
            size = str(row[5] or '').strip()
            avg_idx, total_idx = 6, 7
            try: avg = int(row[avg_idx] or 0)
            except (ValueError, TypeError): avg = 0
        try:
            total = int(row[total_idx] or 0)
        except (ValueError, TypeError):
            total = 0
        per_loc: dict[str, int] = {}
        for loc, idx in loc_cols.items():
            try:
                per_loc[loc] = int(row[idx] or 0)
            except (ValueError, TypeError):
                per_loc[loc] = 0

        # color '-' / 'one' / 빈값 → 의미적으로 색상 없음
        color_text = '' if color in ('', '-', 'one') else color
        # size '-' / 'free' / 'FREE' → 빈값
        size_str = '' if size in ('', '-', 'free', 'FREE') else size

        yield {
            "sku": sku,
            "barcode": barcode,
            "name": pname,
            "brand": brand,
            "model_name": article_no,  # 박스히어로 호환 — model_name 자리에 품번 매핑
            "article_no": article_no,  # 명시적 필드
            "size": size_str,
            "color_text": color_text,
            "quantity": total,
            "purchase_price": avg,
            "category": category,  # [D-6] 신 양식 카테고리
            "per_loc_stock": per_loc,  # 우리 양식 한정 추가 필드
        }


def parse_xlsx_auto(xlsx_path: str) -> Iterator[dict]:
    """양식 자동 감지 + 적절한 parser 호출. 양쪽 모두 같은 record 포맷."""
    fmt = detect_format(xlsx_path)
    if fmt == 'internal':
        yield from parse_internal_xlsx(xlsx_path)
    else:
        yield from parse_boxhero_xlsx(xlsx_path)
