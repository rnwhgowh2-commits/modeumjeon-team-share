# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

DED_LO = {"블랙":"LO2158462485","다크네이비":"LO2158462341","그레이":"LO2158462407",
    "아이보리":"LO2673780784","오렌지":"LO2158462309","스카이블루":"LO2653544137",
    "크림핑크":"LO2446708389","라이트블루":"LO2158462443","올리브그린":None}
HUB_LO = "LO2158462914"
def url(lo): return ("https://www.lotteon.com/p/product/%s" % lo) if lo else ""

# 가짜상품(spdNo 다름) 제거 적용된 보정 데이터 — 라이브 재수집(2026-06-19)
# (color, size, ded_price, ded_stock, hub_price, hub_stock)
ROWS = [
("블랙","220",116910,10,116910,5),("블랙","225",116910,10,116910,10),("블랙","230",116910,50,116910,10),
("블랙","235",116910,50,116910,10),("블랙","240",116910,50,116910,9),("블랙","245",116910,50,116910,9),
("블랙","250",116910,50,116910,10),("블랙","255",116910,9,116910,10),("블랙","260",116910,49,116910,10),
("블랙","265",116910,20,116910,9),("블랙","270",116910,10,116910,10),("블랙","275",116910,10,116910,10),
("블랙","280",116910,10,116910,10),
("다크네이비","220",116910,20,116910,10),("다크네이비","225",116910,20,116910,20),("다크네이비","230",116910,20,116910,10),
("다크네이비","235",116910,20,116910,50),("다크네이비","240",116910,20,116910,30),("다크네이비","245",116910,20,116910,29),
("다크네이비","250",116910,20,116910,10),("다크네이비","255",116910,5,116910,10),("다크네이비","260",116910,"품절",116910,"품절"),
("다크네이비","265",116910,"품절",116910,"품절"),("다크네이비","270",116910,10,116910,10),("다크네이비","275",116910,"품절",116910,"품절"),
("다크네이비","280",116910,10,116910,10),
("그레이","220",116910,15,116910,10),("그레이","225",116910,20,116910,10),("그레이","230",116910,20,116910,10),
("그레이","235",116910,20,116910,9),("그레이","240",116910,20,116910,10),("그레이","245",116910,20,116910,10),
("그레이","250",116910,20,116910,10),("그레이","255",116910,"품절",116910,"품절"),("그레이","260",116910,"품절",116910,"품절"),
("그레이","265",116910,"품절",116910,"품절"),("그레이","270",116910,20,116910,10),("그레이","275",116910,2,116910,2),
("그레이","280",116910,20,116910,10),
("아이보리","220",129900,4,116910,5),("아이보리","225",129900,17,116910,10),("아이보리","230",129900,41,116910,40),
("아이보리","235",129900,30,116910,50),("아이보리","240",129900,30,116910,49),("아이보리","245",129900,30,116910,50),
("아이보리","250",129900,30,116910,50),("아이보리","255",129900,10,116910,10),("아이보리","260",129900,"품절",116910,"품절"),
("아이보리","265",129900,"품절",116910,"품절"),("아이보리","270",129900,10,116910,10),("아이보리","275",129900,"품절",116910,"품절"),
("아이보리","280",129900,5,116910,"품절"),
("오렌지","220",116910,5,116910,10),("오렌지","225",116910,"품절",116910,"품절"),("오렌지","230",116910,"품절",116910,"품절"),
("오렌지","235",116910,"품절",116910,"품절"),("오렌지","240",116910,10,116910,"품절"),("오렌지","245",116910,10,116910,"품절"),
("오렌지","250",116910,2,116910,10),("오렌지","255",116910,"품절",116910,"품절"),
("스카이블루","220",129900,"품절",116910,"품절"),("스카이블루","225",129900,5,116910,5),("스카이블루","230",129900,20,116910,20),
("스카이블루","235",129900,30,116910,30),("스카이블루","240",129900,30,116910,30),("스카이블루","245",129900,30,116910,30),
("스카이블루","250",129900,20,116910,20),("스카이블루","255",129900,10,116910,10),
("크림핑크","220",116910,"품절",116910,10),("크림핑크","225",116910,"품절",116910,10),("크림핑크","230",116910,"품절",116910,10),
("크림핑크","235",116910,"품절",116910,10),("크림핑크","240",116910,"품절",116910,"품절"),("크림핑크","245",116910,20,116910,10),
("크림핑크","250",116910,20,116910,10),("크림핑크","255",116910,"품절",116910,10),
("라이트블루","220",116910,22,116910,10),("라이트블루","225",116910,28,116910,10),("라이트블루","230",116910,38,116910,10),
("라이트블루","235",116910,38,116910,10),("라이트블루","240",116910,39,116910,9),("라이트블루","245",116910,40,116910,10),
("라이트블루","250",116910,40,116910,10),("라이트블루","255",116910,32,116910,10),("라이트블루","260",116910,"품절",None,None),
("라이트블루","270",116910,"품절",None,None),("라이트블루","280",116910,"품절",None,None),
("올리브그린","220",None,None,116910,"품절"),("올리브그린","225",None,None,116910,"품절"),("올리브그린","230",None,None,116910,"품절"),
("올리브그린","235",None,None,116910,"품절"),("올리브그린","240",None,None,116910,"품절"),("올리브그린","245",None,None,116910,"품절"),
("올리브그린","250",None,None,116910,"품절"),("올리브그린","255",None,None,116910,"품절"),("올리브그린","260",None,None,116910,"품절"),
("올리브그린","265",None,None,116910,"품절"),("올리브그린","270",None,None,116910,10),("올리브그린","275",None,None,116910,"품절"),
("올리브그린","280",None,None,116910,10),
]

# 가짜상품 제거로 '재고있음(999)' → '품절' 바뀐 칸 (변경 강조 대상)
FIXED_DED = {("그레이","255"),("오렌지","230"),("오렌지","235")}

wb = Workbook(); ws = wb.active; ws.title = "롯데온 재고·가격(보정)"
F = "맑은 고딕"
thin = Side(style="thin", color="D0D0D0"); border = Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
ded_fill = PatternFill("solid", fgColor="EAF1FB"); hub_fill = PatternFill("solid", fgColor="FCEFE6")
out_fill = PatternFill("solid", fgColor="FDE2E2"); out_font = Font(name=F, color="C00000", bold=True)
chg_fill = PatternFill("solid", fgColor="FFD24D"); chg_font = Font(name=F, color="9C5700", bold=True)
link_font = Font(name=F, color="0563C1", underline="single", size=9)

ws["A1"] = "롯데온 색상별·사이즈별 재고·가격 (보정본 — 가짜상품 제거 적용, 2026-06-19)"
ws["A1"].font = Font(name=F, bold=True, size=14); ws.merge_cells("A1:H1")
ws["A2"] = "★노란칸 = 가짜상품(다른 spdNo) 제거로 '재고있음(999)'→'품절' 바로잡은 칸. 나머지는 라이브 실수집값. URL 클릭→실제값 대조."
ws["A2"].font = Font(name=F, size=9, italic=True, color="9C5700"); ws.merge_cells("A2:H2")

headers = ["색상","사이즈","단품 URL","단품 가격","단품 재고","모음전 URL","모음전 가격","모음전 재고"]
hr = 4
for c,h in enumerate(headers,1):
    cell = ws.cell(row=hr,column=c,value=h)
    cell.font = Font(name=F,bold=True,color="FFFFFF"); cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center",vertical="center"); cell.border = border

r = hr+1
for (color,size,dp,ds,hp,hs) in ROWS:
    is_fixed = (color,size) in FIXED_DED
    ded_lo = DED_LO.get(color)
    ws.cell(row=r,column=1,value=color).font = Font(name=F,bold=True)
    ws.cell(row=r,column=2,value=size)
    c3 = ws.cell(row=r,column=3)
    if ded_lo: c3.value="열기"; c3.hyperlink=url(ded_lo); c3.font=link_font
    else: c3.value="(전용 없음)"; c3.font=Font(name=F,size=9,color="999999")
    ws.cell(row=r,column=4,value=dp if dp is not None else "")
    c5 = ws.cell(row=r,column=5,value=ds if ds is not None else "")
    c6 = ws.cell(row=r,column=6,value="열기"); c6.hyperlink=url(HUB_LO); c6.font=link_font
    if hp is None and hs is None: c6.value="(사이즈 없음)"; c6.hyperlink=None; c6.font=Font(name=F,size=9,color="999999")
    ws.cell(row=r,column=7,value=hp if hp is not None else "")
    c8 = ws.cell(row=r,column=8,value=hs if hs is not None else "")
    for c in range(1,9):
        cell = ws.cell(row=r,column=c)
        if cell.font is None or cell.font.name!=F: cell.font = Font(name=F)
        cell.border = border; cell.alignment = Alignment(horizontal="center",vertical="center")
        if c in (3,4,5): cell.fill = ded_fill
        if c in (6,7,8): cell.fill = hub_fill
    if ds=="품절": c5.font=out_font; c5.fill=out_fill
    if hs=="품절": c8.font=out_font; c8.fill=out_fill
    # 변경 강조 (가짜상품 제거)
    if is_fixed:
        for c in (4,5):
            ws.cell(row=r,column=c).fill = chg_fill; ws.cell(row=r,column=c).font = chg_font
        ws.cell(row=r,column=5).comment = Comment("이전: 재고있음(999)/가격161650 (다른 상품 spdNo). 실제는 품절 → 보정", "검증")
    for pc in (4,7):
        v = ws.cell(row=r,column=pc).value
        if isinstance(v,int): ws.cell(row=r,column=pc).number_format = "#,##0"
    r += 1

for col,w in {"A":11,"B":8,"C":11,"D":12,"E":15,"F":13,"G":13,"H":15}.items(): ws.column_dimensions[col].width=w
ws.freeze_panes = "A5"; ws.row_dimensions[1].height = 22

out = r"C:\dev\모음전 프로젝트\롯데온_색상사이즈_재고가격_검증_20260619_보정.xlsx"
wb.save(out); print("SAVED:", out, "rows:", len(ROWS), "fixed:", len(FIXED_DED))
