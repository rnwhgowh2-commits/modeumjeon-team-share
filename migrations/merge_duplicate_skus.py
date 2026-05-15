"""32개 중복 SKU 머지 (B 정책: qty 합산, 한쪽 0이라 합산 결과=KEEP qty)
Excel 의 (제품명+사이즈) 그룹 → KEEP/DROP 결정 → DB 적용

전략:
1) Excel 30 중복 그룹 → 각 그룹의 SKU 목록
2) DB Option 에서 매칭, 그룹별 정렬 (qty 큰 순)
3) KEEP = 첫번째, DROP = 나머지
4) DROP 옵션의 InventoryProduct(바코드) 삭제
5) DROP 옵션의 InventoryTx 가 있으면 KEEP 으로 이전
6) DROP Option 삭제
"""
import os, sys
from pathlib import Path
from collections import defaultdict
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try: sys.stdout.reconfigure(encoding="utf-8")
    except: pass
NEW = Path("C:/dev/모음전 프로젝트/_시스템")
sys.path.insert(0, str(NEW))
os.chdir(NEW)
from dotenv import load_dotenv
load_dotenv(NEW / ".env", override=True)
from openpyxl import load_workbook
from sqlalchemy import func
import lemouton.sourcing.models, lemouton.sourcing.models_pricing, lemouton.pricing.settings
import lemouton.uploader.models, lemouton.templates.models
from lemouton.sourcing.models import Option
from lemouton.inventory.models import InventoryProduct, InventoryTx
from shared.db import SessionLocal

# === Excel 30 중복 그룹 ===
P = r"C:\Users\seung\Downloads\Items_Export_99LAB_2026-05-15T21-48-04.xlsx"
wb = load_workbook(P, data_only=True)
ws = wb.active
H = [c.value for c in ws[1]]
def i(c): return H.index(c)

groups = defaultdict(list)
for r in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in r): continue
    name = str(r[i("제품명")] or "")
    size = str(r[i("사이즈")] or "")
    sku = r[i("SKU")]
    qty = int(r[i("수량")] or 0)
    groups[(name, size)].append((sku, qty))
dup_groups = {k:v for k,v in groups.items() if len(v) > 1}

# KEEP/DROP boxhero_sku 리스트
keep_bskus = []
drop_bskus = []
for k, lst in dup_groups.items():
    lst_sorted = sorted(lst, key=lambda x: (-x[1], x[0]))
    keep_bskus.append(lst_sorted[0][0])
    for db, dq in lst_sorted[1:]:
        drop_bskus.append(db)
print(f"중복 그룹: {len(dup_groups)}  / KEEP {len(keep_bskus)} / DROP {len(drop_bskus)}")

# === DB 사전 상태 ===
with SessionLocal() as s:
    pre_opt = s.query(func.count(Option.canonical_sku)).scalar()
    pre_ip  = s.query(func.count(InventoryProduct.id)).scalar()
    pre_tx  = s.query(func.count(InventoryTx.id)).scalar()
    pre_stock_sum = s.query(func.coalesce(func.sum(Option.boxhero_stock_total),0)).scalar()
    print(f"\n사전: Option={pre_opt}, IP={pre_ip}, Tx={pre_tx}, boxhero_stock_total 합={pre_stock_sum}")

# === DROP 옵션 식별 ===
with SessionLocal() as s:
    drop_opts = s.query(Option).filter(Option.boxhero_sku.in_(drop_bskus)).all()
    drop_csku = [o.canonical_sku for o in drop_opts]
    print(f"\nDROP 옵션 매칭: {len(drop_opts)} / 예상 {len(drop_bskus)}")
    if len(drop_opts) != len(drop_bskus):
        print("⚠️ 매칭 누락! 중단")
        sys.exit(1)

    # DROP 측 InventoryTx 카운트 (있으면 KEEP 으로 이전 — 현재는 0 예상)
    tx_drop = s.query(InventoryTx).filter(InventoryTx.option_canonical_sku.in_(drop_csku)).all()
    ip_drop = s.query(InventoryProduct).filter(InventoryProduct.canonical_sku.in_(drop_csku)).all()
    print(f"  DROP InventoryTx:        {len(tx_drop)} (KEEP 으로 이전 필요)")
    print(f"  DROP InventoryProduct:   {len(ip_drop)} (삭제 대상)")

    # === 실행 ===
    # 1) Tx 이전 (만약 있다면) — boxhero_sku → KEEP 의 canonical_sku 매핑
    if tx_drop:
        # boxhero_sku → KEEP canonical_sku 매핑 만들기
        # DROP boxhero_sku 의 그룹 → KEEP boxhero_sku → KEEP canonical_sku
        drop_to_keep_bsku = {}
        for k, lst in dup_groups.items():
            lst_sorted = sorted(lst, key=lambda x: (-x[1], x[0]))
            keep_b = lst_sorted[0][0]
            for db, _ in lst_sorted[1:]:
                drop_to_keep_bsku[db] = keep_b
        # DROP canonical_sku → DROP boxhero_sku
        drop_csku_to_bsku = {o.canonical_sku: o.boxhero_sku for o in drop_opts}
        # KEEP boxhero_sku → KEEP canonical_sku
        keep_opts = s.query(Option).filter(Option.boxhero_sku.in_(keep_bskus)).all()
        keep_bsku_to_csku = {o.boxhero_sku: o.canonical_sku for o in keep_opts}
        # Tx 이전
        for tx in tx_drop:
            drop_b = drop_csku_to_bsku[tx.option_canonical_sku]
            keep_b = drop_to_keep_bsku[drop_b]
            keep_csku = keep_bsku_to_csku[keep_b]
            tx.option_canonical_sku = keep_csku
        print(f"  → Tx {len(tx_drop)}건 KEEP 으로 이전")

    # 2) InventoryProduct 삭제
    for ip in ip_drop:
        s.delete(ip)
    # 3) Option 삭제
    for o in drop_opts:
        s.delete(o)
    s.commit()
    print(f"\n✅ 삭제 완료")

# === DB 사후 상태 ===
with SessionLocal() as s:
    post_opt = s.query(func.count(Option.canonical_sku)).scalar()
    post_ip  = s.query(func.count(InventoryProduct.id)).scalar()
    post_tx  = s.query(func.count(InventoryTx.id)).scalar()
    post_stock_sum = s.query(func.coalesce(func.sum(Option.boxhero_stock_total),0)).scalar()
    print(f"\n사후: Option={post_opt}, IP={post_ip}, Tx={post_tx}, boxhero_stock_total 합={post_stock_sum}")
    print(f"     변화: Opt={post_opt-pre_opt:+d}, IP={post_ip-pre_ip:+d}, Tx={post_tx-pre_tx:+d}, Stock={post_stock_sum-pre_stock_sum:+d}")

    # 르무통 검증
    from sqlalchemy import or_
    lemo = s.query(func.count(Option.canonical_sku))\
            .filter(or_(Option.canonical_sku.like("%르무통%"), Option.boxhero_sku.like("%르무통%"))).scalar()
    lemo_stock = s.query(func.coalesce(func.sum(Option.boxhero_stock_total),0))\
            .filter(or_(Option.canonical_sku.like("%르무통%"), Option.boxhero_sku.like("%르무통%"))).scalar()
    print(f"\n=== 검증 (르무통) ===")
    print(f"옵션 수:     {lemo}  (목표 449)  {'✅' if lemo==449 else '❌'}")
    print(f"재고 합:     {lemo_stock}  (목표 448)  {'✅' if lemo_stock==448 else '❌'}")
